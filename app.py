import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from matplotlib.lines import Line2D
import datetime
import pwlf
from PIL import Image
import os
import plotly.graph_objects as go
import tempfile
from fpdf import FPDF
import openpyxl

# --- GLOBAL TIME & SPRINT PARSING HELPERS ---
def parse_time_to_seconds(t):
    if pd.isnull(t): return None
    if isinstance(t, datetime.time): return t.hour * 3600 + t.minute * 60 + t.second
    if isinstance(t, str):
        try:
            parts = t.strip().split(':')
            if len(parts) == 3: return int(parts[0])*3600 + int(parts[1])*60 + float(parts[2])
            if len(parts) == 2: return int(parts[0])*60 + float(parts[1])
        except: pass
    if isinstance(t, (int, float)):
        if 0 < t < 1: return t * 86400 
        return float(t)
    return 0.0

def parse_sprint_excel(file_contents):
    wb = openpyxl.load_workbook(file_contents, data_only=True)
    sheet_name = "Data Sprint" if "Data Sprint" in wb.sheetnames else wb.sheetnames[0]
    sheet = wb[sheet_name]
    
    sprint_powers = []
    sprint_times = []
    sprint_cadences = []
    
    for r in range(1, sheet.max_row + 1):
        val_a = sheet.cell(row=r, column=1).value
        val_c = sheet.cell(row=r, column=3).value
        val_e = sheet.cell(row=r, column=5).value
        
        if val_a is not None and val_e is not None:
            try:
                p = float(val_a)
                t = parse_time_to_seconds(val_e)
                c = float(val_c) if val_c is not None else 0.0
                if t is not None:
                    sprint_powers.append(p)
                    sprint_times.append(t)
                    sprint_cadences.append(c)
            except ValueError:
                continue
                
    return sprint_powers, sprint_times, sprint_cadences


# TEMPORARY SRM DUMP
srm_file_path = "C:\\Users\\marku\\OneDrive\\Markus_HYCYS\\Sonstiges\\Laufauswertung App\\Radsport Auswertungen\\Ralph Ziegaus - 2026-05-22-10-28-31.srm"
if os.path.exists(srm_file_path):
    with open(srm_file_path, "rb") as f:
        data = f.read()
    with open("C:\\Users\\marku\\OneDrive\\Markus_HYCYS\\Sonstiges\\Laufauswertung App\\srm_hex_dump.txt", "w") as out:
        out.write(f"SRM File Size: {len(data)} bytes\n")
        for i in range(0, len(data), 16):
            chunk = data[i:i+16]
            hex_str = " ".join(f"{b:02x}" for b in chunk)
            ascii_str = "".join(chr(b) if 32 <= b < 127 else "." for b in chunk)
            out.write(f"{i:04x} | {hex_str:<47} | {ascii_str}\n")

# Import PDF generator from dedicated module
from pdf_rad import create_pdf_rad


# 1. APP-LAYOUT & CSS
st.set_page_config(page_title="HYCYS - Diagnostik", layout="wide")

st.markdown("""
    <style>
    button[kind="primary"] {
        background-color: #00a1e0 !important;
        color: white !important;
        border: none !important;
        width: 100% !important;
        font-weight: bold !important;
    }
    [data-testid="stFileUploader"] button {
        background-color: #00a1e0 !important;
        color: white !important;
        border: none !important;
    }
    </style>
    """, unsafe_allow_html=True)

# Initialize session state for metadata
metadata_defaults = [
    ("athlete_name", "Michael Wagner"), 
    ("birthdate", "06.03.1996"), 
    ("coach", "Markus Hertlein"), 
    ("test_date", "22.05.2026"), 
    ("sportart", "Bitte wählen..."), 
    ("kategorie", "Amateur"), 
    ("height", 178), 
    ("weight", 72.3), 
    ("body_fat_pct", 15.2), 
    ("last_uploaded_file", None),
    ("last_test_type", "HYCYS Standart (5min / +0,4m/s)"),
    ("start_v", 2.8),
    ("v_increment", 0.4),
    ("anzahl", 5),
    ("vorlauf", 60),
    ("stufendauer", 5.0),
    ("pausendauer", 30),
    ("ppo_val", 436.0),
    ("hfmax_val", 179.0),
    ("slope_val", 0.11),
    ("intercept_val", 138.0),
    ("carb_intake_factor", 1.0),
    ("vo2_override", 4368.0),
    ("t_bel_manual", 13.2),
    ("t_alak_manual", 3.1),
    ("diagnostik_type", "Running BLUE"),
    ("gender", "männlich"),
    ("lauf_auswertung_gestartet", False)
]
for key, val in metadata_defaults:
    if key not in st.session_state:
        st.session_state[key] = val

def update_sf_rad():
    changes = st.session_state.get("edited_sf_rad_key", {})
    if "edited_rows" in changes:
        for row_idx, edit in changes["edited_rows"].items():
            for col_name, new_val in edit.items():
                st.session_state.df_sf.loc[int(row_idx), col_name] = new_val

def update_sf_lauf():
    changes = st.session_state.get("edited_sf_key", {})
    if "edited_rows" in changes:
        for row_idx, edit in changes["edited_rows"].items():
            for col_name, new_val in edit.items():
                st.session_state.df_sf.loc[int(row_idx), col_name] = new_val

def update_sprint_lac():
    changes = st.session_state.get("sprint_lac_editor_key", {})
    if "edited_rows" in changes:
        for row_idx, edit in changes["edited_rows"].items():
            for col_name, new_val in edit.items():
                st.session_state.df_sprint_lac.loc[int(row_idx), col_name] = new_val

def update_lauftest_lac():
    changes = st.session_state.get("lauftest_lac_editor_key", {})
    if "edited_rows" in changes:
        for row_idx, edit in changes["edited_rows"].items():
            for col_name, new_val in edit.items():
                st.session_state.df_lauftest_input.loc[int(row_idx), col_name] = new_val

def sync_lauftest_input_df():
    anzahl = int(st.session_state.anzahl)
    start_v = float(st.session_state.start_v)
    v_increment = float(st.session_state.v_increment)
    
    stufe_names = ["Ruhe"] + [f"Stufe {i+1}" for i in range(anzahl)]
    speeds = [0.0] + [start_v + (i * v_increment) for i in range(anzahl)]
    speeds_kmh = [0.0] + [s * 3.6 for s in speeds[1:]]
    
    default_ruhe_lac = 1.0
    default_ruhe_hr = 70
    default_lac = [1.03, 1.09, 1.19, 1.42, 2.36, 2.32, 6.05, 9.48]
    default_hr = [123, 137, 145, 154, 164, 170, 177, 183]
    
    target_len = anzahl + 1
    
    if "df_lauftest_input" not in st.session_state:
        lac_values = [default_ruhe_lac] + [default_lac[i] if i < len(default_lac) else 2.0 for i in range(anzahl)]
        hr_values = [default_ruhe_hr] + [default_hr[i] if i < len(default_hr) else 130 for i in range(anzahl)]
        st.session_state.df_lauftest_input = pd.DataFrame({
            "Stufe": stufe_names,
            "v (m/s)": np.round(speeds, 2),
            "v (km/h)": np.round(speeds_kmh, 1),
            "Laktat": lac_values,
            "HF": hr_values
        })
        st.session_state.last_sync_params = (anzahl, start_v, v_increment)
    else:
        df = st.session_state.df_lauftest_input
        current_len = len(df)
        
        expected_speeds = [0.0] + [round(start_v + (i * v_increment), 2) for i in range(anzahl)]
        current_speeds = [round(float(s), 2) if pd.notna(s) else 0.0 for s in df["v (m/s)"].values]
        
        if current_len != target_len or current_speeds != expected_speeds:
            if "lauftest_lac_editor_key" in st.session_state:
                del st.session_state["lauftest_lac_editor_key"]
            if current_len != target_len:
                if target_len > current_len:
                    new_rows = []
                    for i in range(current_len, target_len):
                        idx = i - 1
                        lac_val = default_lac[idx] if idx < len(default_lac) else 2.0
                        hr_val = default_hr[idx] if idx < len(default_hr) else 130
                        new_rows.append({
                            "Stufe": stufe_names[i],
                            "v (m/s)": round(speeds[i], 2),
                            "v (km/h)": round(speeds_kmh[i], 1),
                            "Laktat": lac_val,
                            "HF": hr_val
                        })
                    df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
                else:
                    df = df.iloc[:target_len].copy()
            
            df["Stufe"] = stufe_names
            df["v (m/s)"] = np.round(speeds, 2)
            df["v (km/h)"] = np.round(speeds_kmh, 1)
            st.session_state.df_lauftest_input = df
            st.session_state.last_sync_params = (anzahl, start_v, v_increment)

# Primary Sportart Selection (must be the very first sidebar widget)
sportart_options = ["Bitte wählen...", "Laufen", "Radsport"]
sportart_val = st.session_state.sportart if st.session_state.sportart in sportart_options else "Bitte wählen..."
sportart = st.sidebar.selectbox("Sportart", options=sportart_options, index=sportart_options.index(sportart_val))
st.session_state.sportart = sportart

# Dynamically set title based on sportart
if sportart == "Radsport":
    title_text = "HYCYS Radsportauswertung"
elif sportart == "Laufen":
    title_text = "HYCYS Laufauswertung"
else:
    title_text = "HYCYS Diagnostik"

col_title, col_logo = st.columns([4, 1])
with col_title:
    st.title(title_text)
with col_logo:
    logo_candidates = ["image_822d59.png", "logo.png", "image_828f70.png"]
    logo_path = None
    for candidate in logo_candidates:
        if os.path.exists(candidate):
            st.image(candidate, width='stretch')
            logo_path = candidate
            break

st.markdown("---")

# 2. SEITENLEISTE
# Initialize widget keys in session state so they match general session state defaults
widget_keys_to_init = {
    "athlete_name": st.session_state.athlete_name,
    "birthdate": st.session_state.birthdate,
    "test_date": st.session_state.test_date,
    "height": int(st.session_state.height),
    "weight": float(st.session_state.weight),
    "body_fat_pct": float(st.session_state.body_fat_pct),
    "kategorie": st.session_state.kategorie,
    "coach": st.session_state.coach,
    
    "name_lauft_key": st.session_state.athlete_name,
    "birth_lauft_key": st.session_state.birthdate,
    "test_lauft_key": st.session_state.test_date,
    "height_lauft_key": int(st.session_state.height),
    "weight_lauft_key": float(st.session_state.weight),
    "bf_lauft_key": float(st.session_state.body_fat_pct),
    "kategorie_lauft_key": st.session_state.kategorie,
    "coach_lauft_key": st.session_state.coach,
    "gender": st.session_state.gender,
    "gender_lauft_key": st.session_state.gender,
}
for key, val in widget_keys_to_init.items():
    if key not in st.session_state:
        st.session_state[key] = val

def update_base_data(name, height, weight, birthdate, test_date, gender=None):
    if name:
        st.session_state.athlete_name = name
        st.session_state.name_lauft_key = name
        if st.session_state.get("sportart") == "Laufen":
            st.session_state.athlete_name_lauft_widget = name
    if height is not None:
        st.session_state.height = int(height)
        st.session_state.height_lauft_key = int(height)
        if st.session_state.get("sportart") == "Laufen":
            st.session_state.height_lauft_widget = int(height)
    if weight is not None:
        w_rounded = round(float(weight), 1)
        st.session_state.weight = w_rounded
        st.session_state.weight_lauft_key = w_rounded
        if st.session_state.get("sportart") == "Laufen":
            st.session_state.weight_lauft_widget = w_rounded
    if birthdate:
        st.session_state.birthdate = birthdate
        st.session_state.birth_lauft_key = birthdate
        if st.session_state.get("sportart") == "Laufen":
            st.session_state.birthdate_lauft_widget = birthdate
    if test_date:
        st.session_state.test_date = test_date
        st.session_state.test_lauft_key = test_date
        if st.session_state.get("sportart") == "Laufen":
            st.session_state.test_date_lauft_widget = test_date
    if gender:
        st.session_state.gender = gender
        st.session_state.gender_lauft_key = gender
        if st.session_state.get("sportart") == "Laufen":
            st.session_state.gender_lauft_widget = gender

if "df_sf" not in st.session_state:
    sf_names = ["Wange", "Kinn", "Achselfalte vorn", "10. Rippe", "Bauch (Nabel)", "Spina illiaca", "Oberschenkel", "Rücken", "Triceps", "Wade"]
    st.session_state.df_sf = pd.DataFrame({"Falte": sf_names, "M1": [0.0]*10, "M2": [0.0]*10, "M3": [0.0]*10})

# Sportart selection handled at startup
pass

# ------------------ CONDITIONAL SPORTART RENDERING ------------------
if sportart == "Bitte wählen...":
    st.info("Bitte wählen Sie zuerst eine Sportart in der Sidebar aus, um mit der Auswertung zu beginnen.")
    st.stop()

elif sportart == "Radsport":
    # Pre-populate cycling constants and sprint/energetics parameters
    if "ppo_val" not in st.session_state: st.session_state.ppo_val = 436.0
    if "hfmax_val" not in st.session_state: st.session_state.hfmax_val = 179.0
    if "slope_val" not in st.session_state: st.session_state.slope_val = 0.11
    if "intercept_val" not in st.session_state: st.session_state.intercept_val = 138.0
    if "vo2max_spiro" not in st.session_state: st.session_state.vo2max_spiro = 4292.2
    if "t_bel" not in st.session_state: st.session_state.t_bel = 13.2
    if "t_alak" not in st.session_state: st.session_state.t_alak = 3.1
    if "t_glyc" not in st.session_state: st.session_state.t_glyc = 10.1
    if "t_bel_auto" not in st.session_state: st.session_state.t_bel_auto = 13.2
    if "t_alak_auto" not in st.session_state: st.session_state.t_alak_auto = 3.1
    if "t_bel_manual" not in st.session_state: st.session_state.t_bel_manual = 13.2
    if "t_alak_manual" not in st.session_state: st.session_state.t_alak_manual = 3.1
    if "rad_auswertung_gestartet" not in st.session_state: st.session_state.rad_auswertung_gestartet = False
    if "carb_intake_factor" not in st.session_state: st.session_state.carb_intake_factor = 1.0
    
    if "ks1" not in st.session_state: st.session_state.ks1 = 0.0631
    if "ks2" not in st.session_state: st.session_state.ks2 = 1.331
    if "ks4" not in st.session_state: st.session_state.ks4 = 11.7
    if "lac_eq" not in st.session_state: st.session_state.lac_eq = 0.02049
    if "lac_dist" not in st.session_state: st.session_state.lac_dist = 0.4
    if "vo2_rest" not in st.session_state: st.session_state.vo2_rest = 4.0

    coach_options = ["Markus Hertlein", "Marius Trompetter", "Susanne Traser", "Manuel Kuhnle", "Billie Benkel", "Jean Surmont", "Hosea Frick", "Björn Geesmann", "Gregor Eichhorn"]
    kategorie_options = ["Hobby", "Age-Grouper", "Amateur", "Profi"]

    # 0. File Uploaders placed at the very top of the sidebar
    uploaded_sprint = st.sidebar.file_uploader(
        "Sprintdatei hochladen (.xlsx, .srm, .fit)", 
        type=["xlsx", "srm", "fit"], 
        key="sprint_file_key"
    )
    uploaded_spiro = st.sidebar.file_uploader(
        "Spirodatei VO2max Rampe (10s) hochladen (.xlsx, .csv)", 
        type=["xlsx", "csv"], 
        key="spiro_rampe_key"
    )
    uploaded_fit = st.sidebar.file_uploader(
        "VO2max Rampe .fit-Datei hochladen (.fit)", 
        type=["fit"], 
        key="fit_rampe_key"
    )

    st.sidebar.caption(
        "Original SRM-Ergometer Datei (.srm), Garmin/SRM Export (.fit) oder Sprint-Excel-Datei (.xlsx) hochladen. "
        "Power [W] und Trittfrequenz werden automatisch ausgelesen."
    )

    if uploaded_sprint is not None and uploaded_sprint.name != st.session_state.get("last_uploaded_sprint"):
        st.session_state.last_uploaded_sprint = uploaded_sprint.name
        _ext = os.path.splitext(uploaded_sprint.name)[1].lower()
        
        try:
            if _ext == ".xlsx":
                # Excel-Datei parsen
                sprint_powers, sprint_times, sprint_cadences = parse_sprint_excel(uploaded_sprint)
                _src_name = uploaded_sprint.name
                _format_info = "Excel"
                # Rohdaten für Download löschen
                if "sprint_records" in st.session_state:
                    del st.session_state["sprint_records"]
                if "last_uploaded_sprint_raw" in st.session_state:
                    del st.session_state["last_uploaded_sprint_raw"]
            else:
                # SRM/FIT-Datei parsen
                _suffix = _ext
                with tempfile.NamedTemporaryFile(delete=False, suffix=_suffix) as tmp_raw:
                    tmp_raw.write(uploaded_sprint.getvalue())
                    tmp_raw_path = tmp_raw.name
                
                try:
                    import importlib.util as _ilu
                    _spec = _ilu.spec_from_file_location(
                        "convert_sprint", 
                        os.path.join(os.path.dirname(os.path.abspath(__file__)), "convert_sprint_to_excel.py")
                    )
                    _conv = _ilu.module_from_spec(_spec)
                    _spec.loader.exec_module(_conv)
                    
                    _slope  = float(st.session_state.get("srm_slope", 1.0))
                    _offset = float(st.session_state.get("srm_offset", 0.0))
                    
                    if _ext == ".srm":
                        with open(tmp_raw_path, "rb") as _f:
                            _srm_data = _f.read()
                        _records, _hdr = _conv.parse_srm7_data(_srm_data, slope=_slope, zero_offset=_offset, trim_zeros=True)
                        _rec_int_label = f"{_hdr['rec_int']}s (SRM)"
                        _src_name = uploaded_sprint.name
                    else:  # .fit
                        _records, _ath, _date = _conv.parse_fit_file(tmp_raw_path, trim_zeros=True)
                        _rec_int_label = "1s (FIT)"
                        _src_name = uploaded_sprint.name
                    
                    if _records:
                        sprint_powers = [float(r.get("power_w", r.get("power", 0))) for r in _records]
                        sprint_times  = [float(r["elapsed_s"]) for r in _records]
                        sprint_cadences = [float(r.get("cadence", 0)) for r in _records]
                        st.session_state.sprint_records = _records
                        st.session_state.last_uploaded_sprint_raw = uploaded_sprint.name
                        st.session_state.sprint_rec_int_label = _rec_int_label
                        _format_info = "0.1s (SRM)" if _ext == ".srm" else "1s (FIT)"
                    else:
                        raise ValueError("Keine Leistungsdaten gefunden.")
                finally:
                    if os.path.exists(tmp_raw_path):
                        os.unlink(tmp_raw_path)
            
            if len(sprint_powers) > 0:
                p_max = max(sprint_powers)
                t_pmax_list = [t for p, t in zip(sprint_powers, sprint_times) if p == p_max]
                t_pmax_last = max(t_pmax_list) if t_pmax_list else 0.0
                t_alak = None
                p_threshold = p_max * 0.965
                for p, t in zip(sprint_powers, sprint_times):
                    if t > t_pmax_last and p < p_threshold:
                        t_alak = t
                        break
                if t_alak is None:
                    times_after_pmax = [t for t in sprint_times if t > t_pmax_last]
                    t_alak = times_after_pmax[-1] if times_after_pmax else (t_pmax_last + 0.1)
                t_bel = max(sprint_times)
                
                st.session_state.t_bel_auto = round(t_bel, 1)
                st.session_state.t_alak_auto = round(t_alak, 1)
                st.session_state.t_bel_manual = round(t_bel, 1)
                st.session_state.t_alak_manual = round(t_alak, 1)
                st.session_state.t_bel = round(t_bel, 1)
                st.session_state.t_alak = round(t_alak, 1)
                st.session_state.t_glyc = round(t_bel - t_alak, 1)
                st.session_state.sprint_powers = sprint_powers
                st.session_state.sprint_times = sprint_times
                st.session_state.sprint_cadences = sprint_cadences
                st.session_state.sprint_src_name = _src_name
                st.session_state.rad_auswertung_gestartet = False
                
                st.success(f"✅ Geladen [{_format_info}]: {len(sprint_powers)} Punkte, Max {int(p_max)}W, Dauer {t_bel:.1f}s — Nullwerte am Anfang automatisch entfernt.")
                st.rerun()
            else:
                st.warning("Keine Leistungsdaten gefunden.")
        except Exception as e:
            st.error(f"Fehler beim Auslesen der Sprintdatei: {e}")

    # Converted Sprint Excel download button has been moved below the form.

    # Spiro uploader moved to the top of sidebar
    if uploaded_spiro is not None and uploaded_spiro.name != st.session_state.get("last_uploaded_spiro"):
        st.session_state.last_uploaded_spiro = uploaded_spiro.name
        try:
            df_spiro_excel = pd.read_excel(uploaded_spiro, header=None)
            lastname = df_spiro_excel.iloc[1, 1] if df_spiro_excel.shape[0] > 1 and df_spiro_excel.shape[1] > 1 else ""
            firstname = df_spiro_excel.iloc[2, 1] if df_spiro_excel.shape[0] > 2 and df_spiro_excel.shape[1] > 1 else ""
            name_val = f"{firstname} {lastname}".strip()
            
            h_val = None
            if df_spiro_excel.shape[0] > 5 and df_spiro_excel.shape[1] > 1:
                try: h_val = int(float(df_spiro_excel.iloc[5, 1]))
                except: pass
            
            w_val = None
            if df_spiro_excel.shape[0] > 6 and df_spiro_excel.shape[1] > 1:
                try: w_val = round(float(df_spiro_excel.iloc[6, 1]), 1)
                except: pass
                
            b_val = None
            if df_spiro_excel.shape[0] > 7 and df_spiro_excel.shape[1] > 1:
                try:
                    raw_birth = df_spiro_excel.iloc[7, 1]
                    birth_str = str(int(float(raw_birth))).zfill(8)
                    b_val = f"{birth_str[:2]}.{birth_str[2:4]}.{birth_str[4:]}"
                except: pass
                
            t_val = None
            if df_spiro_excel.shape[0] > 0 and df_spiro_excel.shape[1] > 4:
                try:
                    raw_test = df_spiro_excel.iloc[0, 4]
                    test_str = str(int(float(raw_test))).zfill(8)
                    t_val = f"{test_str[:2]}.{test_str[2:4]}.{test_str[4:]}"
                except: pass
                
            update_base_data(name_val, h_val, w_val, b_val, t_val)
                
            times_sec = df_spiro_excel.iloc[3:, 9].apply(parse_time_to_seconds)
            vo2_col = pd.to_numeric(df_spiro_excel.iloc[3:, 13], errors='coerce')
            df_spiro_data = pd.DataFrame({'Time': times_sec, 'VO2': vo2_col}).dropna(subset=['Time', 'VO2'])
            if len(df_spiro_data) >= 2:
                t_diffs = df_spiro_data['Time'].diff().dropna()
                interval = int(round(t_diffs.median()))
            else:
                interval = 10
            window_pts = max(1, int(30 / interval))
            vo2_smooth = df_spiro_data['VO2'].rolling(window=window_pts, center=True).mean()
            st.session_state.vo2max_spiro = round(vo2_smooth.max(), 1)
            st.session_state.rad_auswertung_gestartet = False
            st.rerun()
        except Exception as e:
            st.error(f"Fehler beim Auslesen der Spirodatei: {e}")

    # FIT uploader moved to the top of sidebar
    if uploaded_fit is not None and uploaded_fit.name != st.session_state.get("last_uploaded_fit"):
        st.session_state.last_uploaded_fit = uploaded_fit.name
        with tempfile.NamedTemporaryFile(delete=False, suffix=".fit") as tmp:
            tmp.write(uploaded_fit.getvalue())
            tmp_path = tmp.name
        try:
            import fitparse
            fitfile = fitparse.FitFile(tmp_path)
            records_fit = []
            for record in fitfile.get_messages("record"):
                vals = record.get_values()
                ts = vals.get("timestamp")
                p = vals.get("power")
                hr = vals.get("heart_rate")
                if ts is not None:
                    records_fit.append({
                        "timestamp": ts,
                        "Power": float(p) if p is not None else 0.0,
                        "HR": float(hr) if hr is not None else np.nan
                    })
            
            if records_fit:
                df_raw = pd.DataFrame(records_fit)
                df_raw = df_raw.sort_values("timestamp")
                df_raw = df_raw.drop_duplicates(subset=["timestamp"], keep="last")
                df_raw.set_index("timestamp", inplace=True)
                df_fit = df_raw.resample("1s").ffill()
                df_fit = df_fit.reset_index()
            else:
                df_fit = pd.DataFrame(columns=["timestamp", "Power", "HR"])

            if not df_fit.empty:
                df_fit['Power_30s'] = df_fit['Power'].rolling(window=30, center=False).mean()
                max_p30 = df_fit['Power_30s'].max()
                if pd.isna(max_p30):
                    max_p30 = 0.0
                valid_hrs = df_fit['HR'].dropna().tolist()
                hf_max = max(valid_hrs) if valid_hrs else 180.0
            else:
                max_p30 = 0.0
                hf_max = 180.0
            
            # Only calculate regression once heart rate starts to rise significantly (excluding warmup and lag)
            # In the 3x3min warmup protocol, maximum power is ~150W; constant rise stabilizes around 215W.
            reg_df = df_fit[(df_fit['Power'] >= 215) & (~df_fit['HR'].isna())]
            if len(reg_df) >= 2:
                slope, intercept = np.polyfit(reg_df['Power'], reg_df['HR'], 1)
            else:
                slope, intercept = 0.15, 100.0
                
            st.session_state.ppo_val = round(max_p30, 1)
            st.session_state.hfmax_val = float(hf_max)
            st.session_state.slope_val = round(slope, 4)
            st.session_state.intercept_val = round(intercept, 2)
            st.session_state.rad_auswertung_gestartet = False
            st.rerun()
        except Exception as e:
            st.error(f"Fehler beim Auslesen der FIT-Datei: {e}")
        finally:
            if os.path.exists(tmp_path): os.unlink(tmp_path)

    # Pre-calculate body fat percentage from skinfolds to avoid rerun loops
    if "df_sf" in st.session_state:
        sf_means = st.session_state.df_sf[["M1", "M2", "M3"]].mean(axis=1)
        sum_sf = sf_means.sum()
        if sum_sf > 0:
            calculated_bf = (22.32 * np.log10(sum_sf)) - 29.2
            st.session_state.body_fat_pct = max(0.0, round(calculated_bf, 1))
            
    with st.sidebar.form("cycling_form", enter_to_submit=False):
        # 1. Basisdaten Expander (Default collapsed)
        with st.expander("Basisdaten", expanded=False):
            athlete_name = st.text_input("Name", value=st.session_state.athlete_name)
            st.session_state.athlete_name = athlete_name

            birthdate = st.text_input("Geburtsdatum", value=st.session_state.birthdate)
            st.session_state.birthdate = birthdate

            test_date = st.text_input("Testdatum", value=st.session_state.test_date)
            st.session_state.test_date = test_date

            height = st.number_input("Größe (cm)", value=int(st.session_state.height), step=1, format="%d")
            st.session_state.height = height

            weight = st.number_input("Gewicht (kg)", value=float(st.session_state.weight), step=0.1, format="%.1f")
            st.session_state.weight = round(weight, 1)

            body_fat_pct = st.number_input("Körperfett (%)", value=float(st.session_state.body_fat_pct), step=0.1, format="%.1f")
            st.session_state.body_fat_pct = round(body_fat_pct, 1)

            coach_val = st.session_state.coach
            if coach_val not in coach_options:
                coach_options = coach_options + [coach_val]
            coach = st.selectbox("Coach", options=coach_options, index=coach_options.index(coach_val))
            st.session_state.coach = coach

            kategorie_val = st.session_state.kategorie
            if kategorie_val not in kategorie_options:
                kategorie_options = kategorie_options + [kategorie_val]
            kategorie = st.selectbox("Kategorie", options=kategorie_options, index=kategorie_options.index(kategorie_val))
            st.session_state.kategorie = kategorie

            gender_options = ["männlich", "weiblich"]
            gender_val = st.session_state.get("gender", "männlich")
            gender = st.selectbox("Geschlecht", options=gender_options, index=gender_options.index(gender_val))
            st.session_state.gender = gender

            diag_type_options = ["Cycling WHITE", "Cycling BLUE", "Cycling GOLD"]
            diag_type_val = st.session_state.get("diagnostik_type", "Cycling BLUE")
            if diag_type_val not in diag_type_options:
                diag_type_val = "Cycling BLUE"
            diagnostik_type = st.selectbox("Diagnostik Typ", options=diag_type_options, index=diag_type_options.index(diag_type_val))
            st.session_state.diagnostik_type = diagnostik_type
            
            # Körperfettmessung (Parizkova 10-Falten)
            with st.expander("Körperfettmessung (Parizkova 10-Falten)"):
                if "df_sf" not in st.session_state:
                    sf_names = ["Wange", "Kinn", "Achselfalte vorn", "10. Rippe", "Bauch (Nabel)", "Spina illiaca", "Oberschenkel", "Rücken", "Triceps", "Wade"]
                    st.session_state.df_sf = pd.DataFrame({"Falte": sf_names, "M1": [0.0]*10, "M2": [0.0]*10, "M3": [0.0]*10})
                edited_sf = st.data_editor(
                    st.session_state.df_sf,
                    hide_index=True,
                    width='stretch',
                    key="edited_sf_rad_key"
                )

        # 2. Sprinttest Expander (Default collapsed)
        with st.expander("Sprinttest", expanded=False):
            # Laktatwerte Sprint
            st.subheader("Laktatwerte Sprint")
            if "df_sprint_lac" not in st.session_state:
                times = ["Ruhelaktat 1", "Ruhelaktat 2", "Ruhelaktat 3", "Nachbelastung Min 0", "Nachbelastung Min 2", "Nachbelastung Min 3", "Nachbelastung Min 4", "Nachbelastung Min 5", "Nachbelastung Min 6", "Nachbelastung Min 7", "Nachbelastung Min 8", "Nachbelastung Min 9", "Nachbelastung Min 10"]
                default_lacs = [1.0, 0.94, 0.91, 1.38, 6.87, 7.1, 7.72, 7.75, 8.04, 7.79, 7.58, 7.42, 7.42]
                st.session_state.df_sprint_lac = pd.DataFrame({
                    "Messpunkt": times,
                    "Laktat [mmol/L]": default_lacs
                })
                
            edited_sprint_lac = st.data_editor(
                st.session_state.df_sprint_lac,
                disabled=["Messpunkt"],
                hide_index=True,
                use_container_width=True,
                height=500,
                column_config={
                    "Messpunkt": st.column_config.TextColumn("Messpunkt", width="medium"),
                    "Laktat [mmol/L]": st.column_config.NumberColumn("Laktat", min_value=0.0, max_value=25.0, step=0.01, format="%.2f", width="small")
                },
                key="sprint_lac_editor_key"
            )
        
            st.subheader("Sprinttest Parameter")
            t_bel_auto = float(st.session_state.get("t_bel_auto", 13.2))
            t_alak_auto = float(st.session_state.get("t_alak_auto", 3.1))
            t_glyc_auto = round(t_bel_auto - t_alak_auto, 1)
            
            t_bel_manual = float(st.session_state.get("t_bel_manual", 13.2))
            t_alak_manual = float(st.session_state.get("t_alak_manual", 3.1))
            t_glyc_manual = round(t_bel_manual - t_alak_manual, 1)
            
            st.markdown(f"""
            | Parameter | Auto (Excel) | Manuell |
            | :--- | :--- | :--- |
            | **t_bel [s]** | {t_bel_auto:.1f} | {t_bel_manual:.1f} |
            | **t_alak [s]** | {t_alak_auto:.1f} | {t_alak_manual:.1f} |
            | **t_glyc [s]** | {t_glyc_auto:.1f} | {t_glyc_manual:.1f} |
            """)
            
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                t_bel_manual_input = st.number_input("Manuelle t_bel [s]", value=float(st.session_state.t_bel_manual), step=0.1, format="%.1f")
                st.session_state.t_bel_manual = round(t_bel_manual_input, 1)
            with col_t2:
                t_alak_manual_input = st.number_input("Manuelle t_alak [s]", value=float(st.session_state.t_alak_manual), step=0.1, format="%.1f")
                st.session_state.t_alak_manual = round(t_alak_manual_input, 1)
                    
            sprint_source = st.selectbox(
                "Aktive Quelle für Sprinttest",
                ["Automatisch aus Excel-Rohdaten", "Manuelle Anpassung"],
                key="sprint_source_key"
            )

            st.subheader("SRM Kalibrierung")
            _srm_slope  = st.number_input("Slope [W/count]", 
                                           value=float(st.session_state.get("srm_slope", 1.0)), 
                                           step=0.001, format="%.3f", key="srm_slope_input")
            _srm_offset = st.number_input("Zero-Offset [counts]",
                                           value=float(st.session_state.get("srm_offset", 0.0)),
                                           step=1.0, format="%.1f", key="srm_offset_input")
            st.session_state.srm_slope  = _srm_slope
            st.session_state.srm_offset = _srm_offset

        # 3. Rampentest Expander (Default collapsed)
        with st.expander("Rampentest", expanded=False):
            st.subheader("Rampentest Parameter")
            ppo_val_input = st.number_input("30s PPO [W]", value=float(st.session_state.ppo_val), step=1.0, format="%.1f")
            st.session_state.ppo_val = round(ppo_val_input, 1)
            
            hfmax_val_input = st.number_input("max. Herzfrequenz [bpm]", value=float(st.session_state.hfmax_val), step=1.0, format="%.1f")
            st.session_state.hfmax_val = round(hfmax_val_input, 1)
            
            slope_val_input = st.number_input("Steigung lin. Reg. [bpm/W]", value=float(st.session_state.slope_val), step=0.0001, format="%.4f")
            st.session_state.slope_val = round(slope_val_input, 4)
            
            intercept_val_input = st.number_input("Nullstelle lin. Reg. [bpm]", value=float(st.session_state.intercept_val), step=1.0, format="%.1f")
            st.session_state.intercept_val = round(intercept_val_input, 1)

            st.subheader("VO2max Quellenvergleich")
            weight = float(st.session_state.get("weight", 72.3))
            ppo_val = float(st.session_state.get("ppo_val", 436.0))
            spiro_abs = float(st.session_state.get("vo2max_spiro", 4292.2))
            spiro_rel = spiro_abs / weight if weight > 0 else 0.0
            m_abs = ppo_val * 9.44 + 592.0
            m_rel = m_abs / weight if weight > 0 else 0.0
            w_abs = ppo_val * 9.44 + 275.0
            w_rel = w_abs / weight if weight > 0 else 0.0
            override_abs = float(st.session_state.get("vo2_override", 4368.0))
            override_rel = override_abs / weight if weight > 0 else 0.0
            
            st.markdown(f"""
            | Quelle | Abs. [ml] | Rel. [ml/kg] |
            | :--- | :--- | :--- |
            | **Spiro (gemessen)** | {int(spiro_abs)} | {spiro_rel:.1f} |
            | **PPO (♂)** | {int(m_abs)} | {m_rel:.1f} |
            | **PPO (♀)** | {int(w_abs)} | {w_rel:.1f} |
            | **Manuell (override)** | {int(override_abs)} | {override_rel:.1f} |
            """)
            
            vo2_override_input = st.number_input("Manuelle VO2max [ml/min]", value=float(st.session_state.vo2_override), step=50.0)
            st.session_state.vo2_override = round(vo2_override_input, 1)
            
            vo2_source = st.selectbox(
                "Aktive Quelle für Simulation",
                ["Gemessene Spirodatei", "Berechnet aus 30s PPO (männlich)", "Berechnet aus 30s PPO (weiblich)", "Manuelle Eingabe"],
                key="vo2_source_rad_key"
            )
            if vo2_source == "Manuelle Eingabe":
                vo2_override = float(st.session_state.vo2_override)
            else:
                vo2_override = 0.0

            st.subheader("Energetik Parameter")
            carb_intake_factor_input = st.number_input("KH-Zufuhr Faktor [g/kg KG/h]", value=float(st.session_state.carb_intake_factor), step=0.1, format="%.1f")
            st.session_state.carb_intake_factor = round(carb_intake_factor_input, 1)
            carb_intake_factor = float(st.session_state.carb_intake_factor)

        # 4. Coaching Potential Overrides (Default collapsed)
        with st.expander("Coaching Potential", expanded=False):
            st.caption("Simuliere den Effekt von Anpassungen bei VO2max, VLamax oder Gewicht.")
            pot_vo2_sel = st.selectbox(
                "VO2max Potential",
                ["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"],
                index=["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"].index(st.session_state.get("pot_vo2_select_key", "Keine Änderung")),
                key="pot_vo2_select_key"
            )
            pot_vla_sel = st.selectbox(
                "VLamax Potential",
                ["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"],
                index=["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"].index(st.session_state.get("pot_vla_select_key", "Keine Änderung")),
                key="pot_vla_select_key"
            )
            pot_weight_sel = st.selectbox(
                "Gewicht Potential",
                ["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"],
                index=["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"].index(st.session_state.get("pot_weight_select_key", "Keine Änderung")),
                key="pot_weight_select_key"
            )
            
            if st.session_state.get("rad_auswertung_gestartet", False):
                st.markdown(f"""
                **Simuliertes Coaching-Potential:**
                * **Gewicht:** {st.session_state.get('pot_weight_val', st.session_state.weight):.1f} kg
                * **VO2max:** {st.session_state.get('pot_vo2max', 72.0):.1f} ml/min/kg
                * **VLamax:** {st.session_state.get('pot_vlamax', 0.45):.3f} mmol/L/s
                * **ANS:** {int(st.session_state.get('pot_ans_abs', 295.0))} W ({st.session_state.get('pot_ans_rel', 4.8):.2f} W/kg)
                * **Fatmax:** {st.session_state.get('pot_fatmax', 3.4):.2f} W/kg
                """)

        st.markdown("---")
        start_button_rad = st.form_submit_button("Auswertung starten", type="primary")
        if start_button_rad:
            st.session_state.rad_auswertung_gestartet = True
            st.session_state.df_sprint_lac = edited_sprint_lac
            st.session_state.df_sf = edited_sf
            
    # Download-Button: konvertierte Excel aus SRM/FIT erstellen (outside the form)
    if "sprint_records" in st.session_state and st.session_state.get("last_uploaded_sprint_raw"):
        try:
            import importlib.util as _ilu2
            _spec2 = _ilu2.spec_from_file_location(
                "convert_sprint2",
                os.path.join(os.path.dirname(os.path.abspath(__file__)), "convert_sprint_to_excel.py")
            )
            _conv2 = _ilu2.module_from_spec(_spec2)
            _spec2.loader.exec_module(_conv2)
            
            _dl_bytes = _conv2.create_sprint_excel_bytes(
                records=st.session_state.sprint_records,
                athlete_name=st.session_state.get("athlete_name", ""),
                test_date=st.session_state.get("test_date", ""),
                source_filename=st.session_state.get("sprint_src_name", "sprint"),
                rec_int_label=st.session_state.get("sprint_rec_int_label", ""),
            )
            _dl_name = st.session_state.get("last_uploaded_sprint", "sprint").rsplit(".", 1)[0] + "_Data_Sprint.xlsx"
            st.sidebar.download_button(
                label="📥 Sprint Excel herunterladen",
                data=_dl_bytes,
                file_name=_dl_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_sprint_excel_key"
            )
        except Exception:
            pass  # Download optional
        
    # 6. RENDER MAIN PAGE LAYOUT
    if st.session_state.rad_auswertung_gestartet:
        # Fetch base variables from key-bound session state
        weight = float(st.session_state.get("weight", 72.3))
        height = int(st.session_state.get("height", 178))
        body_fat_pct = float(st.session_state.get("body_fat_pct", 15.2))
        athlete_name = st.session_state.get("athlete_name", "Michael Wagner")
        birthdate = st.session_state.get("birthdate", "06.03.1996")
        test_date = st.session_state.get("test_date", "22.05.2026")
        kategorie = st.session_state.get("kategorie", "Amateur")
        coach = st.session_state.get("coach", "Markus Hertlein")
        
        ppo_val = float(st.session_state.get("ppo_val", 436.0))
        hfmax_val = float(st.session_state.get("hfmax_val", 179.0))
        slope_val = float(st.session_state.get("slope_val", 0.11))
        intercept_val = float(st.session_state.get("intercept_val", 138.0))
        carb_intake_factor = float(st.session_state.get("carb_intake_factor", 1.0))

        # A. Check and compute VO2max
        if vo2_source == "Gemessene Spirodatei":
            abs_vo2max = float(st.session_state.vo2max_spiro)
            vo2_source_lbl = f"Spirodatei ({abs_vo2max:.1f} ml/min)"
        elif vo2_source == "Berechnet aus 30s PPO (männlich)":
            abs_vo2max = float(st.session_state.ppo_val) * 9.44 + 592.0
            vo2_source_lbl = f"30s PPO männlich ({abs_vo2max:.1f} ml/min)"
        elif vo2_source == "Berechnet aus 30s PPO (weiblich)":
            abs_vo2max = float(st.session_state.ppo_val) * 9.44 + 275.0
            vo2_source_lbl = f"30s PPO weiblich ({abs_vo2max:.1f} ml/min)"
        else:
            abs_vo2max = float(vo2_override)
            vo2_source_lbl = f"Manuell überschrieben ({abs_vo2max:.1f} ml/min)"
            
        rel_vo2max = abs_vo2max / weight if weight > 0 else 0.0
        
        # B. Load Sprint parameters based on active source
        sprint_src = st.session_state.get("sprint_source_key", "Automatisch aus Excel-Rohdaten")
        if sprint_src == "Automatisch aus Excel-Rohdaten":
            st.session_state.t_bel = float(st.session_state.get("t_bel_auto", 13.2))
            st.session_state.t_alak = float(st.session_state.get("t_alak_auto", 3.1))
        else:
            st.session_state.t_bel = float(st.session_state.get("t_bel_manual", 13.2))
            st.session_state.t_alak = float(st.session_state.get("t_alak_manual", 3.1))
        st.session_state.t_glyc = round(st.session_state.t_bel - st.session_state.t_alak, 1)
        
        t_bel = float(st.session_state.t_bel)
        t_alak = float(st.session_state.t_alak)
        t_glyc = float(st.session_state.t_glyc)
        
        # Extract Lactates from table
        lac_vals = st.session_state.df_sprint_lac["Laktat [mmol/L]"].values
        ruhe_lacs = lac_vals[0:3]
        avg_rest_lac = np.mean(ruhe_lacs)
        post_lacs = lac_vals[3:]
        max_post_lac = np.max(post_lacs)
        
        # Calculate VLamax
        vlamax = (max_post_lac - avg_rest_lac) / t_glyc if t_glyc > 0 else 0.0
        
        # C. Run metabolic grid simulation (0 to 600 W) using session state constants
        sim_results = []
        ks1 = float(st.session_state.ks1)
        ks2 = float(st.session_state.ks2)
        ks4 = float(st.session_state.ks4)
        lac_eq = float(st.session_state.lac_eq)
        lac_dist = float(st.session_state.lac_dist)
        vo2_rest = float(st.session_state.vo2_rest)
        
        for p in range(601):
            vo2ss = (p * ks4 / weight) + vo2_rest
            
            # ADP calculation (using numpy to handle invalid operations gracefully)
            if rel_vo2max - vo2ss > 0:
                adp = np.sqrt((ks1 * vo2ss) / (rel_vo2max - vo2ss))
            else:
                adp = np.nan
                
            adp_3 = adp ** 3 if not np.isnan(adp) else np.nan
            
            # VLass (mmol/L/min)
            vlass = (vlamax * 60.0) / (1.0 + ks2 / adp_3) if not np.isnan(adp_3) and adp_3 > 0 else np.nan
            
            # VLaoxmax (mmol/L/min)
            vlaoxmax = (lac_eq / lac_dist) * vo2ss
            
            # Pyruvatdefizit (VLaoxmax - VLass)
            pyruvat_def = vlaoxmax - vlass if not np.isnan(vlass) else np.nan
            
            # Fuel splitting (capped for energetics and table display to match physiology)
            kh_pct = (vlass / vlaoxmax * 100.0) if vlaoxmax > 0 and not np.isnan(vlass) else np.nan
            fat_pct = 100.0 - kh_pct if not np.isnan(kh_pct) else np.nan
            
            if not np.isnan(kh_pct):
                kh_pct_capped = max(0.0, min(100.0, kh_pct))
                fat_pct_capped = max(0.0, min(100.0, 100.0 - kh_pct_capped))
            else:
                kh_pct_capped = np.nan
                fat_pct_capped = np.nan
            
            # RQ
            rq = 0.7 + 0.3 * (kh_pct_capped / 100.0) if not np.isnan(kh_pct_capped) else np.nan
            
            # AE (Energy equivalent, kJ/L O2)
            ae = (((rq - 0.7) * 0.05094) * 100.0) + 19.6 if not np.isnan(rq) else np.nan
            
            # EE gesamt [kcal/h]
            vo2_abs_ss = vo2ss * weight
            ee_gesamt = ((vo2_abs_ss / 1000.0) * ae / 4.186) * 60.0 if not np.isnan(ae) else np.nan
            
            # EE Fett [kcal/h]
            ee_fett = ee_gesamt * (fat_pct_capped / 100.0) if not np.isnan(ee_gesamt) and not np.isnan(fat_pct_capped) else np.nan
            
            # EE KH [g/h]
            ee_kh_g = (ee_gesamt * (kh_pct_capped / 100.0)) / 4.063 if not np.isnan(ee_gesamt) and not np.isnan(kh_pct_capped) else np.nan
            
            # HR
            hr_val = p * slope_val + intercept_val
            if hr_val > hfmax_val:
                hr_val = np.nan
                
            sim_results.append({
                'power': p,
                'vo2ss': vo2ss,
                'vlass': vlass,
                'vlaoxmax': vlaoxmax,
                'pyruvat_def': pyruvat_def,
                'kh_pct': kh_pct_capped,
                'fat_pct': fat_pct_capped,
                'rq': rq,
                'ee_gesamt': ee_gesamt,
                'ee_fett': ee_fett,
                'ee_kh_g': ee_kh_g,
                'hr': hr_val
            })
            
        df_sim = pd.DataFrame(sim_results)
        
        # D. Extract Derived Metrics
        # ANS crossover: power step where vlass > vlaoxmax and the positive difference is minimized
        ans_df = df_sim[df_sim['vlass'] > df_sim['vlaoxmax']].copy()
        if not ans_df.empty:
            diff = ans_df['vlass'] - ans_df['vlaoxmax']
            min_diff_idx = diff.idxmin()
            ans_power = int(df_sim.loc[min_diff_idx, 'power'])
        else:
            ans_power = 0
            
        ans_rel = ans_power / weight
        ans_hr = ans_power * slope_val + intercept_val
        
        # Note the Excel formula inconsistency in Report_Cycling!AA16 (adds vo2_rest before weight division)
        ans_vo2_rel = ((ans_power * ks4) + vo2_rest) / weight
        ans_vo2_abs = ans_vo2_rel * weight
        ans_ee_gesamt = df_sim.loc[df_sim['power'] == ans_power, 'ee_gesamt'].values[0]
        ans_ee_kh_g = df_sim.loc[df_sim['power'] == ans_power, 'ee_kh_g'].values[0]
        
        # Fatmax: power where pyruvat_def (vlaoxmax - vlass) is maximized
        fatmax_df = df_sim.dropna(subset=['pyruvat_def']).copy()
        fatmax_idx = fatmax_df['pyruvat_def'].idxmax()
        fatmax_row = df_sim.loc[fatmax_idx]
        fatmax_power = int(fatmax_row['power'])
        fatmax_rel = fatmax_power / weight
        fatmax_ee_gesamt = fatmax_row['ee_gesamt']
        fatmax_ee_fett = fatmax_row['ee_fett']
        
        # MAP (max. Leistung)
        map_val = (abs_vo2max - weight * vo2_rest) / ks4
        
        # Match-Schwelle: power where ee_kh_g matches carb intake
        carb_intake_factor = float(st.session_state.carb_intake_factor)
        carb_intake = carb_intake_factor * weight
        carb_match_df = df_sim[df_sim['ee_kh_g'] <= carb_intake]
        if not carb_match_df.empty:
            carb_match_power = int(carb_match_df['power'].max())
        else:
            carb_match_power = 0
            
        # Body mass indices
        bmi = weight / ((height/100.0)**2) if height > 0 else 0.0
        fett_kg = weight * (body_fat_pct / 100.0)
        fettfrei_kg = weight - fett_kg
        vo2_ffm = abs_vo2max / fettfrei_kg if fettfrei_kg > 0 else 0.0
        
        # E. Coaching Potential Overrides Recalculation
        pot_vo2_sel = st.session_state.get("pot_vo2_select_key", "Keine Änderung")
        pot_vla_sel = st.session_state.get("pot_vla_select_key", "Keine Änderung")
        pot_weight_sel = st.session_state.get("pot_weight_select_key", "Keine Änderung")
        
        def parse_sel_pct(sel_str):
            if "+2" in sel_str: return 0.02
            if "+5" in sel_str: return 0.05
            if "-2" in sel_str: return -0.02
            if "-5" in sel_str: return -0.05
            return 0.0
            
        vo2_change = parse_sel_pct(pot_vo2_sel)
        vla_change = parse_sel_pct(pot_vla_sel)
        weight_change = parse_sel_pct(pot_weight_sel)
        
        pot_weight = weight * (1.0 + weight_change)
        pot_abs_vo2max = abs_vo2max * (1.0 + vo2_change)
        pot_rel_vo2max = pot_abs_vo2max / pot_weight if pot_weight > 0 else 0.0
        pot_vlamax = vlamax * (1.0 + vla_change)
        
        # Save values for display in the expander and pass to PDF
        st.session_state.pot_weight_val = pot_weight
        st.session_state.pot_vlamax = pot_vlamax
        st.session_state.pot_vo2max = pot_rel_vo2max
        st.session_state.pot_fat = body_fat_pct
        
        # Run potential metabolic simulation
        pot_sim_results = []
        for p in range(601):
            vo2ss = (p * ks4 / pot_weight) + vo2_rest
            if pot_rel_vo2max - vo2ss > 0:
                adp = np.sqrt((ks1 * vo2ss) / (pot_rel_vo2max - vo2ss))
            else:
                adp = np.nan
            adp_3 = adp ** 3 if not np.isnan(adp) else np.nan
            vlass = (pot_vlamax * 60.0) / (1.0 + ks2 / adp_3) if not np.isnan(adp_3) and adp_3 > 0 else np.nan
            vlaoxmax = (lac_eq / lac_dist) * vo2ss
            pyruvat_def = vlaoxmax - vlass if not np.isnan(vlass) else np.nan
            kh_pct = (vlass / vlaoxmax * 100.0) if vlaoxmax > 0 and not np.isnan(vlass) else np.nan
            fat_pct = 100.0 - kh_pct if not np.isnan(kh_pct) else np.nan
            
            if not np.isnan(kh_pct):
                kh_pct_capped = max(0.0, min(100.0, kh_pct))
                fat_pct_capped = max(0.0, min(100.0, 100.0 - kh_pct_capped))
            else:
                kh_pct_capped = np.nan
                fat_pct_capped = np.nan
                
            rq = 0.7 + 0.3 * (kh_pct_capped / 100.0) if not np.isnan(kh_pct_capped) else np.nan
            ae = (((rq - 0.7) * 0.05094) * 100.0) + 19.6 if not np.isnan(rq) else np.nan
            vo2_abs_ss = vo2ss * pot_weight
            ee_gesamt = ((vo2_abs_ss / 1000.0) * ae / 4.186) * 60.0 if not np.isnan(ae) else np.nan
            ee_fett = ee_gesamt * (fat_pct_capped / 100.0) if not np.isnan(ee_gesamt) and not np.isnan(fat_pct_capped) else np.nan
            ee_kh_g = (ee_gesamt * (kh_pct_capped / 100.0)) / 4.063 if not np.isnan(ee_gesamt) and not np.isnan(kh_pct_capped) else np.nan
            
            pot_sim_results.append({
                'power': p,
                'vo2ss': vo2ss,
                'vlass': vlass,
                'vlaoxmax': vlaoxmax,
                'pyruvat_def': pyruvat_def,
                'kh_pct': kh_pct_capped,
                'fat_pct': fat_pct_capped,
                'rq': rq,
                'ee_gesamt': ee_gesamt,
                'ee_fett': ee_fett,
                'ee_kh_g': ee_kh_g
            })
            
        df_pot_sim = pd.DataFrame(pot_sim_results)
        
        # Calculate pot_ans_power
        pot_ans_df = df_pot_sim[df_pot_sim['vlass'] > df_pot_sim['vlaoxmax']].copy()
        if not pot_ans_df.empty:
            diff = pot_ans_df['vlass'] - pot_ans_df['vlaoxmax']
            min_diff_idx = diff.idxmin()
            pot_ans_power = int(df_pot_sim.loc[min_diff_idx, 'power'])
        else:
            pot_ans_power = 0
            
        pot_ans_rel = pot_ans_power / pot_weight if pot_weight > 0 else 0.0
        
        # Calculate pot_fatmax_power
        pot_fatmax_df = df_pot_sim.dropna(subset=['pyruvat_def']).copy()
        pot_fatmax_idx = pot_fatmax_df['pyruvat_def'].idxmax()
        pot_fatmax_row = df_pot_sim.loc[pot_fatmax_idx]
        pot_fatmax_power = int(pot_fatmax_row['power'])
        pot_fatmax_rel = pot_fatmax_power / pot_weight if pot_weight > 0 else 0.0
        
        # pot_match_power
        pot_carb_intake = carb_intake_factor * pot_weight
        pot_carb_match_df = df_pot_sim[df_pot_sim['ee_kh_g'] <= pot_carb_intake]
        if not pot_carb_match_df.empty:
            pot_carb_match_power = int(pot_carb_match_df['power'].max())
        else:
            pot_carb_match_power = 0
        pot_match_rel = pot_carb_match_power / pot_weight if pot_weight > 0 else 0.0
        
        # pot_pmax_rel
        pot_sprint_powers = st.session_state.get("sprint_powers", [])
        if not pot_sprint_powers or len(pot_sprint_powers) < 2:
            pot_sprint_powers = [max(0.0, 1200.0 * (1.0 - ((t - 2.0)**2) / 64.0)) for t in [i * 0.1 for i in range(120)]]
        pot_max_sprint_power = np.max(pot_sprint_powers)
        pot_pmax_rel = pot_max_sprint_power / pot_weight if pot_weight > 0 else 0.0
        
        st.session_state.pot_ans_abs = float(pot_ans_power)
        st.session_state.pot_ans_rel = float(pot_ans_rel)
        st.session_state.pot_fatmax = float(pot_fatmax_rel)
        st.session_state.pot_match = float(pot_match_rel)
        st.session_state.pot_pmax = float(pot_pmax_rel)
        
        # Render Main Panels
        st.markdown(f"""
            <div style="font-size: 16px; line-height: 1.5; margin-bottom: 20px;">
                <b>Athlet:</b> {athlete_name}<br>
                <b>Sportart:</b> Radsport<br>
                <b>Test-Datum:</b> {test_date}<br>
                <b>Kategorie:</b> {kategorie} | <b>Coach:</b> {coach}
            </div>
            """, unsafe_allow_html=True)
            
        st.subheader("Stammdaten & Anthropometrie")
        col_base_1, col_base_2, col_base_3 = st.columns(3)
        with col_base_1:
            st.markdown(f"**Gewicht:** {weight:.1f} kg")
            st.markdown(f"**Größe:** {height} cm")
            st.markdown(f"**BMI:** {bmi:.2f} kg/m²")
        with col_base_2:
            st.markdown(f"**Körperfett:** {body_fat_pct:.1f} %")
            st.markdown(f"**Fettmasse:** {fett_kg:.1f} kg")
            st.markdown(f"**Fettfreie Masse:** {fettfrei_kg:.1f} kg")
        with col_base_3:
            st.markdown(f"**Geburtsdatum:** {birthdate}")
            st.markdown(f"**Testdatum:** {test_date}")
            st.markdown(f"**Alter:** {st.session_state.get('birthdate', birthdate)}")
            
        st.markdown("---")
        st.subheader("⚡ Physiologische Kennzahlen & Ergebnisse")
        
        # Grid layout for metrics
        c_1, c_2, c_3, c_4 = st.columns(4)
        with c_1:
            st.metric("VO2max (abs.)", f"{int(abs_vo2max)} ml/min")
            st.caption(f"Quelle: {vo2_source_lbl}")
        with c_2:
            st.metric("VO2max (rel.)", f"{rel_vo2max:.1f} ml/min/kg")
            st.caption(f"FFM: {vo2_ffm:.1f} ml/min/kg FFM")
        with c_3:
            st.metric("VLamax", f"{vlamax:.3f} mmol/l/s")
            st.caption(f"t_alak: {t_alak:.1f}s | t_bel: {t_bel:.1f}s")
        with c_4:
            st.metric("MAP (max. Leistung)", f"{map_val:.1f} W")
            st.caption(f"relativ: {map_val/weight:.2f} W/kg")
            
        c_5, c_6, c_7, c_8 = st.columns(4)
        with c_5:
            st.metric("ANS Schwelle (PMLSSc)", f"{ans_power} W")
            st.caption(f"rel: {ans_rel:.2f} W/kg | EE ges: {int(ans_ee_gesamt)} kcal/h")
        with c_6:
            st.metric("Herzfrequenz @ ANS", f"{int(ans_hr)} bpm")
            st.caption(f"VO2 abs: {int(ans_vo2_abs)} ml/min")
        with c_7:
            st.metric("Fatmax Leistung", f"{fatmax_power} W")
            st.caption(f"rel: {fatmax_rel:.2f} W/kg | EE Fett: {int(fatmax_ee_fett)} kcal/h")
        with c_8:
            st.metric("Match-Schwelle", f"{carb_match_power} W")
            st.caption(f"rel: {carb_match_power/weight:.2f} W/kg | Zufuhr: {int(carb_intake)} g/h")

        st.markdown("---")
        
        # Sprint-Verifizierung Section
        st.subheader("⏱️ Sprint-Verifizierung & VLamax-Berechnung")
        col_sprint_1, col_sprint_2 = st.columns([2, 1])
        
        with col_sprint_1:
            if "sprint_powers" in st.session_state and len(st.session_state.sprint_powers) > 0:
                # Plotly figure for Sprint Verification
                fig_sprint = go.Figure()
                
                # Line for sprint powers
                fig_sprint.add_trace(go.Scatter(
                    x=st.session_state.sprint_times,
                    y=st.session_state.sprint_powers,
                    mode='lines',
                    line=dict(color='#00a1e0', width=3),
                    name='Leistung'
                ))
                
                # Find p_max and t_pmax_last to draw lines
                sprint_pmax = max(st.session_state.sprint_powers)
                sprint_palak = 0.965 * sprint_pmax
                
                # Add horizontal line for P_alak (96.5% Pmax)
                fig_sprint.add_hline(
                    y=sprint_palak,
                    line_dash="dot",
                    line_color="#cdb663",
                    annotation_text=f"96.5% Pmax ({sprint_palak:.0f} W)",
                    annotation_position="bottom left"
                )
                
                # Shaded region for Alactic phase (0 to t_alak)
                fig_sprint.add_vrect(
                    x0=0.0, x1=t_alak,
                    fillcolor="rgba(0, 161, 224, 0.1)",
                    layer="below", line_width=0,
                    annotation_text="Alaktazid (t_alak)",
                    annotation_position="top left"
                )
                
                # Shaded region for Glycolytic phase (t_alak to t_bel)
                fig_sprint.add_vrect(
                    x0=t_alak, x1=t_bel,
                    fillcolor="rgba(205, 182, 99, 0.1)",
                    layer="below", line_width=0,
                    annotation_text="Glykolytisch (t_glyc)",
                    annotation_position="top right"
                )
                
                fig_sprint.update_layout(
                    xaxis_title="Zeit [Sekunden]",
                    yaxis_title="Leistung [Watt]",
                    height=350,
                    margin=dict(l=40, r=40, t=20, b=40),
                    plot_bgcolor='#f9f9f9',
                    legend=dict(x=0.02, y=0.98, bgcolor="rgba(255,255,255,0.7)")
                )
                st.plotly_chart(fig_sprint, use_container_width=True)
            else:
                st.info("ℹ️ Keine Sprint-Rohdaten hochgeladen. Es wird die standardmäßig konfigurierte Belastungsdauer verwendet.")
                
        with col_sprint_2:
            st.markdown(f"""
            <div style="border: 1px solid #e6e6e6; border-radius: 8px; padding: 15px; background-color: #f9f9f9; color: #222;">
                <h4 style="margin-top:0; color: #111;">Dauer der Phasen</h4>
                <table style="width:100%; border-collapse: collapse;">
                    <tr style="border-bottom: 1px solid #e6e6e6;">
                        <td style="padding: 6px 0;"><b>Gesamtdauer (t_bel):</b></td>
                        <td style="text-align:right;">{t_bel:.1f} s</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #e6e6e6;">
                        <td style="padding: 6px 0;"><b>Alaktazide Phase (t_alak):</b></td>
                        <td style="text-align:right;">{t_alak:.1f} s</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #e6e6e6; color: #cdb663;">
                        <td style="padding: 6px 0;"><b>Glykolytische Phase (t_glyc):</b></td>
                        <td style="text-align:right;"><b>{t_glyc:.1f} s</b></td>
                    </tr>
                </table>
                <h4 style="margin-top: 15px; color: #111;">Laktat-Differenz</h4>
                <table style="width:100%; border-collapse: collapse;">
                    <tr style="border-bottom: 1px solid #e6e6e6;">
                        <td style="padding: 6px 0;">Ruhelaktat (Mittelwert):</td>
                        <td style="text-align:right;">{avg_rest_lac:.2f} mmol/L</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #e6e6e6;">
                        <td style="padding: 6px 0;">Peak-Laktat (post-sprint):</td>
                        <td style="text-align:right;">{max_post_lac:.2f} mmol/L</td>
                    </tr>
                    <tr style="border-bottom: 1px solid #e6e6e6; color: #00a1e0;">
                        <td style="padding: 6px 0;"><b>Laktat-Anstieg (&Delta;La):</b></td>
                        <td style="text-align:right;"><b>{(max_post_lac - avg_rest_lac):.2f} mmol/L</b></td>
                    </tr>
                </table>
                <h4 style="margin-top:15px; margin-bottom:5px; color: #111;">VLamax Berechnung</h4>
                <div style="font-family: monospace; font-size:13px; background-color:#ffffff; padding:10px; border-radius:4px; border: 1px solid #e6e6e6; text-align:center; color: #111;">
                    VLamax = &Delta;La / t_glyc <br>
                    VLamax = {max_post_lac - avg_rest_lac:.2f} / {t_glyc:.1f} <br>
                    <b>VLamax = {vlamax:.3f} mmol/l/s</b>
                </div>
            </div>
            """, unsafe_allow_html=True)
            
        st.markdown("---")
        
        # Diagrams layout
        col_plot_1, col_plot_2 = st.columns(2)
        
        with col_plot_1:
            st.subheader("📈 Laktatkinetik (Metabolischer Crossover)")
            
            # Plotly figure for crossover
            fig_cross = go.Figure()
            # Calculate dynamic max W range based on ANS + 50W (rounded to next 10W)
            max_x_plot = int(min(600, max(300, np.ceil((ans_power + 50) / 10.0) * 10.0)))
            df_plot = df_sim[df_sim['power'] <= max_x_plot]
            
            fig_cross.add_trace(go.Scatter(
                x=df_plot['power'], y=df_plot['vlass'],
                mode='lines',
                line=dict(color='#00a1e0', width=3),
                name='Laktatbildung (VLass)'
            ))
            fig_cross.add_trace(go.Scatter(
                x=df_plot['power'], y=df_plot['vlaoxmax'],
                mode='lines',
                line=dict(color='#cdb663', width=3),
                name='max. Laktatabbau (VLaoxmax)'
            ))
            
            # Add vertical line for ANS
            fig_cross.add_vline(
                x=ans_power,
                line_dash="dash",
                line_color="#595a59",
                annotation_text=f"ANS ({ans_power} W)",
                annotation_position="top right"
            )
            
            fig_cross.update_layout(
                xaxis=dict(title="Leistung [Watt]", range=[0, max_x_plot]),
                yaxis_title="Laktatkinetik [mmol/L/min]",
                height=400,
                legend=dict(x=0.02, y=0.98, bgcolor="rgba(255,255,255,0.7)"),
                margin=dict(l=40, r=40, t=20, b=40),
                plot_bgcolor='#f9f9f9'
            )
            st.plotly_chart(fig_cross, use_container_width=True)
            
        with col_plot_2:
            st.subheader("📊 Substratumsatz & Energiefluss")
            
            # Plotly figure with twin y-axes
            fig_met = go.Figure()
            df_plot = df_sim[df_sim['power'] <= max_x_plot]
            
            # EE gesamt
            fig_met.add_trace(go.Scatter(
                x=df_plot['power'], y=df_plot['ee_gesamt'],
                mode='lines',
                line=dict(color='#595a59', width=2.5, dash='dash'),
                name='Gesamtumsatz [kcal/h]',
                yaxis='y'
            ))
            # EE Fett
            fig_met.add_trace(go.Scatter(
                x=df_plot['power'], y=df_plot['ee_fett'],
                mode='lines',
                line=dict(color='#2cb7b9', width=3),
                name='Fettverbrennung [kcal/h]',
                yaxis='y'
            ))
            # KH-Verbrauch
            fig_met.add_trace(go.Scatter(
                x=df_plot['power'], y=df_plot['ee_kh_g'],
                mode='lines',
                line=dict(color='#cdb663', width=3),
                name='KH-Verbrennung [g/h]',
                yaxis='y2'
            ))
            
            # Vertical line for Fatmax
            fig_met.add_vline(
                x=fatmax_power,
                line_dash="dash",
                line_color="#2cb7b9",
                annotation_text=f"Fatmax ({fatmax_power} W)",
                annotation_position="top left"
            )
            
            # Twin axis layout
            fig_met.update_layout(
                xaxis=dict(title="Leistung [Watt]", range=[0, max_x_plot]),
                yaxis=dict(
                    title=dict(text="Energie [kcal/h]", font=dict(color="#595a59")),
                    tickfont=dict(color="#595a59")
                ),
                yaxis2=dict(
                    title=dict(text="Kohlenhydratumsatz [g/h]", font=dict(color="#cdb663")),
                    tickfont=dict(color="#cdb663"),
                    overlaying="y",
                    side="right"
                ),
                height=400,
                legend=dict(x=0.02, y=0.98, bgcolor="rgba(255,255,255,0.7)"),
                margin=dict(l=45, r=45, t=20, b=40),
                plot_bgcolor='#f9f9f9'
            )
            st.plotly_chart(fig_met, use_container_width=True)

        st.markdown("---")
        
        # Detailed Grid Table
        with st.expander("🔍 Detaillierter metabolischer Verlauf (Tabelle in 10W-Schritten)"):
            df_table_rows = []
            for p in range(0, max_x_plot + 1, 10):
                rows = df_sim[df_sim['power'] == p]
                if not rows.empty:
                    row_data = rows.iloc[0]
                    df_table_rows.append({
                        'Leistung [W]': p,
                        'VO2ss [ml/min/kg]': round(row_data['vo2ss'], 1) if not np.isnan(row_data['vo2ss']) else "-",
                        'VLass [mmol/L/min]': round(row_data['vlass'], 3) if not np.isnan(row_data['vlass']) else "-",
                        'VLaoxmax [mmol/L/min]': round(row_data['vlaoxmax'], 3) if not np.isnan(row_data['vlaoxmax']) else "-",
                        'Pyruvatdefizit [mmol/L/min]': round(row_data['pyruvat_def'], 3) if not np.isnan(row_data['pyruvat_def']) else "-",
                        'KH-Anteil [%]': f"{round(row_data['kh_pct'], 1)}%" if not np.isnan(row_data['kh_pct']) else "-",
                        'Fett-Anteil [%]': f"{round(row_data['fat_pct'], 1)}%" if not np.isnan(row_data['fat_pct']) else "-",
                        'kalk. RQ': round(row_data['rq'], 3) if not np.isnan(row_data['rq']) else "-",
                        'EE gesamt [kcal/h]': int(round(row_data['ee_gesamt'], 0)) if not np.isnan(row_data['ee_gesamt']) else "-",
                        'EE Fett [kcal/h]': int(round(row_data['ee_fett'], 0)) if not np.isnan(row_data['ee_fett']) else "-",
                        'KH-Verbrauch [g/h]': round(row_data['ee_kh_g'], 1) if not np.isnan(row_data['ee_kh_g']) else "-",
                        'HF [bpm]': f"{int(round(row_data['hr'], 0))}" if not np.isnan(row_data['hr']) else "-"
                    })
            st.dataframe(pd.DataFrame(df_table_rows), hide_index=True, width='stretch')

        # --- RADSPORT PDF DOWNLOAD ---
        st.markdown("---")
        st.subheader("📄 Report-Generierung")
        
        # 1. Sprint raw values extraction or dummy fallback
        sprint_times = st.session_state.get("sprint_times", [])
        sprint_powers = st.session_state.get("sprint_powers", [])
        sprint_cadences = st.session_state.get("sprint_cadences", [])
        
        if not sprint_powers or len(sprint_powers) < 2:
            # Fallback dummy sprint values if none uploaded
            sprint_times = [i * 0.1 for i in range(120)]
            # Peak power around 1200W
            sprint_powers = [max(0.0, 1200.0 * (1.0 - ((t - 2.0)**2) / 64.0)) for t in sprint_times]
            sprint_cadences = [max(40.0, 110.0 - 2.0 * t) for t in sprint_times]

        # 2. Calculate sprint force metrics using weight and 0.1725 crank length
        avg_sprint_power = np.mean(sprint_powers)
        max_sprint_power = np.max(sprint_powers)
        
        forces = []
        for p, c in zip(sprint_powers, sprint_cadences):
            if c > 0:
                # Force F = Power / (crank_length * cadence * 2*pi / 60)
                f = p / (0.1725 * c * 2.0 * np.pi / 60.0)
            else:
                f = 0.0
            forces.append(f)
            
        max_force = np.max(forces) if forces else 0.0
        
        # Force at Pmax (find index of maximum power)
        pmax_idx = np.argmax(sprint_powers) if sprint_powers else 0
        force_at_pmax = forces[pmax_idx] if forces and pmax_idx < len(forces) else 0.0
        
        force_pct_max = (force_at_pmax / max_force * 100.0) if max_force > 0 else 0.0
        
        # 3. Dynamic percentage of VO2max at Anaerobic Threshold
        ans_vo2_pct = (ans_vo2_abs / abs_vo2max * 100.0) if abs_vo2max > 0 else 0.0
        rel_vo2max = (abs_vo2max / weight) if weight > 0 else 0.0
        
        # 4. Coaching potential overrides
        pot_vo2max = float(st.session_state.get("pot_vo2max", 72.0))
        pot_fat = float(st.session_state.get("pot_fat", 10.5))
        pot_ans_abs = float(st.session_state.get("pot_ans_abs", 295.0))
        pot_ans_rel = float(st.session_state.get("pot_ans_rel", 4.8))
        pot_fatmax = float(st.session_state.get("pot_fatmax", 3.4))
        pot_pmax = float(st.session_state.get("pot_pmax", 14.5))
        pot_match = float(st.session_state.get("pot_match", 3.6))
        pot_vlamax = float(st.session_state.get("pot_vlamax", 0.45))
        
        # Create PDF Bytes
        pdf_bytes = None
        pdf_error = None
        with st.spinner("PDF wird erstellt..."):
            try:
                pdf_bytes = create_pdf_rad(
                    athlete_name=athlete_name,
                    birthdate=birthdate,
                    test_date=test_date,
                    test_type=st.session_state.get("diagnostik_type", "Cycling BLUE"),
                    weight=weight,
                    body_fat_pct=body_fat_pct,
                    rel_vo2max=rel_vo2max,
                    ans_power=ans_power,
                    ans_rel=ans_rel,
                    ans_ee_gesamt=ans_ee_gesamt,
                    ans_hr=ans_hr,
                    ans_vo2_abs=ans_vo2_abs,
                    ans_vo2_pct=ans_vo2_pct,
                    ans_ee_kh_g=ans_ee_kh_g,
                    fatmax_power=fatmax_power,
                    fatmax_rel=fatmax_rel,
                    fatmax_ee_gesamt=fatmax_ee_gesamt,
                    fatmax_ee_fett=fatmax_ee_fett,
                    map_val=map_val,
                    carb_match_power=carb_match_power,
                    carb_intake=carb_intake,
                    vlamax=vlamax,
                    avg_rest_lac=avg_rest_lac,
                    max_post_lac=max_post_lac,
                    t_glyc=t_glyc,
                    t_bel=t_bel,
                    t_alak=t_alak,
                    coach=coach,
                    sportart="Radsport",
                    kategorie=kategorie,
                    height=height,
                    slope_val=slope_val,
                    intercept_val=intercept_val,
                    hfmax_val=hfmax_val,
                    df_sim=df_sim,
                    sprint_times=sprint_times,
                    sprint_powers=sprint_powers,
                    sprint_cadences=sprint_cadences,
                    avg_sprint_power=avg_sprint_power,
                    max_sprint_power=max_sprint_power,
                    max_force=max_force,
                    force_at_pmax=force_at_pmax,
                    force_pct_max=force_pct_max,
                    pot_vo2max=pot_vo2max,
                    pot_fat=pot_fat,
                    pot_ans_abs=pot_ans_abs,
                    pot_ans_rel=pot_ans_rel,
                    pot_fatmax=pot_fatmax,
                    pot_pmax=pot_pmax,
                    pot_match=pot_match,
                    pot_vlamax=pot_vlamax,
                    gender=st.session_state.get("gender", "männlich")
                )
            except Exception as e:
                import traceback
                pdf_error = traceback.format_exc()

        if pdf_bytes is not None:
            st.download_button(
                label="📄 PDF Report herunterladen",
                data=pdf_bytes,
                file_name=f"HYCYS_Radsportdiagnostik_{athlete_name.replace(' ', '_')}.pdf",
                mime="application/pdf",
                type="primary",
                key="radsport_pdf_download_btn"
            )
        else:
            st.error(f"Fehler bei der PDF-Erstellung: Bitte Eingaben prüfen.")
            if pdf_error:
                with st.expander("Fehlerdetails anzeigen"):
                    st.code(pdf_error)

    else:
        # Default view before starting the calculation
        st.markdown(f"""
            <div style="font-size: 16px; line-height: 1.5; margin-bottom: 20px;">
                <b>Athlet:</b> {athlete_name}<br>
                <b>Geschlecht:</b> {st.session_state.get("gender", "männlich")}<br>
                <b>Sportart:</b> Radsport<br>
                <b>Test-Datum:</b> {test_date}
            </div>
            """, unsafe_allow_html=True)
            
        st.info("💡 Radsport-Diagnostik: Bitte laden Sie die gewünschten Dateien hoch und klicken Sie in der Seitenleiste auf **'Auswertung starten'**, um die Stoffwechselprofile zu generieren.")
        
        st.subheader("Hochgeladene Dateien")
        c1, c2, c3 = st.columns(3)
        with c1:
            if uploaded_spiro is not None:
                st.success(f"Spirodatei geladen:\n{uploaded_spiro.name}")
                st.caption(f"Berechnete VO2max: {st.session_state.get('vo2max_spiro', 4292.2):.1f} ml/min")
            else:
                st.warning("Keine Spirodatei hochgeladen")
        with c2:
            if uploaded_fit is not None:
                st.success(f".fit-Datei geladen:\n{uploaded_fit.name}")
                st.caption(f"Steigung: {st.session_state.get('slope_val', 0.11):.4f} | Nullstelle: {st.session_state.get('intercept_val', 138):.1f}")
            else:
                st.warning("Keine .fit-Datei hochgeladen")
        with c3:
            if uploaded_sprint is not None:
                st.success(f"Sprintdatei geladen:\n{uploaded_sprint.name}")
            elif st.session_state.get("sprint_src_name"):
                st.success(f"Sprintdatei geladen:\n{st.session_state.sprint_src_name}")
            else:
                st.warning("Keine Sprintdatei hochgeladen")
                
        st.subheader("Eingegebene Laktatwerte")
        st.dataframe(st.session_state.df_sprint_lac, hide_index=True, width='stretch')
        
    st.stop()


# Testprotokoll selection placed at the very top of the sidebar for Laufauswertung
test_types = [
    "HYCYS Standart (5min / +0,4m/s)",
    "Neues Protokoll (3min / +1km/h bzw. 0,278m/s)"
]

# 1. Upload (Spiro file)
uploaded_file = st.sidebar.file_uploader("Spiro-Datei hochladen (.xlsx oder .csv)", type=["xlsx", "csv"], key="uploaded_file_key")

# 2. Testprotokoll (eingeklappt)
with st.sidebar.expander("Testprotokoll", expanded=False):
    test_type = st.selectbox("Testprotokoll", options=test_types, key="test_type_lauft_widget")

    selected_model_name = "OBLA 4.0 (ANS)"
    if test_type == "Neues Protokoll (3min / +1km/h bzw. 0,278m/s)":
        selected_model_name = st.selectbox(
            "Auswertungs-Schwelle",
            options=["OBLA 4.0 (ANS)", "Dmax (Standard)", "Modified Dmax", "LTP2"],
            index=0,
            key="selected_threshold_model_widget"
        )
    
    # Step parameters (inside the expander)
    start_v = st.number_input("Startgeschwindigkeit (m/s)", value=float(st.session_state.start_v), step=0.1, key="start_v_lauft_widget")
    st.session_state.start_v = start_v

    v_increment = st.number_input("Geschwindigkeitssteigerung (m/s)", value=float(st.session_state.v_increment), step=0.001, format="%.3f", key="v_increment_lauft_widget")
    st.session_state.v_increment = v_increment

    anzahl = st.number_input("Anzahl Stufen", value=int(st.session_state.anzahl), min_value=1, max_value=20, step=1, key="anzahl_lauft_widget")
    st.session_state.anzahl = anzahl

    # Collapsible Protokoll-Setup expander (nested, collapsed)
    with st.expander("Protokoll-Setup", expanded=False):
        vorlauf = st.number_input("Ruhemessung (Sekunden)", value=int(st.session_state.vorlauf), step=10, key="vorlauf_lauft_widget")
        st.session_state.vorlauf = vorlauf
        
        stufendauer = st.number_input("Stufendauer (Minuten)", value=float(st.session_state.stufendauer), step=0.5, key="stufendauer_lauft_widget")
        st.session_state.stufendauer = stufendauer
        
        pausendauer = st.number_input("Pausendauer (Sekunden)", value=int(st.session_state.pausendauer), step=5, key="pausendauer_lauft_widget")
        st.session_state.pausendauer = pausendauer
        
        ausbelastung = st.checkbox("Test bis zur Ausbelastung", value=True, key="ausbelastung_lauft_widget")

if test_type != st.session_state.last_test_type:
    st.session_state.last_test_type = test_type
    if test_type == "HYCYS Standart (5min / +0,4m/s)":
        st.session_state.start_v = 2.8
        st.session_state.v_increment = 0.400
        st.session_state.anzahl = 5
        st.session_state.stufendauer = 5.0
    elif test_type == "Neues Protokoll (3min / +1km/h bzw. 0,278m/s)":
        st.session_state.start_v = 2.22
        st.session_state.v_increment = 0.278
        st.session_state.anzahl = 8
        st.session_state.stufendauer = 3.0
    # Ruhemessung and Pausendauer always 60s and 30s
    st.session_state.vorlauf = 60
    st.session_state.pausendauer = 30
    
    # Programmatically update the widgets' internal session state keys
    st.session_state.start_v_lauft_widget = st.session_state.start_v
    st.session_state.v_increment_lauft_widget = st.session_state.v_increment
    st.session_state.anzahl_lauft_widget = st.session_state.anzahl
    st.session_state.stufendauer_lauft_widget = st.session_state.stufendauer
    st.session_state.vorlauf_lauft_widget = st.session_state.vorlauf
    st.session_state.pausendauer_lauft_widget = st.session_state.pausendauer
    
    # Clear lactate table and editor states to force full rebuild
    if "df_lauftest_input" in st.session_state:
        del st.session_state["df_lauftest_input"]
    if "lauftest_lac_editor_key" in st.session_state:
        del st.session_state["lauftest_lac_editor_key"]
    st.session_state.lauf_auswertung_gestartet = False
    for key in list(st.session_state.keys()):
        if key.startswith("lauftest_end_t_") or key.startswith("lauftest_offset_"):
            del st.session_state[key]
        
    st.rerun()

if uploaded_file is not None and uploaded_file.name != st.session_state.last_uploaded_file:
    st.session_state.last_uploaded_file = uploaded_file.name
    try:
        df_excel = pd.read_csv(uploaded_file, header=None, low_memory=False) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file, header=None)
        
        # Name
        name_val = None
        if df_excel.shape[0] > 2 and df_excel.shape[1] > 1:
            name_val = f"{df_excel.iloc[2, 1]} {df_excel.iloc[1, 1]}"
        
        # Height
        h_val = None
        try:
            if df_excel.shape[0] > 5 and df_excel.shape[1] > 1:
                h_val = int(float(df_excel.iloc[5, 1]))
        except:
            pass
            
        # Weight
        w_val = None
        try:
            if df_excel.shape[0] > 6 and df_excel.shape[1] > 1:
                w_val = round(float(df_excel.iloc[6, 1]), 1)
        except:
            pass
            
        # Birthdate
        b_val = None
        if df_excel.shape[0] > 7 and df_excel.shape[1] > 1:
            raw_date = df_excel.iloc[7, 1] 
            if pd.notna(raw_date):
                date_str = str(int(float(raw_date))).zfill(8)
                b_val = f"{date_str[:2]}.{date_str[2:4]}.{date_str[4:]}"
            
        # Testdatum
        t_val = None
        if df_excel.shape[0] > 0 and df_excel.shape[1] > 4:
            raw_test_date = df_excel.iloc[0, 4]
            if pd.notna(raw_test_date):
                try:
                    t_date_str = str(int(float(raw_test_date))).zfill(8)
                    t_val = f"{t_date_str[:2]}.{t_date_str[2:4]}.{t_date_str[4:]}"
                except:
                    t_val = str(raw_test_date)

        # Geschlecht
        g_val = None
        if df_excel.shape[0] > 3 and df_excel.shape[1] > 1:
            raw_gender = df_excel.iloc[3, 1]
            if pd.notna(raw_gender):
                raw_gender_str = str(raw_gender).strip().lower()
                if "w" in raw_gender_str or "fem" in raw_gender_str or "weib" in raw_gender_str:
                    g_val = "weiblich"
                else:
                    g_val = "männlich"
                    
        update_base_data(name_val, h_val, w_val, b_val, t_val, g_val)
        
        # Parse and save spiro data to session state to prevent slow Excel reloading
        try:
            times_sec = df_excel.iloc[3:, 9].apply(parse_time_to_seconds)
            vo2_col = pd.to_numeric(df_excel.iloc[3:, 13], errors='coerce')
            vco2_col = pd.to_numeric(df_excel.iloc[3:, 14], errors='coerce')
            spiro_df = pd.DataFrame({'Time': times_sec, 'VO2': vo2_col, 'VCO2': vco2_col}).dropna(subset=['Time'])
            st.session_state.lauftest_spiro_df = spiro_df
            
            # Detect stage ends and store in session state
            detected = detect_spiro_stage_ends(spiro_df, 20)
            if detected:
                st.session_state.detected_stage_ends = detected
            else:
                if "detected_stage_ends" in st.session_state:
                    del st.session_state.detected_stage_ends
        except Exception as detection_err:
            pass

        # Reset manual end times and offsets
        for key in list(st.session_state.keys()):
            if key.startswith("lauftest_end_t_") or key.startswith("lauftest_offset_"):
                del st.session_state[key]

        # Reset started flag
        st.session_state.lauf_auswertung_gestartet = False

        st.rerun()
    except Exception as e:
        st.sidebar.error(f"Fehler beim Extrahieren der Stammdaten: {e}")

# Kategorie, Coach, and Basisdaten reorganized
# Pre-calculate body fat percentage if skinfolds table has entries (Laufen)
if "df_sf" in st.session_state:
    sf_means = st.session_state.df_sf[["M1", "M2", "M3"]].mean(axis=1)
    sum_sf = sf_means.sum()
    if sum_sf > 0:
        calculated_bf = (22.32 * np.log10(sum_sf)) - 29.2
        st.session_state.body_fat_pct = max(0.0, round(calculated_bf, 1))

# 3. Basisdaten (collapsed expander)
with st.sidebar.expander("Basisdaten (*Edit)", expanded=False):
    athlete_name = st.text_input("Name", value=st.session_state.athlete_name, key="athlete_name_lauft_widget")
    st.session_state.athlete_name = athlete_name

    birthdate = st.text_input("Geburtsdatum", value=st.session_state.birthdate, key="birthdate_lauft_widget")
    st.session_state.birthdate = birthdate

    test_date = st.text_input("Testdatum", value=st.session_state.test_date, key="test_date_lauft_widget")
    st.session_state.test_date = test_date

    # Größe formatted as 123cm without decimals
    height = st.number_input("Größe (cm)", value=int(st.session_state.height), step=1, format="%d", key="height_lauft_widget")
    st.session_state.height = height

    # Gewicht (float, 1 decimal)
    weight = st.number_input("Gewicht (kg)", value=float(st.session_state.weight), step=0.1, format="%.1f", key="weight_lauft_widget")
    st.session_state.weight = round(weight, 1)

    # Körperfett (%) (float, 1 decimal)
    body_fat_pct = st.number_input("Körperfett (%)", value=float(st.session_state.body_fat_pct), step=0.1, format="%.1f", key="bf_lauft_widget")
    st.session_state.body_fat_pct = round(body_fat_pct, 1)

    # Kategorie selectbox
    kategorie_options = ["Hobby", "Age-Grouper", "Amateur", "Profi"]
    kategorie_val = st.session_state.kategorie
    if kategorie_val not in kategorie_options:
        kategorie_options = kategorie_options + [kategorie_val]
    kategorie = st.selectbox("Kategorie", options=kategorie_options, index=kategorie_options.index(kategorie_val), key="kategorie_lauft_widget")
    st.session_state.kategorie = kategorie

    # Geschlecht selectbox
    gender_options = ["männlich", "weiblich"]
    gender_val = st.session_state.get("gender", "männlich")
    gender = st.selectbox("Geschlecht", options=gender_options, index=gender_options.index(gender_val), key="gender_lauft_widget")
    st.session_state.gender = gender

    # Coach selectbox
    coach_options = ["Markus Hertlein", "Marius Trompetter", "Susanne Traser", "Manuel Kuhnle", "Billie Benkel", "Jean Surmont", "Hosea Frick", "Björn Geesmann", "Gregor Eichhorn"]
    coach_val = st.session_state.coach
    if coach_val not in coach_options:
        coach_options = coach_options + [coach_val]
    coach = st.selectbox("Coach", options=coach_options, index=coach_options.index(coach_val), key="coach_lauft_widget")
    st.session_state.coach = coach

    diag_type_options = ["Running WHITE", "Running BLUE", "Running GOLD"]
    diag_type_val = st.session_state.get("diagnostik_type", "Running BLUE")
    if diag_type_val not in diag_type_options:
        diag_type_val = "Running BLUE"
    diagnostik_type = st.selectbox("Diagnostik Typ", options=diag_type_options, index=diag_type_options.index(diag_type_val), key="diagnostik_type_lauft_widget")
    st.session_state.diagnostik_type = diagnostik_type

    # Körperfettmessung (Parizkova 10-Falten)
    with st.expander("Körperfettmessung (Parizkova 10-Falten)"):
        edited_sf = st.data_editor(
            st.session_state.df_sf,
            hide_index=True,
            width='stretch',
            key="edited_sf_key"
        )

# 4. Stufentest Laktat (collapsed expander)
sync_lauftest_input_df()

with st.sidebar.expander("Stufentest Laktat", expanded=False):
    edited_df = st.data_editor(
        st.session_state.df_lauftest_input,
        disabled=["Stufe", "v (m/s)", "v (km/h)"],
        hide_index=True,
        use_container_width=True,
        height=320,
        column_config={
            "Stufe": st.column_config.TextColumn("Stufe", width="small"),
            "v (m/s)": st.column_config.NumberColumn("v (m/s)", format="%.2f", width="small"),
            "v (km/h)": st.column_config.NumberColumn("v (km/h)", format="%.1f", width="small"),
            "Laktat": st.column_config.NumberColumn("Laktat", min_value=0.0, max_value=25.0, step=0.01, format="%.2f", width="small"),
            "HF": st.column_config.NumberColumn("HF", min_value=30, max_value=240, step=1, format="%d", width="small")
        },
        key="lauftest_lac_editor_key"
    )

# 5. Spiro-Fenster Feineinstellung (collapsed expander, outside any form so it reruns immediately)
with st.sidebar.expander("Spiro-Fenster Feineinstellung", expanded=False):
    vo2max_override = st.number_input(
        "VO2max überschreiben (ml/min)", 
        value=0, 
        step=50, 
        help="Ermöglicht das manuelle Überschreiben des berechneten VO2max-Wertes (z.B. 3652 für Barbara Bauer)"
    )
    
    if uploaded_file is not None:
        st.caption("Endzeitpunkt (in Sekunden) eingeben:")
        unregel = st.checkbox("Es gab während des Tests Unregelmäßigkeiten in der Stufendauer", value=False, key="unregel_stufen_widget")
        
        anz = int(st.session_state.anzahl)
        stage_sec = int(st.session_state.stufendauer * 60)
        pause_sec = int(st.session_state.pausendauer)
        vorlauf = int(st.session_state.vorlauf)
        interval = stage_sec + pause_sec
        
        for i in range(anz):
            if "detected_stage_ends" in st.session_state and i < len(st.session_state.detected_stage_ends):
                end_t_base = int(st.session_state.detected_stage_ends[i])
            else:
                end_t_base = int(vorlauf + (i + 1) * stage_sec + i * pause_sec)
            
            end_t_key = f"lauftest_end_t_{i}"
            
            if i == 0:
                if end_t_key not in st.session_state:
                    st.session_state[end_t_key] = end_t_base
                new_end_t = st.number_input(
                    f"Stufe {i+1} Endzeit (s)",
                    min_value=0,
                    max_value=10000,
                    value=int(st.session_state[end_t_key]),
                    step=1,
                    key=f"lauftest_end_t_widget_{i}"
                )
                st.session_state[end_t_key] = new_end_t
            else:
                calculated_end_t = int(st.session_state["lauftest_end_t_0"]) + i * interval
                
                if unregel:
                    if end_t_key not in st.session_state:
                        st.session_state[end_t_key] = calculated_end_t
                    new_end_t = st.number_input(
                        f"Stufe {i+1} Endzeit (s)",
                        min_value=0,
                        max_value=10000,
                        value=int(st.session_state[end_t_key]),
                        step=1,
                        key=f"lauftest_end_t_widget_{i}"
                    )
                    st.session_state[end_t_key] = new_end_t
                else:
                    st.session_state[end_t_key] = calculated_end_t

# 6. Coaching Potential (collapsed expander)
with st.sidebar.expander("Coaching Potential", expanded=False):
    st.caption("Simuliere den Effekt von Anpassungen bei VO2max, VLamax oder ANS.")
    pot_vo2_sel = st.selectbox(
        "VO2max Potential",
        ["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"],
        index=["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"].index(st.session_state.get("pot_vo2_select_key_lauft", "Keine Änderung")),
        key="pot_vo2_select_key_lauft"
    )
    pot_vla_sel = st.selectbox(
        "VLamax Potential",
        ["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"],
        index=["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"].index(st.session_state.get("pot_vla_select_key_lauft", "Keine Änderung")),
        key="pot_vla_select_key_lauft"
    )
    pot_ans_sel = st.selectbox(
        "ANS Potential",
        ["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"],
        index=["Keine Änderung", "+2 %", "+5 %", "-2 %", "-5 %"].index(st.session_state.get("pot_ans_select_key_lauft", "Keine Änderung")),
        key="pot_ans_select_key_lauft"
    )
    
    if st.session_state.get("lauf_auswertung_gestartet", False):
        st.markdown(f"""
        **Simuliertes Coaching-Potential:**
        * **VO2max:** {st.session_state.get('pot_vo2max', 0.0):.1f} ml/min/kg
        * **VLamax:** {st.session_state.get('pot_vlamax', 0.0):.3f} mmol/L/s
        * **ANS:** {st.session_state.get('pot_ans_rel', 0.0):.2f} m/s ({round(st.session_state.get('pot_ans_rel', 0.0) * 3.6, 1)} km/h)
        * **Fatmax:** {st.session_state.get('pot_fatmax', 0.0):.2f} m/s ({round(st.session_state.get('pot_fatmax', 0.0) * 3.6, 1)} km/h)
        """)

st.sidebar.markdown("---")
start_button = st.sidebar.button("Auswertung starten", type="primary")
if start_button:
    st.session_state.df_lauftest_input = edited_df
    st.session_state.df_sf = edited_sf
    st.session_state.lauf_auswertung_gestartet = True
elif st.session_state.get("lauf_auswertung_gestartet", False):
    st.session_state.df_lauftest_input = edited_df
    st.session_state.df_sf = edited_sf

# --- HILFSFUNKTIONEN ---

def detect_spiro_stage_ends(df_spiro, anzahl):
    n_rows = len(df_spiro)
    if n_rows < 10:
        return []
    t_diffs = df_spiro['Time'].diff().dropna()
    dt = t_diffs.median() if len(t_diffs) > 0 else 10.0
    step_30s = max(1, int(round(30.0 / dt)))
    drops = []
    for i in range(n_rows - step_30s):
        t1 = df_spiro['Time'].iloc[i]
        vo2_1 = df_spiro['VO2'].iloc[i]
        vo2_2 = df_spiro['VO2'].iloc[i + step_30s]
        t2 = df_spiro['Time'].iloc[i + step_30s]
        if 20 <= (t2 - t1) <= 45:
            pct_drop = (vo2_1 - vo2_2) / vo2_1 if vo2_1 > 0 else 0.0
            drops.append((t1, pct_drop, vo2_1))
    candidate_drops = [d for d in drops if d[1] > 0.15 and d[0] > 120]
    candidate_drops.sort(key=lambda x: x[0])
    grouped_ends = []
    for d in candidate_drops:
        t, pct, val = d
        if not grouped_ends or t - grouped_ends[-1][0] > 90:
            grouped_ends.append(d)
        else:
            if pct > grouped_ends[-1][1]:
                grouped_ends[-1] = d
    detected_times = [g[0] for g in grouped_ends]
    detected_times.sort()
    return [float(x) for x in detected_times[:anzahl]]

def calculate_vlamax_for_v(v_thresh, rel_vo2max, speed_arr, vo2_steady_abs, vco2_steady_abs, weight):
    if not (rel_vo2max > 0 and v_thresh > 0 and len(vo2_steady_abs) > 0 and weight > 0): 
        return None, None, None
        
    vo2_rest = 3.5 * weight
    re_abs_m_s = []
    
    # Calculate VO2_CHO100 running economy for exercise stages
    for v, vo2, vco2 in zip(speed_arr, vo2_steady_abs, vco2_steady_abs):
        if v <= 0:
            continue
        vo2_working = max(0.0, vo2 - vo2_rest)
        if vo2 > 0 and vco2 > 0:
            rq = vco2 / vo2
        else:
            rq = 0.85
        cal_eq = (rq - 0.7) * 5.094 + 19.6
        vo2_cho100 = vo2_working * cal_eq / 21.1
        re_abs_m_s.append(vo2_cho100 / v)
        
    ex_speeds = [v for v in speed_arr if v > 0]
    if len(re_abs_m_s) == 0 or len(ex_speeds) == 0:
        return None, None, None
        
    # Discretize speed to 0.01 step using floor (with floating-point tolerance)
    # to match Excel's approximate match behavior in J25:K425 table
    v_discretized = np.floor(v_thresh * 100 + 1e-9) / 100
    re_abs_at_thresh = np.interp(v_discretized, ex_speeds, re_abs_m_s)
    
    # Calculate relative VO2 demand at threshold (adding B28 = 1.0)
    rel_vo2_demand = (v_thresh * re_abs_at_thresh / weight) + 1.0
    
    delta_vo2 = rel_vo2max - rel_vo2_demand
    if delta_vo2 <= 0 or rel_vo2_demand <= 0:
        return None, re_abs_at_thresh / weight, rel_vo2_demand
        
    ks1, ks2, la_eq, vol_dist = 0.0631, 1.3310, 0.02049, 0.4
    vla_ox_max = (rel_vo2_demand * la_eq) / vol_dist
    adp = np.sqrt((ks1 * rel_vo2_demand) / delta_vo2)
    adp_3 = adp ** 3
    term2 = 1 + (ks2 / adp_3)
    vla_sec = (vla_ox_max * term2) / 60
    return vla_sec, re_abs_at_thresh / weight, rel_vo2_demand

def calculate_laufokonomie_hycys(vo2_abs, vco2_abs, weight, speed_ms):
    """
    Laufökonomie nach HYCYS Excel-Methodik.
    Entspricht Spalte Y im 'Report Running' Sheet:
      Y = (Arbeits-VO2 bei 100% KH) / Geschwindigkeit [m/s] / Gewicht [kg]
    
    Formeln aus Excel:
      - Ruhe-VO2 = 3.5 * weight (B14-Formel)
      - Arbeits-VO2 = VO2_gesamt - Ruhe-VO2 (G-Spalte)
      - RQ-Korrekturfaktor (L12): cal_eq = (RQ - 0.7) * 5.094 + 19.6 [kcal/L O2]
      - VO2_CHO100 = Arbeits-VO2 * cal_eq / 21.1  (AB-Spalte / B19)
      - Laufökonomie = VO2_CHO100 / speed / weight
    """
    if weight <= 0 or speed_ms <= 0 or vo2_abs <= 0:
        return None
    # Ruhe-VO2 [ml/min] (B14 = 3.5 * Gewicht)
    vo2_rest = 3.5 * weight
    # Arbeits-VO2 [ml/min] (Spalte G = F - Ruhe)
    vo2_working = max(0.0, vo2_abs - vo2_rest)
    if vo2_working <= 0:
        return None
    # RQ (Spalte K = VCO2 / VO2)
    if vco2_abs > 0 and vo2_abs > 0:
        rq = vco2_abs / vo2_abs
    else:
        rq = 0.85  # Standardwert
    # Kalorisches Äquivalent bei gemessenem RQ [kcal/L O2] (L12-Formel)
    # L12 = ((K12 - B17) * 0.05094 * 100) + B18  mit B17=0.7, B18=19.6
    cal_eq = (rq - 0.7) * 5.094 + 19.6
    # VO2-Äquivalent bei 100% KH-Verbrennung [ml/min] (AB-Spalte)
    b19 = 21.1  # kcal/L O2 bei RQ=1 (reine KH)
    vo2_cho100 = vo2_working * cal_eq / b19
    # Laufökonomie [ml*s/(m*min*kg)] = F44/B7 im Excel
    return vo2_cho100 / speed_ms / weight

# --- PDF GENERATOR FUNKTION ---
def create_pdf(athlete_name, birthdate, test_date, test_type, weight, body_fat_pct, rel_vo2max, df_res, fig_laktat,
               coach, sportart, kategorie, height, speed, lactate, hr, vo2_steady_values, vo2_steady_abs, vco2_steady_abs, uploaded_file):
    pdf = FPDF(unit='pt')
    pdf.set_auto_page_break(auto=False)

    # Standard colors from the template
    c_teal = (44, 183, 185)       # #2cb7b9
    c_gold = (205, 182, 99)       # #cdb663
    c_dark_grey = (89, 90, 89)    # #595a59
    c_black = (0, 0, 0)
    c_white = (255, 255, 255)

    # Helper function to set fonts & colors and draw text
    def draw_text(x, y, text, size=8.3, font_style='', color=c_black, font_name='Arial'):
        pdf.set_font(font_name, font_style, size)
        pdf.set_text_color(*color)
        pdf.text(x, y + size * 0.85, text.encode('latin-1', 'replace').decode('latin-1'))

    def draw_text_centered(cx, y, text, size=8.3, font_style='', color=c_black):
        pdf.set_font('Arial', font_style, size)
        w = pdf.get_string_width(str(text).encode('latin-1', 'replace').decode('latin-1'))
        draw_text(cx - w / 2.0, y, text, size=size, font_style=font_style, color=color)

    # Helper function to draw colored rectangles
    def draw_rect(x, y, w, h, fill_color=None, stroke_color=None, stroke_width=0.5):
        if fill_color:
            pdf.set_fill_color(*fill_color)
        if stroke_color:
            pdf.set_draw_color(*stroke_color)
            pdf.set_line_width(stroke_width)
        style = 'F' if fill_color and not stroke_color else 'D' if stroke_color and not fill_color else 'FD' if fill_color and stroke_color else ''
        pdf.rect(x, y, w, h, style)

    def page_logo(pg_num):
        """Draw small logo top-right (pages 2-5)."""
        asset = f'pdf_assets/page{pg_num}_img1.png'
        if not os.path.exists(asset):
            asset = 'pdf_assets/page2_img1.png'
        if os.path.exists(asset):
            coords = {
                2: (396.23, 68.88, 109.29, 20.47),
                3: (399.91, 69.42, 108.63, 20.09),
                4: (408.48, 68.88, 109.23, 20.58),
                5: (399.14, 68.88, 108.80, 20.58)
            }
            x, y, w, h = coords.get(pg_num, (396.23, 68.88, 109.29, 20.47))
            pdf.image(asset, x=x, y=y, w=w, h=h)

    def format_pace_helper(v):
        if not isinstance(v, (int, float)) or v <= 0:
            return "-"
        sec = 1000.0 / v
        m = int(sec // 60)
        s = int(round(sec % 60))
        if s == 60:
            m += 1
            s = 0
        return f"{m:02d}:{s:02d} min/km"

    def format_time_helper(t_min):
        h = int(t_min // 60)
        m = int(t_min % 60)
        s = int(round((t_min * 60) % 60))
        if s == 60:
            m += 1
            s = 0
        if m == 60:
            h += 1
            m = 0
        return f"{h:02d}:{m:02d}:{s:02d}"

    temp_files = []

    # --- METABOLIC GRID CALCULATIONS FOR REPORT ---
    vla_val = 0.6
    ans_row = pd.DataFrame()
    try:
        if not df_res.empty:
            ans_row = df_res.iloc[[0]]
            vla_raw = ans_row.iloc[0].get('VLamax [mmol/l/s]', 0.6)
            if pd.notna(vla_raw) and vla_raw != 'N/A' and 'bitte' not in str(vla_raw).lower():
                vla_val = float(vla_raw)
    except:
        pass

    v_ans = 3.5
    if not ans_row.empty:
        try:
            v_ans = float(ans_row.iloc[0].get('v', 3.5))
        except:
            pass

    has_spiro = (uploaded_file is not None) and (len(vo2_steady_abs) > 0)
    
    # Defaults/Fallbacks
    fatmax_speed = None
    v_grid = None
    lac_clear_grid = None
    lac_prod_grid = None
    fatmax_pace_str = "-"
    fatmax_total_energy = "-"
    fatmax_fat_energy = "-"
    ans_total_energy = "-"
    ans_carb_consumption = "-"
    marathon_paces = {0: "-", 5: "-", 10: "-", 15: "-"}
    marathon_times = {0: "-", 5: "-", 10: "-", 15: "-"}
    marathon_deltas = {5: "-", 10: "-", 15: "-"}

    if has_spiro:
        try:
            glycogen_store = 250.0  # g
            carb_intake_rate = 45.0  # g/h
            vo2_rest = 3.5 * weight

            # 1. Compute AC for each stage
            ac_stages = []
            for v, vo2, vco2 in zip(speed, vo2_steady_abs, vco2_steady_abs):
                if v <= 0 or vo2 <= 0:
                    ac_stages.append(0.0)
                    continue
                net_vo2 = max(0.0, vo2 - vo2_rest)
                rq = vco2 / vo2 if vco2 > 0 else 0.85
                kal_equiv = (rq - 0.7) * 5.094 + 19.6
                ee_kJ_min = (net_vo2 / 1000.0) * kal_equiv
                vo2_100kh = ee_kJ_min / 21.1 * 1000.0
                ac = vo2_100kh / (v * 100.0)
                ac_stages.append(ac)

            ex_speeds = [v for v in speed if v > 0]
            
            if len(ex_speeds) >= 2:
                v_grid = np.arange(2.6, 8.61, 0.01)
                ac_grid = []
                for v in v_grid:
                    if v <= ex_speeds[0]:
                        m = (ac_stages[1] - ac_stages[0]) / (ex_speeds[1] - ex_speeds[0])
                        c_val = ac_stages[0] - m * ex_speeds[0]
                        ac = m * v + c_val
                    elif v >= ex_speeds[-1]:
                        m = (ac_stages[-1] - ac_stages[-2]) / (ex_speeds[-1] - ex_speeds[-2])
                        c_val = ac_stages[-1] - m * ex_speeds[-1]
                        ac = m * v + c_val
                    else:
                        for i in range(len(ex_speeds) - 1):
                            if ex_speeds[i] <= v < ex_speeds[i+1]:
                                m = (ac_stages[i+1] - ac_stages[i]) / (ex_speeds[i+1] - ex_speeds[i])
                                c_val = ac_stages[i] - m * ex_speeds[i]
                                ac = m * v + c_val
                                break
                    ac_grid.append(ac)

                ac_grid = np.array(ac_grid)
                vo2_rel_demand_grid = (ac_grid * v_grid * 100.0) / weight + 1.0
                delta_vo2_grid = rel_vo2max - vo2_rel_demand_grid

                lac_prod_grid = np.zeros_like(v_grid)
                lac_clear_grid = (0.02049 * vo2_rel_demand_grid) / 0.4
                fat_pct_grid = np.zeros_like(v_grid)
                carb_pct_grid = np.zeros_like(v_grid)
                net_clearance = np.full_like(v_grid, -9999.0)

                ks1, ks2, la_eq, vol_dist = 0.0631, 1.331, 0.02049, 0.4

                for idx, (v, vo2_rel, d_vo2) in enumerate(zip(v_grid, vo2_rel_demand_grid, delta_vo2_grid)):
                    if d_vo2 > 0 and vo2_rel > 0:
                        adp = np.sqrt((ks1 * vo2_rel) / d_vo2)
                        adp_3 = adp ** 3
                        pfk = 1 + (ks2 / adp_3)
                        vla_min = vla_val * 60 / pfk  # mmol/l/min
                        lac_prod_grid[idx] = vla_min
                        
                        fat_pct = 100.0 - (vla_min / lac_clear_grid[idx] * 100.0)
                        fat_pct = max(0.0, min(100.0, fat_pct))
                        net_clearance[idx] = lac_clear_grid[idx] - vla_min
                    else:
                        fat_pct = 0.0
                        net_clearance[idx] = -9999.0
                    
                    fat_pct_grid[idx] = fat_pct
                    carb_pct_grid[idx] = 100.0 - fat_pct

                max_clearance_idx = np.argmax(net_clearance)
                fatmax_speed = v_grid[max_clearance_idx]
                fatmax_pace_str = format_pace_helper(fatmax_speed)

                rq_grid = 0.7 + 0.3 * (carb_pct_grid / 100.0)
                kal_equiv_grid = (rq_grid - 0.7) * 5.094 + 19.6
                ee_kJ_min_grid = (vo2_rel_demand_grid * weight / 1000.0) * kal_equiv_grid
                ee_kJ_h_grid = ee_kJ_min_grid * 60
                ee_kcal_h_grid = ee_kJ_h_grid / 4.186
                ee_fat_kcal_h_grid = ee_kcal_h_grid * (fat_pct_grid / 100.0)
                ee_kh_kcal_h_grid = ee_kcal_h_grid * (carb_pct_grid / 100.0)
                carb_g_h_grid = ee_kh_kcal_h_grid / 4.063

                fatmax_idx = max_clearance_idx
                fatmax_total_energy = f"{ee_kcal_h_grid[fatmax_idx]:.1f}".replace('.', ',')
                fatmax_fat_energy = f"{ee_fat_kcal_h_grid[fatmax_idx]:.1f}".replace('.', ',')

                ans_speed_lookup = np.floor(v_ans * 100 + 1e-9) / 100.0
                ans_idx = np.argmin(np.abs(v_grid - ans_speed_lookup))
                ans_total_energy = f"{ee_kcal_h_grid[ans_idx]:.1f}".replace('.', ',')
                ans_carb_consumption = f"{carb_g_h_grid[ans_idx]:.1f}".replace('.', ',')

                def calc_marathon_speed_0():
                    for idx, v in enumerate(v_grid):
                        t_mara_min = 42195.0 / (v * 60.0)
                        carb_req = ee_kh_kcal_h_grid[idx] / 60.0 * t_mara_min
                        carb_avail = (carb_intake_rate / 60.0 * t_mara_min * 4.063) + (glycogen_store * 4.063)
                        deficit = carb_req - carb_avail
                        if deficit > 0:
                            return v_grid[idx - 1]
                    return v_grid[-1]

                v_mara_0 = calc_marathon_speed_0()
                v_mara_5 = v_mara_0 / 0.95
                v_mara_10 = v_mara_0 / 0.90
                v_mara_15 = v_mara_0 / 0.85

                for name, v_m in [("0%", v_mara_0), ("5%", v_mara_5), ("10%", v_mara_10), ("15%", v_mara_15)]:
                    t_min = 42195.0 / (v_m * 60.0)
                    h = int(t_min // 60)
                    m = int(t_min % 60)
                    s = int(round((t_min * 60) % 60))
                    if s == 60:
                        m += 1
                        s = 0
                    if m == 60:
                        h += 1
                        m = 0
                    
                    pct_key = int(name.replace('%', ''))
                    marathon_paces[pct_key] = format_pace_helper(v_m)
                    marathon_times[pct_key] = f"{h:02d}:{m:02d}:{s:02d}"

                t_min_0 = 42195.0 / (v_mara_0 * 60.0)
                for pct_key in [5, 10, 15]:
                    t_min_p = 42195.0 / ((v_mara_0 / (1.0 - pct_key / 100.0)) * 60.0)
                    delta_min = t_min_0 - t_min_p
                    dm = int(delta_min)
                    ds = int(round((delta_min - dm) * 60.0))
                    if ds == 60:
                        dm += 1
                        ds = 0
                    marathon_deltas[pct_key] = f"-{dm:02d}:{ds:02d} min"
        except Exception as e:
            pass

    # --- SEITE 1: DECKBLATT ---
    pdf.add_page()
    draw_rect(50.9, 54.5, 479.1, 663.6, fill_color=c_white)
    
    if os.path.exists('pdf_assets/page1_img3.png'):
        pdf.image('pdf_assets/page1_img3.png', x=86.8, y=194.4, w=400, h=124)
        
    draw_rect(50.9, 460.4, 479.1, 24.5, fill_color=c_gold)
    draw_text(111.7, 464.7, f"HYCYS {test_type}", size=13.8, font_style='B', color=c_white)

    labels_s1 = ["Name", "Geburtsdatum", "Coach", "Testdatum", "Sportart", "Kategorie"]
    vals_s1 = [athlete_name, birthdate, coach, test_date, sportart, kategorie]
    y_coords_s1 = [490.7, 502.7, 542.6, 555.9, 582.6, 595.9]
    for lbl, val, y_coord in zip(labels_s1, vals_s1, y_coords_s1):
        draw_text(111.0, y_coord, lbl, size=8.3, color=c_black)
        draw_text(214.5, y_coord, val, size=8.3, font_style='B', color=c_black)

    if os.path.exists('pdf_assets/page1_img1.png'):
        pdf.image('pdf_assets/page1_img1.png', x=114.4, y=681.4, w=15.6, h=14.4)
    draw_text(162.6, 685.8, "www.hycys.de", size=7.6, font_style='B', color=c_teal)
    
    if os.path.exists('pdf_assets/page1_img2.png'):
        pdf.image('pdf_assets/page1_img2.png', x=137.3, y=680.6, w=15.6, h=16.1)
    draw_text(266.1, 685.8, "contact@hycys.de", size=7.6, font_style='B', color=c_teal)

    # --- SEITE 2: ANTHROPOMETRIE & LAKTATKINETIK ---
    pdf.add_page()
    draw_rect(50.9, 54.5, 471.6, 663.6, fill_color=c_white)
    
    page_logo(2)

    draw_rect(57.6, 133.9, 171.9, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 171.9 / 2.0, 135.0, "Anthropologie", size=8.2, font_style='B', color=c_white)

    draw_text(59.3, 173.2, "Gewicht", size=8.3, font_style='B', color=c_black)
    draw_text(199.9, 173.2, f"{weight:.1f} kg".replace('.', ','), size=8.3, font_style='B', color=c_dark_grey)
    draw_text(59.3, 183.8, "Größe", size=8.3, font_style='B', color=c_black)
    draw_text(199.9, 183.8, f"{int(height)}cm", size=8.3, font_style='B', color=c_dark_grey)
    
    bmi = weight / ((height/100.0)**2) if height > 0 else 0.0
    draw_text(59.3, 209.7, "Body Mass Index", size=8.3, font_style='B', color=c_black)
    draw_text(187.5, 209.7, f"{bmi:.1f} kg/m²".replace('.', ','), size=8.3, font_style='B', color=c_dark_grey)
    
    draw_rect(113.18, 245.58, 112.94, 10.44, fill_color=c_gold)
    draw_text_centered(141.42, 246.8, "Masse", size=8.2, font_style='B', color=c_white)
    draw_text_centered(197.89, 246.8, "%",     size=8.2, font_style='B', color=c_white)
    
    fett_kg = weight * (body_fat_pct / 100.0)
    fettfrei_kg = weight - fett_kg
    draw_text(59.3, 259.6, "Fett", size=8.3, font_style='B', color=c_black)
    draw_text_centered(141.42, 259.6, f"{fett_kg:.1f} kg".replace('.', ','), size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(197.89, 259.6, f"{body_fat_pct:.1f} %".replace('.', ','), size=8.3, font_style='B', color=c_dark_grey)
    
    draw_text(59.3, 270.1, "Fettfrei", size=8.3, font_style='B', color=c_black)
    draw_text_centered(141.42, 270.1, f"{fettfrei_kg:.1f} kg".replace('.', ','), size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(197.89, 270.1, f"{100.0 - body_fat_pct:.1f} %".replace('.', ','), size=8.3, font_style='B', color=c_dark_grey)

    # Exploded Pie Chart matching Cycling Report
    fig, ax = plt.subplots(figsize=(3.856, 2.0), dpi=300)
    pct_fett = max(0.1, body_fat_pct)
    pct_frei = max(0.1, 100.0 - body_fat_pct)
    wedges, texts, autotexts = ax.pie(
        [pct_fett, pct_frei],
        explode=(0.12, 0),
        colors=['#2cb7b9', '#cdb663'],
        autopct=lambda p: f"{p:.1f} %".replace('.', ','),
        startangle=90,
        counterclock=False,
        pctdistance=0.50,
        textprops=dict(size=8, weight='bold', color='white'),
        wedgeprops=dict(edgecolor='white', linewidth=2),
    )
    
    for i, (wedge, autotext) in enumerate(zip(wedges, autotexts)):
        theta = np.deg2rad(wedge.theta1 + (wedge.theta2 - wedge.theta1) / 2.0)
        e = 0.12 if i == 0 else 0.0
        d = 0.65 if i == 0 else 0.50
        x = (e + d) * np.cos(theta)
        y = (e + d) * np.sin(theta)
        autotext.set_position((x, y))
        autotext.set_color('white')

    ax.axis('equal')
    from matplotlib.patches import Patch
    handles = [Patch(color='#2cb7b9', label='Fett'), Patch(color='#cdb663', label='Fettfrei')]
    legend = ax.legend(handles=handles, loc='center left', bbox_to_anchor=(1.05, 0.5), fontsize=8, frameon=False)
    for text in legend.get_texts():
        text.set_weight('bold')
        text.set_color('#595a59')
    fig.patch.set_facecolor('white')
    plt.tight_layout(pad=0.3)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_pie:
        fig.savefig(tmp_pie.name, format="png", bbox_inches="tight", facecolor='white')
        plt.close(fig)
        pdf.image(tmp_pie.name, x=268.05, y=134.92, w=227.56, h=118.01)
        temp_files.append(tmp_pie.name)

    # VLamax Section
    draw_rect(57.6, 293.2, 225.8, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 225.8 / 2.0, 294.3, "Anaerober Stoffwechsel - max. Laktatbildungsrate", size=8.2, font_style='B', color=c_white)
    
    draw_rect(114.02, 314.9, 112.94, 10.44, fill_color=c_gold)
    draw_text_centered(114.02 + 112.94 / 2.0, 316.1, "VLamax", size=8.2, font_style='B', color=c_white)
    
    vla_ans_str = f"{vla_val:.2f} mmol/l/s".replace('.', ',') if vla_val > 0 else "-"
    draw_text_centered(114.02 + 112.94 / 2.0, 314.9 + 10.44 + 0.8, vla_ans_str, size=8.2, font_style='B', color=c_dark_grey)

    # VO2max Section
    draw_rect(57.6, 355.0, 343.6, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 343.6 / 2.0, 356.1, "Aerober Stoffwechsel - max. Sauerstoffaufnahme", size=8.2, font_style='B', color=c_white)
    
    vo2_abs = rel_vo2max * weight
    vo2_abs_str = f"{int(vo2_abs):,}".replace(",", ".") + " ml/min" if vo2_abs > 0 else "-"
    vo2_rel_str = f"{rel_vo2max:.1f} ml/min/kg".replace('.', ',') if rel_vo2max > 0 else "-"
    
    ffm_pct = 100.0 - body_fat_pct
    ffm_kg = weight * (ffm_pct / 100.0)
    vo2_ffm = (rel_vo2max * weight) / ffm_kg if ffm_kg > 0 else 0.0
    vo2_ffm_str = f"{vo2_ffm:.1f} ml/min/kg".replace('.', ',') if vo2_ffm > 0 else "-"

    draw_rect(57.6, 375.0, 143.6, 10.44, fill_color=c_gold)
    draw_text_centered(57.6 + 143.6 / 2.0, 376.2, "VO2max abs.", size=8.2, font_style='B', color=c_white)
    draw_text_centered(57.6 + 143.6 / 2.0, 375.0 + 10.44 + 0.8, vo2_abs_str, size=8.2, font_style='B', color=c_dark_grey)
    
    draw_rect(201.2, 375.0, 143.6, 10.44, fill_color=c_gold)
    draw_text_centered(201.2 + 143.6 / 2.0, 376.2, "VO2max rel.", size=8.2, font_style='B', color=c_white)
    draw_text_centered(201.2 + 143.6 / 2.0, 375.0 + 10.44 + 0.8, vo2_rel_str, size=8.2, font_style='B', color=c_dark_grey)
    
    draw_rect(344.8, 375.0, 143.4, 10.44, fill_color=c_gold)
    draw_text_centered(344.8 + 143.4 / 2.0, 376.2, "VO2max rel. FFM", size=8.2, font_style='B', color=c_white)
    draw_text_centered(344.8 + 143.4 / 2.0, 375.0 + 10.44 + 0.8, vo2_ffm_str, size=8.2, font_style='B', color=c_dark_grey)

    # Lactate kinetics plot
    draw_rect(57.6, 415.0, 458.1, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 458.1 / 2.0, 416.1, "Interaktion Laktatauf- & abbau", size=8.2, font_style='B', color=c_white)
    
    fig, ax = plt.subplots(figsize=(6, 3.2), dpi=300)
    v_start = speed[0] if len(speed) > 0 else 2.9
    v_end = speed[-1] if len(speed) > 0 else 4.5
    v_arr = np.linspace(v_start, v_end, 100)
    
    if uploaded_file is not None and len(vo2_steady_abs) > 0:
        ax.plot(speed[:len(vo2_steady_abs)], vo2_steady_abs, color='#595a59', linewidth=2.5)
        ax.set_ylabel('VO2 [ml/min]', color='#595a59', fontsize=7, fontweight='bold')
        ax.tick_params(axis='y', labelcolor='#595a59', labelsize=6, width=1.2)
        ax.grid(True, which='both', linestyle='-', linewidth=1.0, color='#d8d8d8')
        ax.set_axisbelow(True)
    else:
        ax.set_ylabel('VO2 [ml/min]', color='#595a59', fontsize=7, fontweight='bold')
        ax.tick_params(axis='y', labelcolor='#595a59', labelsize=6, width=1.2)
        ax.grid(True, which='both', linestyle='-', linewidth=1.0, color='#d8d8d8')
        ax.set_axisbelow(True)
        
    ax2 = ax.twinx()
    if v_grid is not None and lac_clear_grid is not None and lac_prod_grid is not None:
        ax2.plot(v_grid, lac_clear_grid, color='#2cb7b9', linewidth=2.5)
        ax2.plot(v_grid, lac_prod_grid, color='#cdb663', linewidth=2.5, linestyle='-')
    else:
        poly_coeffs = np.polyfit(speed, lactate, 3)
        poly_func = np.poly1d(poly_coeffs)
        ax2.plot(v_arr, poly_func(v_arr), color='#2cb7b9', linewidth=2.5)
        
        lac_prod = lactate[0] + (lactate[-1] - lactate[0]) * ((v_arr - v_start)/(v_end - v_start))**3.5
        ax2.plot(v_arr, lac_prod, color='#cdb663', linewidth=2.5, linestyle='-')
    
    ax2.set_ylabel('Laktatkinetik [mmol/L/min]', color='#595a59', fontsize=7, fontweight='bold')
    ax2.tick_params(axis='y', labelcolor='#595a59', labelsize=6, width=1.2)
    
    ax.set_xlabel('Geschwindigkeit [m/s]', color='#595a59', fontsize=7, fontweight='bold')
    ax.set_xlim(v_start, v_end)
    ax.tick_params(axis='x', labelcolor='#595a59', labelsize=6, width=1.2)
    
    for s in ax.spines.values():
        s.set_color('#595a59')
        s.set_linewidth(1.2)
    for s in ax2.spines.values():
        s.set_color('#595a59')
        s.set_linewidth(1.2)

    for label in ax.get_xticklabels(): label.set_weight('bold')
    for label in ax.get_yticklabels(): label.set_weight('bold')
    for label in ax2.get_yticklabels(): label.set_weight('bold')

    bbox_props = dict(boxstyle='square,pad=0.2', facecolor='white', edgecolor='none', alpha=1.0)
    ax.text(0.24, 0.43, 'Sauerstoffaufnahme - VO2', color='#595a59', fontsize=6.5, fontweight='bold', ha='center', va='center', transform=ax.transAxes, bbox=bbox_props, zorder=10)
    ax2.text(0.66, 0.81, 'Laktatproduktion', color='#cdb663', fontsize=6.5, fontweight='bold', ha='center', va='center', transform=ax2.transAxes, bbox=bbox_props, zorder=10)
    ax2.text(0.76, 0.33, 'Laktatabbau', color='#2cb7b9', fontsize=6.5, fontweight='bold', ha='center', va='center', transform=ax2.transAxes, bbox=bbox_props, zorder=10)
    
    plt.tight_layout()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_lactate:
        fig.savefig(tmp_lactate.name, format="png", bbox_inches="tight")
        plt.close(fig)
        pdf.image(tmp_lactate.name, x=55.2, y=432.0, w=455, h=248)
        temp_files.append(tmp_lactate.name)

    # --- SEITE 3: LAUFÖKONOMIE ---
    pdf.add_page()
    draw_rect(50.9, 54.5, 479.1, 663.6, fill_color=c_white)
    
    page_logo(3)
        
    draw_rect(57.6, 133.9, 207.0, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 207.0 / 2.0, 135.0, "Laufökonomie", size=8.2, font_style='B', color=c_white)

    # Table Header
    draw_rect(57.6, 159.1, 207.0, 10.44, fill_color=c_gold)
    draw_text_centered(57.6 + 103.5 / 2.0, 160.3, "Geschwindigkeit", size=8.2, font_style='B', color=c_white)
    draw_text_centered(57.6 + 103.5 + 103.5 / 2.0, 160.3, "Laufökonomie", size=8.2, font_style='B', color=c_white)

    # Populate running economy values dynamically
    speeds_s3 = []
    eco_vals_s3 = []
    y_coords_s3 = [173.2, 183.8, 197.1, 211.8, 225.1]
    
    display_steps = min(5, len(speed))
    for i in range(display_steps):
        v_ms = speed[i]
        vo2_r = vo2_steady_values[i] if i < len(vo2_steady_values) else 0.0
        
        if vo2_r > 0:
            re_val = vo2_r / v_ms
            eco_str = f"{re_val:.1f} ml/min/kg*m/s".replace('.', ',')
        else:
            eco_str = "-"
            
        speeds_s3.append(f"{v_ms:.2f} m/s".replace('.', ','))
        eco_vals_s3.append(eco_str)
        
    for i in range(len(speeds_s3)):
        y_coord = y_coords_s3[i]
        draw_text_centered(57.6 + 103.5 / 2.0, y_coord, speeds_s3[i], size=8.3, font_style='B', color=c_black)
        draw_text_centered(57.6 + 103.5 + 103.5 / 2.0, y_coord, eco_vals_s3[i], size=8.3, font_style='B', color=c_black)

    # Chart 3a: Running Economy plot
    fig, ax = plt.subplots(figsize=(3, 1.8), dpi=300)
    valid_re = []
    valid_speeds_v = []
    for i in range(len(speed)):
        v_ms = speed[i]
        vo2_r = vo2_steady_values[i] if i < len(vo2_steady_values) else 0.0
        if vo2_r > 0:
            re_v = vo2_r / v_ms
            valid_re.append(re_v)
            valid_speeds_v.append(v_ms)
            
    if len(valid_re) > 0:
        bar_w = (valid_speeds_v[-1] - valid_speeds_v[0]) / max(len(valid_speeds_v), 1) * 0.85
        ax.bar(valid_speeds_v, valid_re, width=bar_w, color='#2cb7b9', zorder=3)
        ax.set_ylim(min(valid_re)*0.9, max(valid_re)*1.1)
        
    ax.set_ylabel('RE [ml/min/kg*m/s]', color='#595a59', fontsize=6, fontweight='bold')
    ax.set_xlabel('Geschwindigkeit [m/s]', color='#595a59', fontsize=6, fontweight='bold')
    if len(speed) > 0:
        ax.set_xlim(speed[0]*0.95, speed[-1]*1.05)
    ax.tick_params(axis='both', labelsize=5, width=1.2)
    ax.grid(True, axis='y', linestyle='-', linewidth=1.0, color='#d8d8d8')
    ax.set_axisbelow(True)
    for s in ax.spines.values():
        s.set_color('#595a59')
        s.set_linewidth(1.2)
    plt.tight_layout()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_eco_chart:
        fig.savefig(tmp_eco_chart.name, format="png", bbox_inches="tight")
        plt.close(fig)
        pdf.image(tmp_eco_chart.name, x=266.8, y=156.2, w=257, h=135.6)
        temp_files.append(tmp_eco_chart.name)

    # Marathon predictions table populating calculated values
    draw_rect(57.6, 334.1, 414.0, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 414.0 / 2.0, 335.2, "Einfluss Laufökonomie auf Marathon-Zeit", size=8.2, font_style='B', color=c_white)

    draw_rect(57.6, 355.1, 414.0, 10.44, fill_color=c_gold)
    draw_text_centered(109.35, 356.2, "Verbesserung", size=8.2, font_style='B', color=c_white)
    draw_text_centered(212.85, 356.2, "Marathon-Pace", size=8.2, font_style='B', color=c_white)
    draw_text_centered(316.35, 356.2, "Zielzeit", size=8.2, font_style='B', color=c_white)
    draw_text_centered(419.85, 356.2, "Delta zu 0%", size=8.2, font_style='B', color=c_white)

    y_coords_s3_m = [386.0, 398.9, 412.3]
    improves = ["0%", "5%", "10%"]
    for imp, y_coord in zip(improves, y_coords_s3_m):
        pct_key = int(imp.replace('%', ''))
        pace_val = marathon_paces[pct_key]
        time_val = marathon_times[pct_key]
        delta_val = marathon_deltas[pct_key] if pct_key != 0 else "-"
        
        draw_text_centered(109.35, y_coord, imp, size=9.0, font_style='B', color=c_dark_grey)
        draw_text_centered(212.85, y_coord, pace_val, size=9.0, font_style='B', color=c_dark_grey)
        draw_text_centered(316.35, y_coord, time_val, size=9.0, font_style='B', color=c_dark_grey)
        if pct_key != 0:
            draw_text_centered(419.85, y_coord, delta_val, size=9.0, font_style='B', color=c_dark_grey)

    # Marathon prediction chart with line
    fig, ax = plt.subplots(figsize=(5, 2.5), dpi=300)
    
    if has_spiro and fatmax_speed is not None:
        y_vals = [0, 5, 10]
        t_0_h = 42195.0 / (v_mara_0 * 3600.0)
        t_5_h = 42195.0 / (v_mara_5 * 3600.0)
        t_10_h = 42195.0 / (v_mara_10 * 3600.0)
        x_vals = [t_0_h, t_5_h, t_10_h]
        
        x_min = min(x_vals) - 0.2
        x_max = max(x_vals) + 0.1
        
        bar_height = 2.5
        widths = [x - x_min for x in x_vals]
        ax.barh(y_vals, widths, height=bar_height, left=x_min, color='#2cb7b9')
        
        ax.set_ylim(-3, 13)
        ax.set_yticks([0, 5, 10])
        ax.set_yticklabels(['0%', '5%', '10%'])
        
        ax.set_xlim(x_min, x_max)
        
        import matplotlib.ticker as ticker
        def format_hours(x, pos):
            h = int(x)
            m = int(round((x - h) * 60))
            if m == 60:
                h += 1
                m = 0
            return f"{h}:{m:02d}"
        ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_hours))
        
        for label in ax.get_xticklabels(): label.set_weight('bold'); label.set_size(5)
        for label in ax.get_yticklabels(): label.set_weight('bold'); label.set_size(5)
    else:
        ax.set_xlim(2.0, 3.5)
        ax.set_ylim(-3, 13)
        ax.set_yticks([0, 5, 10])
        ax.set_yticklabels(['0%', '5%', '10%'])
        for label in ax.get_xticklabels(): label.set_weight('bold'); label.set_size(5)
        for label in ax.get_yticklabels(): label.set_weight('bold'); label.set_size(5)

    ax.set_xlabel('Marathon-Zielzeit [h:mm]', color='#595a59', fontsize=6, fontweight='bold')
    ax.set_ylabel('Verbesserung Laufökonomie [%]', color='#595a59', fontsize=6, fontweight='bold')
    ax.tick_params(axis='both', labelcolor='#595a59', width=1.2)
    for s in ax.spines.values(): s.set_color('#595a59'); s.set_linewidth(1.2)
    ax.grid(True, which='both', linestyle='-', linewidth=0.5, color='#d8d8d8')
    ax.set_axisbelow(True)
    plt.tight_layout()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_mara_chart:
        fig.savefig(tmp_mara_chart.name, format="png", bbox_inches="tight")
        plt.close(fig)
        pdf.image(tmp_mara_chart.name, x=66.1, y=475.1, w=449.2, h=209.4)
        temp_files.append(tmp_mara_chart.name)

    # --- SEITE 4: ANAEROBE SCHWELLE & ENERGETIK ---
    pdf.add_page()
    draw_rect(50.9, 54.5, 479.1, 663.6, fill_color=c_white)
    
    page_logo(4)
        
    draw_rect(57.6, 133.9, 465.7, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 465.7 / 2.0, 135.0, "Anaerobe Schwelle (ANS) - Geschwindigkeit & Herzfrequenz", size=8.2, font_style='B', color=c_white)

    # Box 1
    draw_rect(57.6, 159.1, 207.0, 10.44, fill_color=c_gold)
    draw_text_centered(57.6 + 207.0 / 2.0, 160.3, "Laufgeschwindigkeit", size=8.2, font_style='B', color=c_white)
    
    draw_rect(368.0, 159.1, 103.6, 10.44, fill_color=c_gold)
    draw_text_centered(368.0 + 103.6 / 2.0, 160.3, "Herzfrequenz", size=8.2, font_style='B', color=c_white)

    ans_ms_str, ans_pace_str, ans_hf_str = "-", "-", "-"
    v_ans_val = 0.0
    if not ans_row.empty:
        raw_v = ans_row.iloc[0].get('v', 0.0)
        try:
            v_ans_val = float(raw_v)
        except (ValueError, TypeError):
            v_ans_val = 0.0
        if v_ans_val > 0:
            ans_ms_str = f"{v_ans_val:.2f} m/s".replace('.', ',')
            ans_pace_str = format_pace_helper(v_ans_val)
            try:
                raw_hf = ans_row.iloc[0].get('HF', 0)
                ans_hf_str = f"{int(float(raw_hf))} bpm"
            except (ValueError, TypeError):
                ans_hf_str = "-"

    draw_text_centered(109.35, 170.34, ans_ms_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(212.85, 170.34, ans_pace_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(368.0 + 103.6 / 2.0, 170.34, ans_hf_str, size=8.3, font_style='B', color=c_dark_grey)
 
    # Box 2: VO2
    draw_rect(57.6, 206.9, 465.7, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 465.7 / 2.0, 208.0, "Anaerobe Schwelle (ANS) - VO2", size=8.2, font_style='B', color=c_white)
 
    draw_rect(57.6, 234.9, 103.6, 10.44, fill_color=c_gold)
    draw_text_centered(57.6 + 103.6 / 2.0, 236.1, "VO2 abs.", size=8.2, font_style='B', color=c_white)
 
    draw_rect(212.8, 234.9, 103.6, 10.44, fill_color=c_gold)
    draw_text_centered(212.8 + 103.6 / 2.0, 236.1, "VO2 rel.", size=8.2, font_style='B', color=c_white)
 
    draw_rect(368.0, 234.9, 103.6, 10.44, fill_color=c_gold)
    draw_text_centered(368.0 + 103.6 / 2.0, 236.1, "% VO2max", size=8.2, font_style='B', color=c_white)
 
    ans_vo2_abs_str, ans_vo2_rel_str, ans_vo2_pct_str = "-", "-", "-"
    if uploaded_file is not None and v_ans_val > 0 and len(vo2_steady_values) > 0:
        ans_vo2_rel = np.interp(v_ans_val, speed, vo2_steady_values)
        ans_vo2_abs = ans_vo2_rel * weight
        ans_vo2_pct = (ans_vo2_rel / rel_vo2max * 100.0) if rel_vo2max > 0 else 0.0
        
        ans_vo2_abs_str = f"{int(ans_vo2_abs):,}".replace(",", ".") + " ml/min"
        ans_vo2_rel_str = f"{ans_vo2_rel:.1f} ml/min/kg".replace('.', ',')
        ans_vo2_pct_str = f"{int(ans_vo2_pct)} %"
 
    draw_text_centered(57.6 + 103.6 / 2.0, 246.14, ans_vo2_abs_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(212.8 + 103.6 / 2.0, 246.14, ans_vo2_rel_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(368.0 + 103.6 / 2.0, 246.14, ans_vo2_pct_str, size=8.3, font_style='B', color=c_dark_grey)
 
    # Energetik Box with calculated total energy & KH-Verbrauch
    draw_rect(57.6, 279.9, 465.7, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 465.7 / 2.0, 281.0, "Anaerobe Schwelle (ANS) - Energetik", size=8.2, font_style='B', color=c_white)
 
    draw_rect(57.6, 307.8, 103.6, 10.44, fill_color=c_gold)
    draw_text_centered(57.6 + 103.6 / 2.0, 309.0, "Energieverbrauch", size=8.2, font_style='B', color=c_white)
 
    draw_rect(212.8, 307.8, 103.6, 10.44, fill_color=c_gold)
    draw_text_centered(212.8 + 103.6 / 2.0, 309.0, "KH-Verbrauch", size=8.2, font_style='B', color=c_white)
 
    ans_total_energy_str = f"{ans_total_energy} kcal/h" if ans_total_energy != "-" else "-"
    ans_carb_consumption_str = f"{ans_carb_consumption} g/h" if ans_carb_consumption != "-" else "-"
    draw_text_centered(57.6 + 103.6 / 2.0, 319.04, ans_total_energy_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(212.8 + 103.6 / 2.0, 319.04, ans_carb_consumption_str, size=8.3, font_style='B', color=c_dark_grey)
 
    # Fatmax Box with calculated values
    draw_rect(57.6, 355.1, 465.7, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 465.7 / 2.0, 356.2, "Fettstoffwechsel - Fatmax", size=8.2, font_style='B', color=c_white)
 
    draw_rect(57.6, 383.1, 207.0, 10.44, fill_color=c_gold)
    draw_text_centered(57.6 + 207.0 / 2.0, 384.3, "Laufgeschwindigkeit", size=8.2, font_style='B', color=c_white)
 
    draw_rect(316.3, 383.1, 207.0, 10.44, fill_color=c_gold)
    draw_text_centered(316.3 + 207.0 / 4.0, 384.3, "Energie", size=8.2, font_style='B', color=c_white)
    draw_text_centered(316.3 + 3 * 207.0 / 4.0, 384.3, "Energie Fett", size=8.2, font_style='B', color=c_white)
 
    fatmax_speed_str = f"{fatmax_speed:.2f} m/s".replace('.', ',') if fatmax_speed is not None else "-"
    fatmax_total_energy_str = f"{fatmax_total_energy} kcal/h" if fatmax_total_energy != "-" else "-"
    fatmax_fat_energy_str = f"{fatmax_fat_energy} kcal/h" if fatmax_fat_energy != "-" else "-"
    
    draw_text_centered(109.35, 394.34, fatmax_speed_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(212.85, 394.34, fatmax_pace_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(316.3 + 207.0 / 4.0, 394.34, fatmax_total_energy_str, size=8.3, font_style='B', color=c_dark_grey)
    draw_text_centered(316.3 + 3 * 207.0 / 4.0, 394.34, fatmax_fat_energy_str, size=8.3, font_style='B', color=c_dark_grey)

    # Energetik chart with dual-Y axis and plotted curves
    fig, ax = plt.subplots(figsize=(6, 3.2), dpi=300)
    ax.set_xlabel('Geschwindigkeit [m/s]', color='#595a59', fontsize=6, fontweight='bold')
    ax.set_ylabel('Energie [kcal/h]', color='#595a59', fontsize=6, fontweight='bold')
    
    v_start = speed[0]
    v_end = speed[-1]
    x_lo = v_start * 0.95
    x_hi = v_end * 1.05
    ax.set_xlim(x_lo, x_hi)
    
    ax2 = ax.twinx()
    ax2.set_ylabel('Kohlenhydratverbrauch [g/h]', color='#595a59', fontsize=6, fontweight='bold')
    
    if has_spiro and fatmax_speed is not None:
        ax.plot(v_grid, ee_kcal_h_grid, color='#595a59', linewidth=2.5)
        ax.plot(v_grid, ee_fat_kcal_h_grid, color='#2cb7b9', linewidth=2.5)
        ax2.plot(v_grid, carb_g_h_grid, color='#cdb663', linewidth=2.5, linestyle='-')
        
        if fatmax_speed > 0 and x_lo <= fatmax_speed <= x_hi:
            ax.axvspan(fatmax_speed - 0.1, fatmax_speed + 0.1, color='#2cb7b9', alpha=0.25)
        
        # Compute Y-axis limits based on the VISIBLE speed range only
        import math
        vis_mask = (v_grid >= x_lo) & (v_grid <= x_hi)
        vis_ee = ee_kcal_h_grid[vis_mask]
        vis_fat = ee_fat_kcal_h_grid[vis_mask]
        vis_carb = carb_g_h_grid[vis_mask]
        
        y1_data_max = max(vis_ee.max(), vis_fat.max())
        y2_data_max = vis_carb.max()
        
        # Round up to nice tick values with ~10% headroom
        y1_step = 100 if y1_data_max < 1500 else 200 if y1_data_max < 3000 else 500
        y1_max = math.ceil(y1_data_max * 1.12 / y1_step) * y1_step
        
        y2_step = 50 if y2_data_max < 400 else 100 if y2_data_max < 800 else 200
        y2_max = math.ceil(y2_data_max * 1.12 / y2_step) * y2_step
        
        ax.set_ylim(0, y1_max)
        ax2.set_ylim(0, y2_max)
        
        # --- Smart label positioning ---
        # Evaluate curve values at representative x-positions for label placement
        bbox_props = dict(boxstyle='square,pad=0.2', facecolor='white', edgecolor='none', alpha=1.0)
        
        # Fatmax label: place at the fatmax speed if visible
        if fatmax_speed > 0 and x_lo <= fatmax_speed <= x_hi:
            fatmax_idx_vis = np.argmin(np.abs(v_grid - fatmax_speed))
            fatmax_y_pos = ee_fat_kcal_h_grid[fatmax_idx_vis]
            ax.text(fatmax_speed, fatmax_y_pos * 0.5 if fatmax_y_pos > y1_max * 0.15 else y1_max * 0.08,
                    "Fatmax", color='white', fontsize=6, fontweight='bold', ha='center', va='center',
                    bbox=dict(boxstyle='square,pad=0.2', facecolor='#2cb7b9', edgecolor='none'), zorder=12)
        
        # Gesamtenergieverbrauch: place at ~60% of x-range, slightly above the total energy curve
        label_x_frac = 0.6
        label_v = x_lo + (x_hi - x_lo) * label_x_frac
        label_idx = np.argmin(np.abs(v_grid - label_v))
        ee_at_label = ee_kcal_h_grid[label_idx]
        # Position label above the total energy curve
        gesamt_y = min(ee_at_label + y1_max * 0.08, y1_max * 0.95)
        ax.text(label_v, gesamt_y, 'Gesamtenergieverbrauch', color='#595a59', fontsize=6.5,
                fontweight='bold', ha='center', va='center', bbox=bbox_props, zorder=10)
        
        # Verbrauch von Fetten: place at ~35% of x-range, below the fat curve  
        fat_label_v = x_lo + (x_hi - x_lo) * 0.35
        fat_label_idx = np.argmin(np.abs(v_grid - fat_label_v))
        fat_at_label = ee_fat_kcal_h_grid[fat_label_idx]
        # Position label below the fat curve
        fat_y = max(fat_at_label - y1_max * 0.08, y1_max * 0.05)
        ax.text(fat_label_v, fat_y, 'Verbrauch von Fetten', color='#2cb7b9', fontsize=6.5,
                fontweight='bold', ha='center', va='center', bbox=bbox_props, zorder=10)
        
        # Kohlenhydratverbrauch: place at ~75% of x-range, below the carb curve
        carb_label_v = x_lo + (x_hi - x_lo) * 0.75
        carb_label_idx = np.argmin(np.abs(v_grid - carb_label_v))
        carb_at_label = carb_g_h_grid[carb_label_idx]
        # Position below the carb curve in ax2 coordinates
        carb_y = max(carb_at_label - y2_max * 0.08, y2_max * 0.05)
        ax2.text(carb_label_v, carb_y, 'Kohlenhydratverbrauch', color='#cdb663', fontsize=6.5,
                 fontweight='bold', ha='center', va='center', bbox=bbox_props, zorder=10)
    else:
        ax.set_ylim(0, 1500)
        ax2.set_ylim(0, 400)
        
    ax.tick_params(axis='both', labelcolor='#595a59', labelsize=5, width=1.2)
    ax2.tick_params(axis='both', labelcolor='#595a59', labelsize=5, width=1.2)
    for s in ax.spines.values(): s.set_color('#595a59'); s.set_linewidth(1.2)
    for s in ax2.spines.values(): s.set_color('#595a59'); s.set_linewidth(1.2)
    ax.grid(True, which='both', linestyle='-', linewidth=0.5, color='#d8d8d8')
    ax.set_axisbelow(True)
    plt.tight_layout()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_ener_chart:
        fig.savefig(tmp_ener_chart.name, format="png", bbox_inches="tight")
        plt.close(fig)
        pdf.image(tmp_ener_chart.name, x=55.9, y=449.5, w=472, h=254)
        temp_files.append(tmp_ener_chart.name)

    # --- SEITE 5: TRAININGSBEREICHE & REFERENZDATEN ---
    pdf.add_page()
    draw_rect(50.9, 54.5, 479.1, 663.6, fill_color=c_white)
    
    page_logo(5)
        
    draw_rect(57.6, 133.9, 155.3, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 155.3 / 2.0, 135.0, "Referenzdaten", size=8.2, font_style='B', color=c_white)

    # Cartesian Radar chart for Reference data
    fig, ax = plt.subplots(figsize=(5.0, 5.0), dpi=300)
    ax.set_aspect('equal')
    
    radar_cats = ['VO2max', 'Fettanteil', 'Laufökonomie', 'ANS', 'Fatmax', 'KH-Match', 'VLamax']
    radar_labels = [
        'VO2max\n[ml/min/kg]',
        'Fettanteil\n[%]',
        'Laufökonomie\n[ml/min/kg*m/s]',
        'ANS\n[m/s]',
        'Fatmax\n[m/s]',
        'KH-Match\n[m/s]',
        'VLamax\n[mmol/l/s]'
    ]
    N = len(radar_cats)
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    angles += angles[:1]
    
    # Background rings (Sehr gut > Gut > Zu verbessern > Schwach)
    x_100 = [100 * np.cos(t) for t in angles]
    y_100 = [100 * np.sin(t) for t in angles]
    ax.fill(x_100, y_100, color='#cdb663', edgecolor='none', zorder=1, label='Sehr gut')
    
    x_80 = [80 * np.cos(t) for t in angles]
    y_80 = [80 * np.sin(t) for t in angles]
    ax.fill(x_80, y_80, color='#2cb7b9', edgecolor='none', zorder=2, label='Gut')
    
    x_60 = [60 * np.cos(t) for t in angles]
    y_60 = [60 * np.sin(t) for t in angles]
    ax.fill(x_60, y_60, color='#ffffff', edgecolor='none', zorder=3, label='Zu verbessern')
    
    x_35 = [35 * np.cos(t) for t in angles]
    y_35 = [35 * np.sin(t) for t in angles]
    ax.fill(x_35, y_35, color='#595a59', edgecolor='none', zorder=4, label='Schwach')
    
    # Normalize running reference values
    def normalize_ref_running(val, cat):
        ref_params = {
            'VO2max':       (60.0,  5.0,   False),
            'Fettanteil':   (13.0,  2.5,   True),
            'Laufökonomie': (210.0, 15.0,  True),
            'ANS':          (3.5,   0.4,   False),
            'Fatmax':       (2.7,   0.3,   False),
            'KH-Match':     (3.0,   0.3,   False),
            'VLamax':       (0.6,   0.1,   True),
        }
        if cat not in ref_params:
            return 50.0
        mean, sd, inverted = ref_params[cat]
        span = 6 * sd
        if span == 0:
            return 0.0
        rel = (val - (mean - 3 * sd)) / span * 100.0
        if inverted:
            rel = 100.0 - rel
        return max(0.0, min(100.0, rel))

    import streamlit as st
    pot_vo2max = float(st.session_state.get("pot_vo2max", 72.0))
    pot_fat = float(st.session_state.get("pot_fat", 10.5))
    pot_ans_rel = float(st.session_state.get("pot_ans_rel", 4.8))
    pot_fatmax = float(st.session_state.get("pot_fatmax", 3.4))
    pot_match = float(st.session_state.get("pot_match", 3.6))
    pot_vlamax = float(st.session_state.get("pot_vlamax", 0.45))

    re_val = 210.0
    valid_re = []
    for i in range(len(speed)):
        v_ms = speed[i]
        vo2_r = vo2_steady_values[i] if i < len(vo2_steady_values) else 0.0
        if vo2_r > 0:
            valid_re.append(vo2_r / (v_ms * 3.6) * 60.0)
    if len(valid_re) > 0:
        re_val = np.mean(valid_re)

    v_fatmax = v_ans * 0.78
    v_kh_match = v_ans * 0.85

    athlete_scores = [
        normalize_ref_running(rel_vo2max,    'VO2max'),
        normalize_ref_running(body_fat_pct,  'Fettanteil'),
        normalize_ref_running(re_val,        'Laufökonomie'),
        normalize_ref_running(v_ans,         'ANS'),
        normalize_ref_running(v_fatmax,      'Fatmax'),
        normalize_ref_running(v_kh_match,    'KH-Match'),
        normalize_ref_running(vla_val,       'VLamax'),
    ]
    athlete_scores += athlete_scores[:1]

    pot_v_ans = pot_ans_rel if pot_ans_rel < 10 else v_ans * 1.15
    pot_v_fatmax = pot_fatmax if pot_fatmax < 10 else v_fatmax * 1.15
    pot_v_match = pot_match if pot_match < 10 else v_kh_match * 1.15

    ans_change_str = st.session_state.get("pot_ans_select_key_lauft", "Keine Änderung")
    ans_change = 0.0
    if "+2" in ans_change_str: ans_change = 0.02
    elif "+5" in ans_change_str: ans_change = 0.05
    elif "-2" in ans_change_str: ans_change = -0.02
    elif "-5" in ans_change_str: ans_change = -0.05
    pot_re_val = re_val / (1.0 + ans_change)

    potential_scores = [
        normalize_ref_running(pot_vo2max,  'VO2max'),
        normalize_ref_running(pot_fat,     'Fettanteil'),
        normalize_ref_running(pot_re_val,  'Laufökonomie'),
        normalize_ref_running(pot_v_ans,   'ANS'),
        normalize_ref_running(pot_v_fatmax, 'Fatmax'),
        normalize_ref_running(pot_v_match, 'KH-Match'),
        normalize_ref_running(pot_vlamax,  'VLamax'),
    ]
    potential_scores += potential_scores[:1]

    # Draw spokes (radial lines)
    for t in angles[:-1]:
        ax.plot([0, 100 * np.cos(t)], [0, 100 * np.sin(t)], color='#595a59', linestyle='-', linewidth=0.8, zorder=5)

    # Draw outer heptagon spine boundary
    ax.plot(x_100, y_100, color='#595a59', linewidth=1.2, zorder=8)

    # Draw data lines
    x_ath = [r * np.cos(t) for r, t in zip(athlete_scores, angles)]
    y_ath = [r * np.sin(t) for r, t in zip(athlete_scores, angles)]
    ax.plot(x_ath, y_ath, color='#1a1a1a', linewidth=3.0, zorder=6)
    ax.fill(x_ath, y_ath, color='#1a1a1a', alpha=0.08, zorder=5)

    x_pot = [r * np.cos(t) for r, t in zip(potential_scores, angles)]
    y_pot = [r * np.sin(t) for r, t in zip(potential_scores, angles)]
    ax.plot(x_pot, y_pot, color='#595a59', linewidth=2.0, linestyle='--', zorder=6)

    # Add text labels manually
    for i, label in enumerate(radar_labels):
        t = angles[i]
        r_label = 112
        x = r_label * np.cos(t)
        y = r_label * np.sin(t)
        
        ha = 'center'
        if np.cos(t) > 0.1:
            ha = 'left'
        elif np.cos(t) < -0.1:
            ha = 'right'
            
        va = 'center'
        if np.sin(t) > 0.1:
            va = 'bottom'
        elif np.sin(t) < -0.1:
            va = 'top'
            
        if abs(np.cos(t)) < 0.1:
            if np.sin(t) > 0:
                y += 2
            else:
                y -= 2
                
        ax.text(x, y, label, fontsize=5.0, fontweight='bold', color='#595a59', ha=ha, va=va)

    ax.axis('off')
    ax.set_xlim(-140, 140)
    ax.set_ylim(-140, 140)
    
    from matplotlib.lines import Line2D
    legend_handles = [
        Patch(facecolor='#cdb663', alpha=1.0,  label='Sehr gut'),
        Patch(facecolor='#2cb7b9', alpha=1.0,  label='Gut'),
        Patch(facecolor='#ffffff', edgecolor='#595a59', alpha=1.0,  label='Zu verbessern'),
        Patch(facecolor='#595a59', alpha=1.0,  label='Schwach'),
        Line2D([0], [0], color='#1a1a1a', lw=2, label='Daten Sportler'),
        Line2D([0], [0], color='#595a59', lw=1.5, ls='--', label='Coaching Potential'),
    ]
    legend = ax.legend(handles=legend_handles, loc='lower center', bbox_to_anchor=(0.5, -0.15),
                       ncol=3, fontsize=5, frameon=True, framealpha=0.9)
    for text in legend.get_texts():
        text.set_color('#595a59')
        text.set_weight('bold')
        
    fig.patch.set_facecolor('white')
    plt.tight_layout()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_radar:
        fig.savefig(tmp_radar.name, format="png", bbox_inches="tight", facecolor='white')
        plt.close(fig)
        
        img = Image.open(tmp_radar.name)
        w_img, h_img = img.size
        max_dim = max(w_img, h_img)
        new_img = Image.new("RGB", (max_dim, max_dim), (255, 255, 255))
        new_img.paste(img, ((max_dim - w_img) // 2, (max_dim - h_img) // 2))
        new_img.save(tmp_radar.name)
        
        pdf.image(tmp_radar.name, x=160.1, y=155.0, w=275.0, h=275.0)
        temp_files.append(tmp_radar.name)

    # Trainingsbereiche Table
    draw_rect(57.6, 471.4, 207.0, 10.44, fill_color=c_teal)
    draw_text_centered(57.6 + 207.0 / 2.0, 472.5, "Trainingsbereiche", size=8.2, font_style='B', color=c_white)

    draw_text(99.3, 497.0, "Trainingsbereich", size=9.0, font_style='B', color=c_dark_grey)
    draw_text(254.8, 497.0, "Geschwindigkeit", size=9.0, font_style='B', color=c_dark_grey)
    draw_text(417.4, 497.0, "Herzfrequenz", size=9.0, font_style='B', color=c_dark_grey)
    draw_text(275.3, 510.7, "[min/km]", size=8.3, color=c_dark_grey)
    draw_text(435.8, 510.7, "[bpm]", size=8.3, color=c_dark_grey)
    
    draw_text(231.8, 523.9, "min", size=8.3, font_style='B', color=c_dark_grey)
    draw_text(282.6, 523.9, "max", size=8.3, font_style='B', color=c_dark_grey)
    draw_text(335.5, 523.9, "Ziel", size=8.3, color=c_dark_grey)
    draw_text(387.0, 523.9, "min", size=8.3, font_style='B', color=c_dark_grey)
    draw_text(437.9, 523.9, "max", size=8.3, font_style='B', color=c_dark_grey)
    draw_text(490.7, 523.9, "Ziel", size=8.3, color=c_dark_grey)

    def format_pace(v):
        if not isinstance(v, (int, float)) or v <= 0:
            return "-"
        seconds_per_km = 1000.0 / v
        minutes = int(seconds_per_km // 60)
        seconds = int(round(seconds_per_km % 60))
        if seconds == 60:
            minutes += 1
            seconds = 0
        return f"{minutes:02d}:{seconds:02d}"

    def format_hr(hr_val):
        if not isinstance(hr_val, (int, float)) or hr_val <= 0:
            return "-"
        return str(int(round(hr_val)))

    ans_hf = 160.0
    if not ans_row.empty:
        try:
            ans_hf = float(ans_row.iloc[0].get('HF', 160.0))
        except:
            pass

    if v_ans > 0 and ans_hf > 0:
        zones_data = [
            ("regenerativer Dauerlauf (DLreg)",
             "-", format_pace(v_ans * 0.68), format_pace(v_ans * 0.65),
             "-", format_hr(ans_hf * 0.77), format_hr(ans_hf * 0.70)),
            
            ("extensiver Dauerlauf (DLext)",
             format_pace(v_ans * 0.68), format_pace(v_ans * 0.79), format_pace(v_ans * 0.77),
             format_hr(ans_hf * 0.77), format_hr(ans_hf * 0.85), format_hr(ans_hf * 0.80)),
            
            ("intensiver  Dauerlauf (DLint)",
             format_pace(v_ans * 0.79), format_pace(v_ans * 0.93), format_pace(v_ans * 0.86),
             format_hr(ans_hf * 0.85), format_hr(ans_hf * 0.95), format_hr(ans_hf * 0.91)),
            
            ("Tempodauerlauf (DLTempo)",
             format_pace(v_ans * 0.93), format_pace(v_ans * 1.05), format_pace(v_ans * 0.99),
             format_hr(ans_hf * 0.95), format_hr(ans_hf * 1.03), format_hr(ans_hf * 0.99)),
            
            ("extensives Tempotraining (TText)",
             format_pace(v_ans * 1.05), format_pace(v_ans * 1.23), format_pace(v_ans * 1.13),
             format_hr(ans_hf * 1.03), format_hr(ans_hf * 1.05), format_hr(ans_hf * 1.05)),
            
            ("intensives Tempotraining (TTint)",
             format_pace(v_ans * 1.23), "max", format_pace(v_ans * 1.25),
             format_hr(ans_hf * 1.05), "max", format_hr(ans_hf * 1.06)),
        ]
    else:
        zones_data = [
            ("regenerativer Dauerlauf (DLreg)", "-", "-", "-", "-", "-", "-"),
            ("extensiver Dauerlauf (DLext)",     "-", "-", "-", "-", "-", "-"),
            ("intensiver  Dauerlauf (DLint)",     "-", "-", "-", "-", "-", "-"),
            ("Tempodauerlauf (DLTempo)",          "-", "-", "-", "-", "-", "-"),
            ("extensives Tempotraining (TText)",  "-", "-", "-", "-", "-", "-"),
            ("intensives Tempotraining (TTint)",  "-", "-", "-", "-", "-", "-"),
        ]

    y_coords_s5 = [536.1, 549.4, 562.7, 576.0, 602.7, 616.0]
    for (zn, p_min, p_max, p_zl, hr_mn, hr_mx, hr_zl), y_coord in zip(zones_data, y_coords_s5):
        draw_text(59.3, y_coord, zn, size=8.3, font_style='B', color=c_dark_grey)
        
        if p_min == "-":
            draw_text_centered(231.8 + 6.0, y_coord, "-", size=8.3, color=c_dark_grey)
        else:
            draw_text(231.8, y_coord, p_min, size=8.3, color=c_dark_grey)
            
        draw_text(282.6, y_coord, p_max, size=8.3, color=c_dark_grey)
        draw_text(335.5, y_coord, p_zl, size=8.3, color=c_dark_grey)
        
        if hr_mn == "-":
            draw_text_centered(387.0 + 6.0, y_coord, "-", size=8.3, color=c_dark_grey)
        else:
            draw_text(387.0, y_coord, hr_mn, size=8.3, color=c_dark_grey)
            
        draw_text(437.9, y_coord, hr_mx, size=8.3, color=c_dark_grey)
        draw_text(490.7, y_coord, hr_zl, size=8.3, color=c_dark_grey)

    for x in [212.9, 368.8, 524.7]:
        draw_rect(x, 522.0, 0.6, 104.0, fill_color=c_dark_grey)
    draw_rect(212.9, 625.4, 311.8, 0.6, fill_color=c_dark_grey)

    pdf_output = pdf.output(dest='S')
    
    for f in temp_files:
        try:
            os.unlink(f)
        except Exception:
            pass
            
    return pdf_output.encode('latin-1', 'replace')


# 3. BERECHNUNG & ANZEIGE
if st.session_state.get("lauf_auswertung_gestartet", False):
    raw_speeds = st.session_state.df_lauftest_input["v (m/s)"].values
    raw_lactate = st.session_state.df_lauftest_input["Laktat"].values
    raw_hr = st.session_state.df_lauftest_input["HF"].values
    
    # Exclude Ruhe (first row) for exercise steps calculations
    speed = raw_speeds[1:]
    lactate = raw_lactate[1:]
    hr = raw_hr[1:]
    
    ruhe_lac = raw_lactate[0]
    ruhe_hr = raw_hr[0]
    abs_vo2max, rel_vo2max = 0.0, 0.0
    steady_fenster = 120      # steady state window in seconds
    vo2_steady_values = []
    vo2_steady_abs = []      # absoluter VO2-Steady-State [ml/min] je Stufe
    vco2_steady_abs = []     # absoluter VCO2-Steady-State [ml/min] je Stufe
    window_coords = []
    spiro_df = pd.DataFrame()

    sf_means = st.session_state.df_sf[["M1", "M2", "M3"]].mean(axis=1)
    sum_sf = sf_means.sum()
    if sum_sf > 0:
        body_fat_pct = (22.32 * np.log10(sum_sf)) - 29.2

    spiro_interval_sec = None   # wird beim Einlesen ermittelt (10 oder 30)
    spiro_interval_note = ""    # Hinweis bei 30s für VO2max-Anzeige

    if uploaded_file is not None:
        try:
            if "lauftest_spiro_df" in st.session_state:
                spiro_df = st.session_state.lauftest_spiro_df
            else:
                df_excel = pd.read_csv(uploaded_file, header=None, low_memory=False) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file, header=None)
                times_sec = df_excel.iloc[3:, 9].apply(parse_time_to_seconds)
                vo2_col = pd.to_numeric(df_excel.iloc[3:, 13], errors='coerce')
                vco2_col = pd.to_numeric(df_excel.iloc[3:, 14], errors='coerce')
                spiro_df = pd.DataFrame({'Time': times_sec, 'VO2': vo2_col, 'VCO2': vco2_col}).dropna(subset=['Time'])
                st.session_state.lauftest_spiro_df = spiro_df

            # --- Zeitintervall automatisch erkennen ---
            if len(spiro_df) >= 2:
                t_diffs = spiro_df['Time'].diff().dropna()
                spiro_interval_sec = int(round(t_diffs.median()))
            else:
                spiro_interval_sec = 10  # default

            if spiro_interval_sec == 30:
                spiro_interval_note = "(aus 30s-Auflösung, Genauigkeit eingeschränkt)"
            elif spiro_interval_sec == 10:
                spiro_interval_note = ""
            else:
                spiro_interval_note = f"({spiro_interval_sec}s-Auflösung)"

            if not spiro_df.empty:
                # VO2max Rolling-Window: 30s rolling average (3 points for 10s, 1 point for 30s)
                window_pts = max(1, int(30 / spiro_interval_sec))
                abs_vo2max = spiro_df['VO2'].rolling(window=window_pts, center=True).mean().max()
                rel_vo2max = abs_vo2max / weight if weight > 0 else 0

                stufendauer = st.session_state.get("stufendauer", 5.0)
                pausendauer = st.session_state.get("pausendauer", 30)
                vorlauf = st.session_state.get("vorlauf", 60)
                stage_sec = int(stufendauer * 60)
                pause_sec = int(pausendauer)
                for i in range(anzahl):
                    if "detected_stage_ends" in st.session_state and i < len(st.session_state.detected_stage_ends):
                        end_t_base = st.session_state.detected_stage_ends[i]
                    else:
                        end_t_base = vorlauf + (i + 1) * stage_sec + i * pause_sec
                    
                    end_t_val = st.session_state.get(f"lauftest_end_t_{i}", end_t_base)
                    end_t = float(end_t_val)
                    start_t = end_t - steady_fenster
                    
                    mask = (spiro_df['Time'] > start_t) & (spiro_df['Time'] <= end_t)
                    window_coords.append((start_t, end_t))
                    stage_data = spiro_df.loc[mask]
                    vo2_val  = stage_data['VO2'].mean()  if not stage_data.empty else 0.0
                    vco2_val = stage_data['VCO2'].mean() if not stage_data.empty else 0.0
                    vo2_steady_values.append(vo2_val / weight if weight > 0 else 0.0)
                    vo2_steady_abs.append(vo2_val)
                    vco2_steady_abs.append(vco2_val)
        except Exception as e:
            st.error(f"Fehler beim Einlesen: {e}")

    # Override VO2max if manual value is provided
    if vo2max_override > 0:
        abs_vo2max = float(vo2max_override)
        rel_vo2max = abs_vo2max / weight if weight > 0 else 0.0


    # --- SCHWELLENBERECHNUNG ---
    v_start, v_end = speed[0], speed[-1]
    results = []

    # 1. Fits
    # Fit polynomial on exercise stages for Dmax/ModDmax:
    poly_coeffs_no = np.polyfit(speed, lactate, 3)
    poly_func_no = np.poly1d(poly_coeffs_no)
    
    # Fit polynomial on all data (including adjusted baseline) for OBLA and LTP:
    speed_with_ruhe = raw_speeds.copy()
    if len(raw_speeds) > 2 and raw_speeds[0] == 0.0:
        step_size = raw_speeds[2] - raw_speeds[1]
        speed_with_ruhe[0] = raw_speeds[1] - step_size
    lactate_with_ruhe = raw_lactate.copy()
    poly_coeffs_with = np.polyfit(speed_with_ruhe, lactate_with_ruhe, 3)
    poly_func_with = np.poly1d(poly_coeffs_with)

    # 2. OBLA 4.0 (ANS)
    if 4.0 < lactate[0] or 4.0 > lactate[-1]:
        v_exact_40 = None
    else:
        v_exact_40 = np.interp(4.0, lactate, speed)
        
    if v_exact_40 is not None:
        results.append({'Modell': 'OBLA 4.0 (ANS)', 'v': v_exact_40, 'Laktat': 4.0})
    else:
        results.append({'Modell': 'OBLA 4.0 (ANS)', 'v': 'Laktatwerte zu niedrig um OBLA zu kalkulieren', 'Laktat': None})

    # 3. Dmax (Standard)
    # Solves perpendicular distance to line from first exercise stage speed[0] to last stage speed[-1]
    c3, c2, c1, _ = poly_coeffs_no
    m_dmax = (lactate[-1] - lactate[0]) / (speed[-1] - speed[0])
    d_roots = np.roots([3 * c3, 2 * c2, c1 - m_dmax])
    valid_d_roots = [r.real for r in d_roots if np.isreal(r) and speed[0] <= r.real <= speed[-1]]
    if valid_d_roots:
        v_exact_dmax = max(valid_d_roots)
        results.append({'Modell': 'Dmax (Standard)', 'v': v_exact_dmax, 'Laktat': poly_func_no(v_exact_dmax)})
    else:
        results.append({'Modell': 'Dmax (Standard)', 'v': 'Kalkulation nicht möglich', 'Laktat': None})

    # 4. Modified Dmax
    # Line starts at point preceding first rise in lactate >= 0.4 mmol/L
    diffs = lactate[1:] - lactate[:-1]
    idx_first_rise = None
    for i, d in enumerate(diffs):
        if d >= 0.4:
            idx_first_rise = i
            break
            
    if not ausbelastung:
        results.append({'Modell': 'Modified Dmax', 'v': 'Aufgrund fehlender Ausbelastung keine Kalkulation der ModDmax möglich', 'Laktat': None})
    elif idx_first_rise is None:
        results.append({'Modell': 'Modified Dmax', 'v': 'Keine Laktat-Steigerung >= 0.4 mmol/L gefunden', 'Laktat': None})
    else:
        speed_start = speed[idx_first_rise]
        lac_start = lactate[idx_first_rise]
        m_mod = (lactate[-1] - lac_start) / (speed[-1] - speed_start)
        mod_roots = np.roots([3 * c3, 2 * c2, c1 - m_mod])
        valid_mod_roots = [r.real for r in mod_roots if np.isreal(r) and speed_start <= r.real <= speed[-1]]
        if valid_mod_roots:
            v_exact_mod = max(valid_mod_roots)
            results.append({'Modell': 'Modified Dmax', 'v': v_exact_mod, 'Laktat': poly_func_no(v_exact_mod)})
        else:
            results.append({'Modell': 'Modified Dmax', 'v': 'Kalkulation nicht möglich', 'Laktat': None})

    # 5. LTP1 & LTP2
    try:
        # Interpolate raw speed and lactate onto 0.1 m/s grid (excluding Ruhe)
        grid_speed = np.arange(speed[0], speed[-1] + 1e-9, 0.1)
        interpolated_lactate = np.interp(grid_speed, speed, lactate)
        
        def fit_3seg_fixed_bps(bp1, bp2, x, y):
            if bp1 <= x[0] or bp2 >= x[-1] or bp1 >= bp2:
                return 1e10
            A = np.column_stack([
                np.ones_like(x),
                x,
                np.maximum(0.0, x - bp1),
                np.maximum(0.0, x - bp2)
            ])
            try:
                coefs, residuals, rank, s_vals = np.linalg.lstsq(A, y, rcond=None)
                if len(residuals) > 0:
                    return residuals[0]
                else:
                    pred = A @ coefs
                    return np.sum((y - pred)**2)
            except:
                return 1e10

        # Coarse grid search (30 x 30)
        best_err = 1e10
        best_bps = None
        bp1_vals = np.linspace(speed[0] + 0.05, speed[-1] - 0.05, 30)
        bp2_vals = np.linspace(speed[0] + 0.05, speed[-1] - 0.05, 30)
        for bp1 in bp1_vals:
            for bp2 in bp2_vals:
                if bp1 < bp2:
                    err = fit_3seg_fixed_bps(bp1, bp2, grid_speed, interpolated_lactate)
                    if err < best_err:
                        best_err = err
                        best_bps = (bp1, bp2)

        # Refine search with Nelder-Mead
        import scipy.optimize as opt
        def loss_func(bps):
            return fit_3seg_fixed_bps(bps[0], bps[1], grid_speed, interpolated_lactate)
            
        res_ltp = opt.minimize(loss_func, x0=best_bps, method='Nelder-Mead')
        v_exact_ltp1, v_exact_ltp2 = res_ltp.x[0], res_ltp.x[1]
        
        # Associated lactates from polynomial with baseline
        lac_ltp1 = poly_func_with(v_exact_ltp1)
        lac_ltp2 = poly_func_with(v_exact_ltp2)
        
        results.append({'Modell': 'LTP1', 'v': v_exact_ltp1, 'Laktat': lac_ltp1})
        results.append({'Modell': 'LTP2', 'v': v_exact_ltp2, 'Laktat': lac_ltp2})
    except Exception as e:
        pass


    # Ergebnisanzeige Matrix aufbauen mit Error-Handling
    df_res_list = []
    for res in results:
        row = {'Modell': res['Modell'], 'v': res['v']}
        
        if isinstance(res['v'], str):
            row['m/s'] = res['v']
            row['km/h'] = ""
            row['Laktat'] = ""
            row['HF'] = ""
            row['VLamax [mmol/l/s]'] = ""
        else:
            row['m/s'] = round(res['v'], 2)
            row['km/h'] = round(res['v'] * 3.6, 1)
            row['Laktat'] = round(res['Laktat'], 2)
            
            # Linear model for HF instead of interp
            hr_mask = (speed > 0) & (hr > 0) & (~np.isnan(speed)) & (~np.isnan(hr))
            if hr_mask.sum() >= 2:
                hr_slope, hr_intercept = np.polyfit(speed[hr_mask], hr[hr_mask], 1)
                row['HF'] = int(round(hr_slope * res['v'] + hr_intercept, 0))
            else:
                row['HF'] = int(round(np.interp(res['v'], speed, hr), 0))
            
            if uploaded_file is None and vo2max_override == 0:
                row['VLamax [mmol/l/s]'] = "Bitte lade eine Spirodatei hoch"
            else:
                vla, _, _ = calculate_vlamax_for_v(res['v'], rel_vo2max, speed, vo2_steady_abs, vco2_steady_abs, weight)
                row['VLamax [mmol/l/s]'] = f"{vla:.2f}" if vla is not None else "N/A"
                
        df_res_list.append(row)

    df_res = pd.DataFrame(df_res_list)
    
    # 1. Determine active threshold model name
    selected_model_name = "OBLA 4.0 (ANS)"
    if "Neues Protokoll" in test_type:
        selected_model_name = st.session_state.get("selected_threshold_model_widget", "OBLA 4.0 (ANS)")
        
    # 2. Extract v_ans and vla_val for downstream calculations
    v_ans = 3.5
    vla_val = 0.6
    hf_ans_val = "-"
    ans_pace_str = "-"
    vla_display = "N/A"
    
    if not df_res.empty:
        selected_row = df_res[df_res['Modell'] == selected_model_name]
        if selected_row.empty:
            selected_row = df_res[df_res['Modell'].str.contains('OBLA 4.0', regex=False, na=False)]
        if not selected_row.empty:
            v_val = selected_row.iloc[0]['v']
            if not isinstance(v_val, str) and pd.notna(v_val):
                v_ans = float(v_val)
                p_s = 1000.0 / v_ans if v_ans > 0 else 0
                ans_pace_str = f"{int(p_s // 60)}:{int(p_s % 60):02d}"
                
            hf_ans_val = selected_row.iloc[0]['HF']
            vla_raw = selected_row.iloc[0]['VLamax [mmol/l/s]']
            vla_display = str(vla_raw).replace('.', ',') if isinstance(vla_raw, (int, float)) else str(vla_raw)
            if pd.notna(vla_raw) and vla_raw != 'N/A' and 'bitte' not in str(vla_raw).lower():
                try:
                    vla_val = float(vla_raw)
                except ValueError:
                    pass

    # 3. Reorder df_res to put the active model row at the very top (index 0)
    if not df_res.empty:
        selected_idx = df_res[df_res['Modell'] == selected_model_name].index
        if not selected_idx.empty:
            idx = selected_idx[0]
            df_res = pd.concat([df_res.iloc[[idx]], df_res.drop(idx)]).reset_index(drop=True)

    # Header-Metriken & Körperfett-Tacho
    st.markdown(f"""
        <div style="font-size: 16px; line-height: 1.5; margin-bottom: 20px;">
            <b>Athlet:</b> {athlete_name}<br>
            <b>Protokoll:</b> {test_type}<br>
            <b>Test-Datum:</b> {test_date}
        </div>
        """, unsafe_allow_html=True)
        
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Gewicht", f"{weight} kg")
    
    with m3:
        st.metric("VO2max (rel)", f"{rel_vo2max:.1f} ml/min/kg")
        st.metric("VO2max (abs.)", f"{int(abs_vo2max)} ml/min")
        st.metric("VLamax", f"{vla_display}")
        if spiro_interval_note:
            st.caption(f"⚠️ {spiro_interval_note}")
            
    with m4:
        st.metric("ANS Pace", f"{ans_pace_str} min/km")
        st.metric("HF @ ANS", f"{hf_ans_val} bpm")
        
        v_fatmax = v_ans * 0.78
        if v_fatmax > 0:
            p_f_s = 1000.0 / v_fatmax
            fatmax_pace_str = f"{int(p_f_s // 60)}:{int(p_f_s % 60):02d}"
        else:
            fatmax_pace_str = "-"
        st.metric("Geschw. @ Fatmax", f"{fatmax_pace_str} min/km")
    
    with m2:
        fig_bf = go.Figure(go.Indicator(
            mode = "gauge+number",
            value = max(0, body_fat_pct),
            number = {'suffix': "%", 'valueformat': ".1f", 'font': {'color': '#00a1e0', 'size': 30}},
            title = {'text': "Körperfett", 'font': {'size': 14}},
            gauge = {
                'axis': {'range': [0, 30], 'tickwidth': 1, 'tickcolor': "darkblue"},
                'bar': {'color': "#00a1e0"},
                'bgcolor': "white",
                'borderwidth': 2,
                'bordercolor': "gray",
                'steps': [
                    {'range': [0, 10], 'color': "rgba(0, 161, 224, 0.1)"},
                    {'range': [10, 20], 'color': "rgba(0, 161, 224, 0.3)"},
                    {'range': [20, 30], 'color': "rgba(0, 161, 224, 0.5)"}],
            }
        ))
        fig_bf.update_layout(height=150, margin=dict(l=10, r=10, t=30, b=10))
        st.plotly_chart(fig_bf, width='stretch')

    # 4. Coaching Potential Override Simulation
    def parse_sel_pct(sel_str):
        if "+2" in sel_str: return 0.02
        if "+5" in sel_str: return 0.05
        if "-2" in sel_str: return -0.02
        if "-5" in sel_str: return -0.05
        return 0.0

    vo2_change = parse_sel_pct(st.session_state.get("pot_vo2_select_key_lauft", "Keine Änderung"))
    vla_change = parse_sel_pct(st.session_state.get("pot_vla_select_key_lauft", "Keine Änderung"))
    ans_change = parse_sel_pct(st.session_state.get("pot_ans_select_key_lauft", "Keine Änderung"))

    pot_vo2max = rel_vo2max * (1.0 + vo2_change)
    pot_vlamax = vla_val * (1.0 + vla_change)

    has_spiro = (uploaded_file is not None) and (len(vo2_steady_abs) > 0)
    if has_spiro:
        try:
            vo2_rest = 3.5 * weight
            ac_stages = []
            for v, vo2, vco2 in zip(speed, vo2_steady_abs, vco2_steady_abs):
                if v <= 0 or vo2 <= 0:
                    ac_stages.append(0.0)
                    continue
                net_vo2 = max(0.0, vo2 - vo2_rest)
                rq = vco2 / vo2 if vco2 > 0 else 0.85
                kal_equiv = (rq - 0.7) * 5.094 + 19.6
                ee_kJ_min = (net_vo2 / 1000.0) * kal_equiv
                vo2_100kh = ee_kJ_min / 21.1 * 1000.0
                ac = vo2_100kh / (v * 100.0)
                ac_stages.append(ac)

            ex_speeds = [v for v in speed if v > 0]
            if len(ex_speeds) >= 2:
                v_grid = np.arange(2.6, 8.61, 0.01)
                ac_grid = []
                for v in v_grid:
                    if v <= ex_speeds[0]:
                        m = (ac_stages[1] - ac_stages[0]) / (ex_speeds[1] - ex_speeds[0])
                        c_val = ac_stages[0] - m * ex_speeds[0]
                        ac = m * v + c_val
                    elif v >= ex_speeds[-1]:
                        m = (ac_stages[-1] - ac_stages[-2]) / (ex_speeds[-1] - ex_speeds[-2])
                        c_val = ac_stages[-1] - m * ex_speeds[-1]
                        ac = m * v + c_val
                    else:
                        for i in range(len(ex_speeds) - 1):
                            if ex_speeds[i] <= v < ex_speeds[i+1]:
                                m = (ac_stages[i+1] - ac_stages[i]) / (ex_speeds[i+1] - ex_speeds[i])
                                c_val = ac_stages[i] - m * ex_speeds[i]
                                ac = m * v + c_val
                                break
                    ac_grid.append(ac)
                ac_grid = np.array(ac_grid)

                # Improve running economy if ANS is overridden
                eco_factor = 1.0 / (1.0 + ans_change)
                pot_ac_grid = ac_grid * eco_factor

                pot_vo2_rel_demand_grid = (pot_ac_grid * v_grid * 100.0) / weight + 1.0
                pot_delta_vo2_grid = pot_vo2max - pot_vo2_rel_demand_grid

                pot_lac_clear_grid = (0.02049 * pot_vo2_rel_demand_grid) / 0.4
                pot_net_clearance = np.full_like(v_grid, -9999.0)
                pot_carb_pct_grid = np.zeros_like(v_grid)

                ks1, ks2 = 0.0631, 1.331
                for idx, (v, vo2_rel, d_vo2) in enumerate(zip(v_grid, pot_vo2_rel_demand_grid, pot_delta_vo2_grid)):
                    if d_vo2 > 0 and vo2_rel > 0:
                        adp = np.sqrt((ks1 * vo2_rel) / d_vo2)
                        adp_3 = adp ** 3
                        pfk = 1 + (ks2 / adp_3)
                        vla_min = pot_vlamax * 60 / pfk
                        pot_net_clearance[idx] = pot_lac_clear_grid[idx] - vla_min
                        fat_pct = max(0.0, min(100.0, 100.0 - (vla_min / pot_lac_clear_grid[idx] * 100.0)))
                        pot_carb_pct_grid[idx] = 100.0 - fat_pct
                    else:
                        pot_net_clearance[idx] = -9999.0
                        pot_carb_pct_grid[idx] = 100.0

                pot_ans_idx = None
                for idx in range(len(v_grid)):
                    if pot_net_clearance[idx] < 0 and idx > 0 and pot_net_clearance[idx-1] >= 0:
                        pot_ans_idx = idx - 1
                        break
                if pot_ans_idx is None:
                    pot_ans_idx = np.argmin(np.abs(pot_net_clearance))

                pot_v_ans = v_grid[pot_ans_idx]

                pot_fatmax_idx = np.argmax(pot_net_clearance)
                pot_v_fatmax = v_grid[pot_fatmax_idx]

                pot_rq_grid = 0.7 + 0.3 * (pot_carb_pct_grid / 100.0)
                pot_kal_equiv_grid = (pot_rq_grid - 0.7) * 5.094 + 19.6
                pot_ee_kcal_h_grid = (((pot_vo2_rel_demand_grid * weight / 1000.0) * pot_kal_equiv_grid) * 60.0) / 4.186
                pot_carb_g_h_grid = (pot_ee_kcal_h_grid * (pot_carb_pct_grid / 100.0)) / 4.063

                pot_match_idx = np.argmin(np.abs(pot_carb_g_h_grid - 45.0))
                pot_v_match = v_grid[pot_match_idx]
            else:
                pot_v_ans = v_ans * (1.0 + ans_change)
                pot_v_fatmax = pot_v_ans * 0.78
                pot_v_match = pot_v_ans * 0.85
        except:
            pot_v_ans = v_ans * (1.0 + ans_change)
            pot_v_fatmax = pot_v_ans * 0.78
            pot_v_match = pot_v_ans * 0.85
    else:
        pot_v_ans = v_ans * (1.0 + ans_change)
        pot_v_fatmax = pot_v_ans * 0.78
        pot_v_match = pot_v_ans * 0.85

    st.session_state.pot_weight_val = weight
    st.session_state.pot_vo2max = pot_vo2max
    st.session_state.pot_vlamax = pot_vlamax
    st.session_state.pot_ans_rel = pot_v_ans
    st.session_state.pot_fatmax = pot_v_fatmax
    st.session_state.pot_match = pot_v_match
    st.session_state.pot_fat = body_fat_pct

    poly_func = poly_func_no
    fig_laktat = None

    if not df_res.empty:
        c1, c2 = st.columns([1, 1])
        with c1:
            st.subheader("Schwellen & VLamax Matrix")
            # Style the active threshold row (index 0) with a light gray background and bold text
            styled_df = df_res[['Modell', 'm/s', 'km/h', 'Laktat', 'HF', 'VLamax [mmol/l/s]']]
            
            def fmt_2_decimals(x):
                if pd.isna(x): return ""
                if isinstance(x, (int, float, np.number)): return f"{x:.2f}".replace('.', ',')
                try: return f"{float(x):.2f}".replace('.', ',')
                except ValueError: return str(x)

            def make_row_gray_bold(row):
                if row.name == 0:
                    return ['background-color: #f0f2f6; font-weight: bold; color: #1a1a1a;'] * len(row)
                return [''] * len(row)
                
            styled_table = styled_df.style.format({
                'm/s': fmt_2_decimals,
                'km/h': fmt_2_decimals,
                'Laktat': fmt_2_decimals,
                'HF': fmt_2_decimals,
                'VLamax [mmol/l/s]': fmt_2_decimals
            }).apply(make_row_gray_bold, axis=1)

            st.dataframe(styled_table, hide_index=True, width='stretch')
        with c2:
            fig, ax = plt.subplots(figsize=(8, 4))
            v_smooth = np.linspace(v_start, v_end, 200)
            ax.plot(speed, lactate, 'ko', label='Messwerte')
            ax.plot(v_smooth, poly_func(v_smooth), color='#00a1e0', label='Laktatkurve')
            
            for _, r in df_res.iterrows():
                if isinstance(r['v'], (int, float, np.number)):
                    ax.plot(r['v'], r['Laktat'], 'X', markersize=8, label=r['Modell'])
                    
            ax.set_xlabel("Geschwindigkeit (m/s)")
            ax.set_ylabel("Laktat (mmol/L)")
            ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
            st.pyplot(fig)
            fig_laktat = fig # Graph speichern für das PDF

    # LAUFÖKONOMIE SECTION
    if not spiro_df.empty and len(vo2_steady_abs) > 0 and any(v > 0 for v in vo2_steady_abs):
        st.markdown("---")
        st.subheader("🏃 Laufökonomie & VLamax je Stufe")

        lauf_rows = []
        for i in range(anzahl):
            v_ms   = speed[i]
            v_kmh  = round(v_ms * 3.6, 1)
            vo2_a  = vo2_steady_abs[i]  if i < len(vo2_steady_abs)  else 0.0
            vco2_a = vco2_steady_abs[i] if i < len(vco2_steady_abs) else 0.0
            vo2_r  = vo2_steady_values[i]   # ml/min/kg

            # RE in ml/kg/km (Literaturstandard): VO2ss_rel [ml/min/kg] / v [km/h] * 60
            re_ml_kg_km = (vo2_r / v_kmh * 60.0) if (vo2_r > 0 and v_kmh > 0) else None

            # RQ
            rq_val = (vco2_a / vo2_a) if (vco2_a > 0 and vo2_a > 0) else None

            # VLamax je Stufe via bestehender Funktion
            vla_stage = None
            if rel_vo2max > 0 and vo2_r > 0 and len(vo2_steady_values) > 1:
                vla_stage, _, _ = calculate_vlamax_for_v(
                    v_ms, rel_vo2max, speed, vo2_steady_abs, vco2_steady_abs, weight
                )

            pace_sec = (1000.0 / v_ms) if v_ms > 0 else 0
            pace_str = f"{int(pace_sec // 60)}:{int(pace_sec % 60):02d}"

            lauf_rows.append({
                'v_ms':       v_ms,
                'v_kmh':      v_kmh,
                'pace':       pace_str,
                'vo2_abs':    vo2_a,
                'vo2_rel':    vo2_r,
                'rq':         rq_val,
                're_ml_kg_km':re_ml_kg_km,
                'vla_stage':  vla_stage
            })

        valid_rows = [r for r in lauf_rows if r['re_ml_kg_km'] is not None]

        if len(valid_rows) >= 1:
            col_lt, col_lc = st.columns([1, 1.3])

            with col_lt:
                st.markdown("**Messwerte je Stufe**")
                df_lauf_display = pd.DataFrame([{
                    'v [km/h]':           r['v_kmh'],
                    'Pace [min/km]':      r['pace'],
                    'VO2ss [ml/min/kg]':  round(r['vo2_rel'], 2) if r['vo2_rel'] > 0 else '-',
                    'RQ':                 round(r['rq'], 3)      if r['rq']       else '-',
                    'RE [ml/kg/km]':      round(r['re_ml_kg_km'], 1) if r['re_ml_kg_km'] else '-',
                    'VLamax [mmol/L/s]':  round(r['vla_stage'], 3)   if r['vla_stage']   else '-'
                } for r in lauf_rows])
                st.dataframe(df_lauf_display, hide_index=True, width='stretch')
                st.caption(
                    "RE [ml/kg/km] = VO2ss [ml/min/kg] ÷ v [km/h] × 60  "
                    "| VLamax je Stufe aus aerober Metabolik-Gleichung"
                )

            with col_lc:
                # Chart: RE [ml/kg/km] und VO2ss auf zwei Y-Achsen
                speeds_v    = [r['v_kmh']       for r in valid_rows]
                re_vals     = [r['re_ml_kg_km'] for r in valid_rows]
                vo2_r_vals  = [r['vo2_rel']     for r in lauf_rows if r['vo2_rel'] > 0]
                speeds_vo2  = [r['v_kmh']       for r in lauf_rows if r['vo2_rel'] > 0]

                fig_lauf = go.Figure()

                if len(speeds_v) >= 2:
                    v_interp  = np.linspace(min(speeds_v), max(speeds_v), 300)
                    re_interp = np.interp(v_interp, speeds_v, re_vals)
                    fig_lauf.add_trace(go.Scatter(
                        x=v_interp, y=re_interp,
                        mode='lines',
                        line=dict(color='#00a1e0', width=2.5),
                        name='RE [ml/kg/km]',
                        yaxis='y1'
                    ))

                fig_lauf.add_trace(go.Scatter(
                    x=speeds_v, y=re_vals,
                    mode='markers',
                    marker=dict(color='#00a1e0', size=10, symbol='circle',
                                line=dict(color='white', width=1.5)),
                    name='RE Messp.',
                    yaxis='y1'
                ))

                if len(speeds_vo2) >= 1:
                    v_vo2_interp = np.linspace(min(speeds_vo2), max(speeds_vo2), 300)
                    vo2_interp   = np.interp(v_vo2_interp, speeds_vo2, vo2_r_vals)
                    fig_lauf.add_trace(go.Scatter(
                        x=v_vo2_interp, y=vo2_interp,
                        mode='lines',
                        line=dict(color='#ff7f0e', width=2, dash='dot'),
                        name='VO2ss [ml/min/kg]',
                        yaxis='y2'
                    ))
                    fig_lauf.add_trace(go.Scatter(
                        x=speeds_vo2, y=vo2_r_vals,
                        mode='markers',
                        marker=dict(color='#ff7f0e', size=8),
                        name='VO2ss Messp.',
                        yaxis='y2'
                    ))

                fig_lauf.update_layout(
                    title='Laufökonomie (ml/kg/km) & VO2ss',
                    xaxis_title='Geschwindigkeit [km/h]',
                    yaxis=dict(
                        title=dict(text='RE [ml/kg/km]', font=dict(color='#00a1e0')),
                        tickfont=dict(color='#00a1e0')
                    ),
                    yaxis2=dict(
                        title=dict(text='VO2ss [ml/min/kg]', font=dict(color='#ff7f0e')),
                        tickfont=dict(color='#ff7f0e'),
                        overlaying='y',
                        side='right'
                    ),
                    legend=dict(x=0.01, y=0.99, bgcolor='rgba(255,255,255,0.7)'),
                    height=380,
                    margin=dict(l=60, r=70, t=50, b=50),
                    plot_bgcolor='#f9f9f9'
                )
                st.plotly_chart(fig_lauf, width='stretch')

    # EXPERTEN TOOLS
    st.markdown("---")
    with st.expander("Experten Tools zum Vergleich mit einer bestehenden HYCYS Diagnostik"):
        if not spiro_df.empty:
            st.subheader("Spiro-Synchronisation & Rohdaten")
            import plotly.graph_objects as go
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(
                x=spiro_df['Time'],
                y=spiro_df['VO2'],
                mode='lines+markers',
                name='VO2',
                line=dict(color='gray', width=1.5),
                marker=dict(size=4),
                hovertemplate='Zeit: %{x}s<br>VO2: %{y:.1f} ml/min<extra></extra>'
            ))
            
            shapes = []
            for i, (s_t, e_t) in enumerate(window_coords):
                shapes.append(dict(
                    type="rect",
                    xref="x",
                    yref="paper",
                    x0=s_t,
                    y0=0,
                    x1=e_t,
                    y1=1,
                    fillcolor="#00a1e0",
                    opacity=0.3,
                    layer="below",
                    line_width=0,
                ))
            
            fig2.update_layout(
                xaxis=dict(title=dict(text="Zeit (s)")),
                yaxis=dict(title=dict(text="VO2 (ml/min)")),
                shapes=shapes,
                height=350,
                margin=dict(l=60, r=60, t=20, b=50),
                plot_bgcolor='#f9f9f9',
                showlegend=False
            )
            st.plotly_chart(fig2, use_container_width=True)
            
            col_diag1, col_diag2 = st.columns(2)
            with col_diag1:
                st.write("**Metabolische Vergleichswerte:**")
                st.write(f"- Berechnete VO2max: {rel_vo2max:.2f} ml/min/kg")
            with col_diag2:
                st.write("**Stufen-VO2 (relativ):**")
                st.dataframe(pd.DataFrame({"Stufe": [f"S{i+1}" for i in range(anzahl)], "VO2 rel": [f"{v:.2f}" for v in vo2_steady_values]}), hide_index=True)
        
        st.subheader("Stufenauswertung")
        if not spiro_df.empty:
            step_data = []
            for i in range(anzahl):
                v_ms   = speed[i]
                v_kmh  = round(v_ms * 3.6, 1)
                pace_s = (1000.0 / v_ms) if v_ms > 0 else 0
                pace_str = f"{int(pace_s // 60)}:{int(pace_s % 60):02d}"
                vo2_a  = vo2_steady_abs[i]  if i < len(vo2_steady_abs)  else 0.0
                vco2_a = vco2_steady_abs[i] if i < len(vco2_steady_abs) else 0.0
                vo2_r  = vo2_steady_values[i]
                rq_val   = (vco2_a / vo2_a) if (vco2_a > 0 and vo2_a > 0) else None
                rq_str   = f"{rq_val:.3f}" if rq_val else '-'
                re_kg_km = (vo2_r / (v_ms * 3.6) * 60.0) if (vo2_r > 0 and v_ms > 0) else None
                step_data.append({
                    "Stufe":              f"{i+1}",
                    "v [km/h]":           v_kmh,
                    "Pace [min/km]":      pace_str,
                    "VO2ss [ml/min]": round(vo2_a, 1)  if vo2_a > 0 else '-',
                    "VO2ss [ml/min/kg]": round(vo2_r, 2) if vo2_r > 0 else '-',
                    "RQ":                 rq_str,
                    "RE [ml/kg/km]":     round(re_kg_km, 1) if re_kg_km else '-'
                })
            st.dataframe(pd.DataFrame(step_data), hide_index=True, width='stretch')
            st.caption("RE [ml/kg/km] = VO2ss [ml/min/kg] ÷ v [km/h] × 60  |  Literaturstandard Laufökonomie")

    # --- PDF DOWNLOAD BUTTON ---
    st.markdown("---")
    if fig_laktat is not None:
        pdf_bytes = create_pdf(
            athlete_name, birthdate, test_date, st.session_state.get("diagnostik_type", "Running BLUE"), weight, body_fat_pct, rel_vo2max, df_res, fig_laktat,
            coach, sportart, kategorie, height, speed, lactate, hr, vo2_steady_values, vo2_steady_abs, vco2_steady_abs, uploaded_file
        )
        
        st.download_button(
            label="📄 PDF Report herunterladen",
            data=pdf_bytes,
            file_name=f"HYCYS_Laufdiagnostik_{athlete_name.replace(' ', '_')}.pdf",
            mime="application/pdf",
            type="primary"
        )

