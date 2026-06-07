"""
convert_sprint_to_excel.py  (v2 - SRM + FIT Support)

Konvertiert eine Sprint-Messdatei (.srm oder .fit) in das Excel-Format
wie im "Sprint Kohmann" (Data_Sprint Sheet), das von app.py eingelesen werden kann.

Das SRM7 Format speichert die Leistungswerte bereits in Watt.
Für die Konvertierung werden Slope und Zero-Offset benötigt, die auf dem
SRM-Gerät kalibriert wurden.

WICHTIG: Der SRM-Wert ist der bereits berechnete Leistungswert (Watt).
Die Formel lautet: P [W] = (raw - zero_offset) * slope
wobei raw die Rohfrequenz des SRM-Kraftsensors in Hz ist.

Standard-Werte (falls unbekannt):
  Slope: 16.964 Hz/Nm  (typisch für SRM PC7)
  Zero-Offset: 488 Hz  (muss vom Gerät abgelesen werden!)

Ausgabe-Format (entspricht Sprint Kohman.xlsx):
  Spalte A: Leistung [W]
  Spalte B: Herzfrequenz [bpm]  
  Spalte C: Trittfrequenz [rpm]
  Spalte D: Geschwindigkeit [km/h]
  Spalte E: Zeit als String "0:00:00.10"
  Spalte F: Zeilennummer (Index)

Zeilen 1-3: Header-Informationen
Datendaten beginnen ab Zeile 4.

Verwendung:
  python convert_sprint_to_excel.py <input.srm> [--slope 16.964] [--offset 488] [output.xlsx]
  python convert_sprint_to_excel.py <input.fit> [output.xlsx]
"""

import sys
import os
import struct
import datetime
import io
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# SRM7 Parser
# ---------------------------------------------------------------------------

def parse_srm7_header(data):
    """Liest den SRM7-Header und gibt ein Dict mit Metadaten zurück."""
    if data[:4] != b'SRM7':
        raise ValueError("Keine gültige SRM7-Datei (falsche Magic Bytes)")
    
    days_raw   = struct.unpack_from('<H', data, 4)[0]
    wheel_circ = struct.unpack_from('<H', data, 6)[0]   # mm
    rec_num    = data[8]    # Aufzeichnungsintervall Zähler
    rec_den    = data[9]    # Aufzeichnungsintervall Nenner
    rec_int    = rec_num / rec_den  # Sekunden pro Sample (z.B. 0.1)
    n_chunks   = struct.unpack_from('<H', data, 10)[0]
    n_blocks   = struct.unpack_from('<H', data, 12)[0]
    n_markers  = data[14]
    
    # Datum berechnen (Tage seit 01.01.1880)
    base_date = datetime.date(1880, 1, 1)
    record_date = base_date + datetime.timedelta(days=days_raw)
    
    # Athletenname aus Marker 0 (offset 86, erste 64 Bytes)
    athlete_name = data[86:150].rstrip(b'\x00').decode('latin1', errors='replace').strip()
    # Notiz aus Marker 0 ab Offset 16
    notes = data[16:86].rstrip(b'\x00').decode('latin1', errors='replace').strip()
    
    return {
        'rec_int': rec_int,
        'wheel_circ_mm': wheel_circ,
        'n_markers': n_markers,
        'n_blocks': n_blocks,
        'n_chunks': n_chunks,
        'record_date': record_date,
        'athlete_name': athlete_name,
        'notes': notes,
        'marker_offset': 86,
        'marker_size': 270,
        'block_size': 6,
    }


def parse_srm7_data(data, slope=1.0, zero_offset=0.0, trim_zeros=True):
    """
    Liest alle Datenpunkte aus einer SRM7-Datei.
    
    Die SRM7-Datei speichert Rohwerte in 14-Byte-Records:
    - Bytes 0: 0x00 (Marker)
    - Bytes 1-2: Rohleistung (uint16 LE) 
    - Bytes 3-4: Rohkadenz (uint16 LE)
    - Bytes 5-6: weitere Daten
    - Bytes 7-13: Trennzeichen-Muster (0x00 0x82 ...)
    
    Formel: P [W] = (raw - zero_offset) * slope
    
    HINWEIS: Das SRM PowerControl speichert bereits berechnete Wattwerte.
    Mit den Standard-Werten slope=1.0, offset=0 werden die raw-Werte direkt
    als Watt interpretiert. Slope und Offset müssen vom SRM-Gerät abgelesen werden.
    
    trim_zeros=True: Führende Null-Leistungswerte (Vorlauf ohne Treten)
    werden automatisch entfernt. Die Zeit wird neu ab 0 gezählt.
    
    Returns: Liste von Dicts mit elapsed_s, power_w, cadence, hr
    """
    hdr = parse_srm7_header(data)
    rec_int = hdr['rec_int']
    n_markers = hdr['n_markers']
    n_blocks = hdr['n_blocks']
    
    # Datensection berechnen
    data_offset = 86 + n_markers * 270 + n_blocks * 6
    data_section = data[data_offset:]
    
    record_size = 14  # empirisch ermittelt für SRM7
    n_records = len(data_section) // record_size
    
    records = []
    for i in range(n_records):
        offset = i * record_size
        b = data_section[offset:offset + record_size]
        if len(b) < record_size:
            break
        
        # Rohwerte auslesen
        raw_power   = struct.unpack_from('<H', b, 1)[0]  # bytes 1-2
        raw_cadence = struct.unpack_from('<H', b, 3)[0]  # bytes 3-4
        
        # Leistung berechnen
        if raw_power > 0:
            power_w = (raw_power - zero_offset) * slope
            power_w = max(0.0, power_w)
        else:
            power_w = 0.0
        
        elapsed = (i + 1) * rec_int
        
        records.append({
            'elapsed_s':   elapsed,
            'power_w':     round(power_w),
            'raw_power':   raw_power,
            'raw_cadence': raw_cadence,
            'heart_rate':  0,
            'cadence':     0,
        })
    
    # Führende Null-Leistungswerte abschneiden (Vorlauf)
    if trim_zeros and records:
        first_power_idx = next(
            (i for i, r in enumerate(records) if r['power_w'] > 0), None
        )
        if first_power_idx is not None and first_power_idx > 0:
            records = records[first_power_idx:]
        
        # Zeit neu ab rec_int starten
        if records:
            for j, r in enumerate(records):
                r['elapsed_s'] = round((j + 1) * rec_int, 3)
    
    return records, hdr


def parse_fit_file(filepath, trim_zeros=True):
    """
    Liest eine .fit Datei (1s-Intervalle vom Garmin/SRM-Export).
    
    trim_zeros=True: Führende Null-Leistungswerte (Vorlauf) werden entfernt.
    Gibt alle Records mit Power, HR, Cadence zurück.
    """
    try:
        import fitparse
    except ImportError:
        raise ImportError("fitparse nicht installiert. Bitte: pip install fitparse")

    fitfile = fitparse.FitFile(filepath)
    records = []
    t0 = None

    for msg in fitfile.get_messages("record"):
        vals = msg.get_values()
        ts = vals.get("timestamp")
        if ts is None:
            continue
        if t0 is None:
            t0 = ts

        elapsed = (ts - t0).total_seconds()
        power   = int(vals.get("power") or 0)
        cadence = int(vals.get("cadence") or 0)
        hr      = int(vals.get("heart_rate") or 0)
        speed   = float(vals.get("speed") or 0.0)
        speed_kmh = round(speed * 3.6, 2)

        records.append({
            "elapsed_s":  elapsed,
            "power_w":    power,
            "cadence":    cadence,
            "heart_rate": hr,
            "speed_kmh":  speed_kmh,
        })

    # Führende Null-Leistungswerte abschneiden
    if trim_zeros and records:
        first_power_idx = next(
            (i for i, r in enumerate(records) if r['power_w'] > 0), None
        )
        if first_power_idx is not None and first_power_idx > 0:
            records = records[first_power_idx:]
            # Zeit neu ab 1s starten
            for j, r in enumerate(records):
                r['elapsed_s'] = float(j + 1)

    athlete_name, test_date = _get_fit_info(filepath)
    return records, athlete_name, test_date


def _get_fit_info(filepath):
    """Extrahiert Athletenname und Datum aus dem Dateinamen."""
    basename = os.path.basename(filepath).replace(".fit", "").replace(".FIT", "")
    if " - " in basename:
        athlete_name = basename.split(" - ")[0].strip()
        date_part = basename.split(" - ")[1].strip()
        try:
            parts = date_part.split("-")
            test_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
        except Exception:
            test_date = date_part
    else:
        athlete_name = basename
        test_date = datetime.datetime.now().strftime("%d.%m.%Y")
    return athlete_name, test_date


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------

def seconds_to_timestr(s):
    """Konvertiert Sekunden in '0:00:00.10' Format."""
    m = int(s) // 60
    whole_sec = int(s) % 60
    hundredths = round((s % 1) * 100)
    if hundredths >= 100:
        hundredths = 99
    return f"0:{m:02d}:{whole_sec:02d}.{hundredths:02d}"


def create_sprint_excel(records, output_path, athlete_name="", test_date="",
                        source_filename="", rec_int_label="0.1s (SRM)"):
    """
    Erstellt Excel-Datei im Sprint-Kohmann-Format.
    
    Spalten: A=Power[W], B=HR[bpm], C=Cadence[rpm], D=Speed[km/h], E=Zeit, F=Index
    Zeilen 1-3: Header-Info
    Ab Zeile 4: Daten
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Sheet-Name wie Original (Initialen + Datum)
    if athlete_name and test_date:
        parts = athlete_name.split()
        initials = "".join(p[0] for p in parts if p)
        try:
            d_parts = test_date.replace(".", "-").split("-")
            date_short = d_parts[0].zfill(2) + d_parts[1].zfill(2) + d_parts[2][-2:]
            ws.title = initials + date_short
        except Exception:
            ws.title = "Data_Sprint"
    else:
        ws.title = "Data_Sprint"

    # Header-Zeilen (Rows 1-3) -- wie in Sprint Kohman.xlsx
    ws.cell(row=1, column=2).value = len(records)
    ws.cell(row=1, column=4).value = rec_int_label
    ws.cell(row=1, column=5).value = athlete_name
    ws.cell(row=1, column=6).value = test_date

    ws.cell(row=2, column=2).value = source_filename
    ws.cell(row=2, column=5).value = athlete_name

    ws.cell(row=3, column=2).value = len(records)
    ws.cell(row=3, column=4).value = rec_int_label
    ws.cell(row=3, column=5).value = "ERGOTEST:"

    # Datenzeilen ab Row 4
    for i, rec in enumerate(records):
        row = i + 4
        elapsed  = rec["elapsed_s"]
        time_str = seconds_to_timestr(elapsed)
        power    = rec.get("power_w", rec.get("power", 0))
        hr       = rec.get("heart_rate", 0)
        cadence  = rec.get("cadence", 0)
        speed    = rec.get("speed_kmh", 0.0)

        ws.cell(row=row, column=1).value = int(round(power))   # A: Leistung
        ws.cell(row=row, column=2).value = int(hr)             # B: HF
        ws.cell(row=row, column=3).value = int(cadence)        # C: Kadenz
        ws.cell(row=row, column=4).value = round(speed, 2)     # D: Speed
        ws.cell(row=row, column=5).value = time_str            # E: Zeit
        ws.cell(row=row, column=6).value = i + 2               # F: Index

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8

    wb.save(output_path)
    return ws.title


def create_sprint_excel_bytes(records, athlete_name="", test_date="",
                              source_filename="", rec_int_label="0.1s (SRM)"):
    """Wie create_sprint_excel, aber gibt Bytes zurück (für Download in Streamlit)."""
    buf = io.BytesIO()
    
    wb = openpyxl.Workbook()
    ws = wb.active

    if athlete_name and test_date:
        parts = athlete_name.split()
        initials = "".join(p[0] for p in parts if p)
        try:
            d_parts = test_date.replace(".", "-").split("-")
            date_short = d_parts[0].zfill(2) + d_parts[1].zfill(2) + d_parts[2][-2:]
            ws.title = initials + date_short
        except Exception:
            ws.title = "Data_Sprint"
    else:
        ws.title = "Data_Sprint"

    ws.cell(row=1, column=2).value = len(records)
    ws.cell(row=1, column=4).value = rec_int_label
    ws.cell(row=1, column=5).value = athlete_name
    ws.cell(row=1, column=6).value = test_date
    ws.cell(row=2, column=2).value = source_filename
    ws.cell(row=2, column=5).value = athlete_name
    ws.cell(row=3, column=2).value = len(records)
    ws.cell(row=3, column=4).value = rec_int_label
    ws.cell(row=3, column=5).value = "ERGOTEST:"

    for i, rec in enumerate(records):
        row = i + 4
        elapsed  = rec["elapsed_s"]
        time_str = seconds_to_timestr(elapsed)
        power    = rec.get("power_w", rec.get("power", 0))
        hr       = rec.get("heart_rate", 0)
        cadence  = rec.get("cadence", 0)
        speed    = rec.get("speed_kmh", 0.0)

        ws.cell(row=row, column=1).value = int(round(power))
        ws.cell(row=row, column=2).value = int(hr)
        ws.cell(row=row, column=3).value = int(cadence)
        ws.cell(row=row, column=4).value = round(speed, 2)
        ws.cell(row=row, column=5).value = time_str
        ws.cell(row=row, column=6).value = i + 2

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CLI Entry Point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Konvertiert .srm oder .fit Sprint-Dateien ins Excel Data_Sprint Format"
    )
    parser.add_argument("input", help="Eingabedatei (.srm oder .fit)")
    parser.add_argument("output", nargs="?", help="Ausgabe-Excel (.xlsx). Standard: <input>_Data_Sprint.xlsx")
    parser.add_argument("--slope", type=float, default=1.0,
                        help="SRM Kalibrierungs-Slope [W/count] (Standard: 1.0 = direkte Wattwerte)")
    parser.add_argument("--offset", type=float, default=0.0,
                        help="SRM Zero-Offset [counts] (Standard: 0)")
    args = parser.parse_args()

    input_path = args.input
    output_path = args.output or (os.path.splitext(input_path)[0] + "_Data_Sprint.xlsx")
    ext = os.path.splitext(input_path)[1].lower()

    print(f"Eingabedatei: {input_path}")
    print(f"Format:       {ext.upper()}")
    print()

    if ext == ".srm":
        with open(input_path, "rb") as f:
            data = f.read()
        
        records, hdr = parse_srm7_data(data, slope=args.slope, zero_offset=args.offset)
        athlete_name = hdr["athlete_name"]
        test_date = hdr["record_date"].strftime("%d.%m.%Y")
        rec_int_label = f"{hdr['rec_int']}s (SRM)"
        print(f"SRM-Version:  SRM7")
        print(f"Aufzeichnung: {hdr['rec_int']}s Intervall")
        print(f"Slope:        {args.slope}")
        print(f"Zero-Offset:  {args.offset}")

    elif ext == ".fit":
        records, athlete_name, test_date = parse_fit_file(input_path)
        rec_int_label = "1s (FIT)"
    else:
        print(f"FEHLER: Unbekanntes Format '{ext}'. Nur .srm und .fit werden unterstützt.")
        sys.exit(1)

    if not records:
        print("FEHLER: Keine Datenpunkte gefunden!")
        sys.exit(1)

    powers = [r.get("power_w", r.get("power", 0)) for r in records]
    print(f"Athlet:       {athlete_name}")
    print(f"Testdatum:    {test_date}")
    print(f"Messpunkte:   {len(records)}")
    print(f"Dauer:        {records[-1]['elapsed_s']:.2f}s")
    print(f"Max Power:    {max(powers):.0f}W")
    print()

    sheet_title = create_sprint_excel(
        records=records,
        output_path=output_path,
        athlete_name=athlete_name,
        test_date=test_date,
        source_filename=os.path.basename(input_path),
        rec_int_label=rec_int_label,
    )

    print(f"Excel gespeichert: {output_path}")
    print(f"  Sheet: '{sheet_title}'  ({len(records)} Datenzeilen ab Row 4)")
    print()
    print("Fertig! Datei kann in der App unter 'Sprint-Rohdaten Excel hochladen' verwendet werden.")


if __name__ == "__main__":
    main()
